import asyncio
from datetime import date, datetime, time, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

from sardoba_bot.db.queries import ReportQueries
from sardoba_bot.db.connection import DatabaseConnection

TASHKENT_TZ = ZoneInfo("Asia/Tashkent")


class ExportUtils:
    def __init__(self, db=None):
        self.db = db or DatabaseConnection()
        self._owns_db = db is None

    async def connect(self):
        if self._owns_db:
            await self.db.connect()

    def _datetime_bounds(self, start_date, end_date):
        start_day = datetime.fromisoformat(str(start_date)).date()
        end_day = datetime.fromisoformat(str(end_date)).date()
        start_bound = datetime.combine(start_day, datetime.min.time(), tzinfo=TASHKENT_TZ)
        end_bound = datetime.combine(end_day + timedelta(days=1), datetime.min.time(), tzinfo=TASHKENT_TZ)
        return start_bound, end_bound

    async def _fetch_excel_dataset(self, report_type="daily", start_date=None, end_date=None):
        if report_type == "daily":
            query = ReportQueries.DAILY_EXCEL_BASE
            params = None
            if start_date and end_date:
                query += " AND s.opened_at >= %s AND s.opened_at < %s"
                params = self._datetime_bounds(start_date, end_date)

            columns = [
                ("Ism", "first_name"),
                ("Familiya", "last_name"),
                ("Filial", "location"),
                ("Smena ochilgan vaqt", "opened_at"),
                ("Smena yopilgan vaqt", "closed_at"),
                ("Savdo", "sales_amount"),
                ("Kelgan qarz", "debt_received"),
                ("Chiqim", "expenses"),
                ("Uzcard", "uzcard_amount"),
                ("Humo", "humo_amount"),
                ("Uzcard vozvrat", "uzcard_refund"),
                ("Humo vozvrat", "humo_refund"),
                ("Boshqa to'lovlar", "other_payments"),
                ("Qarzga berilgan to'lovlar", "debt_payments"),
                ("Vozvrat qarzlar", "debt_refunds"),
                ("Sof summa", "total_balance"),
            ]
        elif report_type == "cashier_performance":
            query = ReportQueries.CASHIER_PERFORMANCE_EXCEL
            params = None
            columns = [
                ("Ism", "first_name"),
                ("Familiya", "last_name"),
                ("Telefon", "phone_number"),
                ("Smenalar soni", "shifts_count"),
                ("Jami savdo", "total_sales"),
                ("O'rtacha savdo", "avg_sales"),
                ("Jami chiqim", "total_expenses"),
            ]
        else:
            raise ValueError(f"Unsupported report type: {report_type}")

        rows = await self.db.fetch_all(query, params)
        headers = [header for header, _ in columns]
        values = [
            [self._normalize_excel_value(row.get(key)) for _, key in columns]
            for row in rows
        ]
        return headers, values

    def _normalize_excel_value(self, value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            try:
                return value.astimezone(TASHKENT_TZ).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                return str(value)
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, time):
            return value.strftime("%H:%M:%S")
        return value

    def _fit_excel_columns(self, ws):
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, min(ws.max_row, 51) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 40)

    def _build_excel_report(self, headers, rows):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Hisobotlar"
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in rows:
            ws.append(row)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        self._fit_excel_columns(ws)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    async def generate_excel_report(self, report_type="daily", start_date=None, end_date=None):
        """Generate Excel report based on report type."""
        headers, rows = await self._fetch_excel_dataset(report_type, start_date, end_date)
        return await asyncio.to_thread(self._build_excel_report, headers, rows)

    async def _fetch_pdf_dataset(self, report_type="daily", start_date=None, end_date=None):
        if report_type == "daily":
            query = ReportQueries.DAILY_PDF_BASE
            if start_date and end_date:
                query += " AND s.opened_at >= %s AND s.opened_at < %s"
                return await self.db.fetch_all(query, self._datetime_bounds(start_date, end_date))
            return await self.db.fetch_all(query)

        if report_type == "cashier_performance":
            query = ReportQueries.CASHIER_PERFORMANCE_PDF
            return await self.db.fetch_all(query)

        raise ValueError(f"Unsupported report type: {report_type}")

    def _apply_pdf_style(self, table):
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

    def _build_pdf_report(self, report_type, start_date, end_date, data):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []

        report_titles = {
            "daily": "Kunlik hisobot",
            "cashier_performance": "Kassirlar bo'yicha hisobot",
        }
        title_text = report_titles.get(report_type, report_type.replace("_", " ").title())
        title = Paragraph(f"Sardoba Restoran - {title_text}", styles["Title"])
        elements.append(title)

        if start_date and end_date:
            elements.append(Paragraph(f"Date Range: {start_date} to {end_date}", styles["Normal"]))

        elements.append(Paragraph(" ", styles["Normal"]))

        if data:
            if report_type == "daily":
                headers = [
                    "Kassir",
                    "Filial",
                    "Sana",
                    "Ochilish vaqti",
                    "Savdo",
                    "Kelgan qarz",
                    "Chiqim",
                    "Uzcard",
                    "Humo",
                    "Uzcard vozvrat",
                    "Humo vozvrat",
                    "Boshqa to'lov",
                    "Qarzga berilgan",
                    "Vozvrat qarz",
                    "Sof summa",
                ]
                part1_idx = list(range(0, 8))
                part2_idx = list(range(8, 15))

                table_data_1 = [[headers[i] for i in part1_idx]]
                table_data_2 = [[headers[i] for i in part2_idx]]

                for row in data:
                    full = [
                        row["cashier_name"],
                        row["location"],
                        str(row["date"]),
                        str(row["open_time"])[:5],
                        f"{row['sales_amount']:,.0f}",
                        f"{row['debt_received']:,.0f}",
                        f"{row['expenses']:,.0f}",
                        f"{row['uzcard_amount']:,.0f}",
                        f"{row['humo_amount']:,.0f}",
                        f"{row['uzcard_refund']:,.0f}",
                        f"{row['humo_refund']:,.0f}",
                        f"{row['other_payments']:,.0f}",
                        f"{row['debt_payments']:,.0f}",
                        f"{row['debt_refunds']:,.0f}",
                        f"{row['total_balance']:,.0f}",
                    ]
                    table_data_1.append([full[i] for i in part1_idx])
                    table_data_2.append([full[i] for i in part2_idx])

                table1 = Table(table_data_1, repeatRows=1)
                self._apply_pdf_style(table1)
                elements.append(table1)
                elements.append(Paragraph(" ", styles["Normal"]))

                table2 = Table(table_data_2, repeatRows=1)
                self._apply_pdf_style(table2)
                elements.append(table2)
            else:
                headers = ["Kassir", "Telefon", "Smenalar", "Jami savdo", "O'rtacha savdo", "Jami chiqim"]
                table_data = [headers]
                for row in data:
                    table_data.append(
                        [
                            row["cashier_name"],
                            row["phone_number"],
                            row["shifts_count"],
                            f"{row['total_sales']:,.0f}",
                            f"{row['avg_sales']:,.0f}",
                            f"{row['total_expenses']:,.0f}",
                        ]
                    )
                table = Table(table_data, repeatRows=1)
                self._apply_pdf_style(table)
                elements.append(table)
        else:
            elements.append(Paragraph("Tanlangan davr uchun ma'lumot topilmadi.", styles["Normal"]))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    async def generate_pdf_report(self, report_type="daily", start_date=None, end_date=None):
        """Generate PDF report based on report type."""
        data = await self._fetch_pdf_dataset(report_type, start_date, end_date)
        return await asyncio.to_thread(self._build_pdf_report, report_type, start_date, end_date, data)

    async def close_connection(self):
        """Close database connection."""
        if self._owns_db:
            await self.db.disconnect()
