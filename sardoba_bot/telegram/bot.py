# -*- coding: utf-8 -*-
import asyncio
import json
import logging
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional
import warnings
from io import BytesIO
from collections import defaultdict
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    KeyboardButton,
    InputFile,
    InputMediaPhoto,
    InputMediaDocument,
)
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler, CallbackQueryHandler
from telegram.request import HTTPXRequest
from sardoba_bot.core.constants import (
    ADMIN_DIRECT_ACTIONS,
    ADMIN_MENU_TEXTS,
    ADMIN_MENU_ROWS,
    ADMIN_REPORT_TEXTS,
    ADMIN_REPORT_PERIODS,
    ADMIN_REPORTS_MENU_ROWS,
    CASHIER_DIRECT_ACTIONS,
    CASHIER_MENU_ROWS,
    CASHIER_MENU_TEXTS,
    EDIT_REPORT_SELECT,
    EDIT_REPORT_VALUE,
    EXPORT_MENU_ROWS,
    EXPORT_MENU_TEXTS,
    KNOWN_MENU_TEXTS,
    MAIN_MENU,
    OPEN_SHIFT_AMOUNT,
    REGISTER_FIRSTNAME,
    REGISTER_LASTNAME,
    REGISTER_PASSWORD,
    REGISTER_PHONE,
    REPORT_DEBT_PAYMENTS,
    REPORT_DEBT_RECEIVED,
    REPORT_DEBT_REFUNDS,
    REPORT_EXPENSES,
    REPORT_HUMO,
    REPORT_P2P,
    REPORT_HUMO_REFUND,
    REPORT_OTHER_PAYMENTS,
    REPORT_SALES,
    REPORT_TAX_INFO,
    REPORT_UZCARD,
    REPORT_UZCARD_REFUND,
    SELECT_LOCATION,
    SELECT_PAYMENT_IMAGE,
    SELECT_ROLE,
    SUBMIT_DAILY_REPORT,
    UPLOAD_OPENING_NOTIFICATION,
    UPLOAD_PAYMENT_IMAGE,
    UPLOAD_RECEIPT_ROLL,
    UPLOAD_TERMINAL_POWER,
    UPLOAD_WORKPLACE_STATUS,
    UPLOAD_ZERO_REPORT,
    VERIFY_PASSWORD,
    ADMIN_LOGIN,
    ADMIN_REGISTER_PASSWORD,
    ADMIN_REGISTER_PHONE,
    ADMIN_VERIFY_PASSWORD,
    CLOSE_SHIFT,
    CLOSE_SHIFT_NOTE,
)
from sardoba_bot.db.queries import AdminQueries, CashierQueries, CommonQueries
from sardoba_bot.db.connection import DatabaseConnection
from sardoba_bot.common.utils import hash_password, verify_password, validate_phone_number
from sardoba_bot.services.export_utils import ExportUtils
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

logger = logging.getLogger(__name__)

POSTGRES_SCHEMA_PATH = Path(__file__).resolve().parents[2] / "sql" / "postgres" / "schema.sql"
TASHKENT_TZ = ZoneInfo("Asia/Tashkent")
CASHIER_RESUME_ACTION_TTL_MINUTES = 15

# Silence PTB ConversationHandler per_message warnings (non-fatal)
warnings.filterwarnings(
    "ignore",
    message=".*CallbackQueryHandler.*per_message.*ConversationHandler.*",
    category=UserWarning,
)

class SardobaBot:
    def __init__(self):
        self.db = DatabaseConnection()
        self.export_utils = ExportUtils(self.db)
        self._group_chat_id_cache = None
        self._locations_cache = None

    def _build_contact_request_keyboard(self) -> ReplyKeyboardMarkup:
        """Create a keyboard for sharing the user's own contact."""
        keyboard = [[KeyboardButton("Telefon raqamni ulashish", request_contact=True)]]
        return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)

    def _build_cashier_menu_keyboard(self) -> ReplyKeyboardMarkup:
        """Create the persistent cashier menu keyboard."""
        return ReplyKeyboardMarkup(
            [[KeyboardButton(label) for label in row] for row in CASHIER_MENU_ROWS],
            resize_keyboard=True,
        )

    def _build_expense_payment_type_keyboard(self) -> ReplyKeyboardMarkup:
        """Create a keyboard for expense payment type selection."""
        return ReplyKeyboardMarkup(
            [[KeyboardButton("Naqd"), KeyboardButton("Karta"), KeyboardButton("P2P")]],
            resize_keyboard=True,
            one_time_keyboard=True,
        )

    def _build_expense_entry_keyboard(self) -> ReplyKeyboardMarkup:
        """Create a keyboard for multi-entry expense collection."""
        return ReplyKeyboardMarkup(
            [[KeyboardButton("➕ Yana qo'shish"), KeyboardButton("✅ Yakunlash")]],
            resize_keyboard=True,
        )

    def _prime_cashier_password_setup(self, context: ContextTypes.DEFAULT_TYPE, telegram_id: int) -> None:
        """Mark a cashier so their next message starts the password setup flow."""
        application = getattr(context, "application", None)
        app_user_data = getattr(application, "_user_data", None)
        if app_user_data is None:
            return

        user_state = app_user_data[telegram_id]
        user_state["cashier_set_password"] = True
        user_state["cashier_set_password_confirm"] = False
        user_state["cashier_pending_password"] = False
        user_state["cashier_authenticated"] = False
        user_state.pop("new_password_hash", None)

    def _set_cashier_resume_action(self, context: ContextTypes.DEFAULT_TYPE, action_name: Optional[str]) -> None:
        """Remember cashier intent so auth can resume it."""
        if not action_name:
            return
        context.user_data["cashier_resume_action"] = action_name
        context.user_data["cashier_resume_action_at"] = self._now_tashkent().isoformat()

    def _clear_cashier_resume_action(self, context: ContextTypes.DEFAULT_TYPE) -> None:
        context.user_data.pop("cashier_resume_action", None)
        context.user_data.pop("cashier_resume_action_at", None)

    def _consume_cashier_resume_action(self, context: ContextTypes.DEFAULT_TYPE) -> tuple[Optional[str], bool]:
        """Return pending action and whether it expired."""
        action_name = context.user_data.get("cashier_resume_action")
        action_at_raw = context.user_data.get("cashier_resume_action_at")
        self._clear_cashier_resume_action(context)

        if not action_name:
            return None, False

        if not action_at_raw:
            return action_name, False

        try:
            action_at = datetime.fromisoformat(str(action_at_raw))
            if action_at.tzinfo is None:
                action_at = action_at.replace(tzinfo=TASHKENT_TZ)
        except ValueError:
            return None, True

        is_expired = (self._now_tashkent() - action_at) > timedelta(minutes=CASHIER_RESUME_ACTION_TTL_MINUTES)
        if is_expired:
            return None, True
        return action_name, False

    async def _resume_cashier_post_auth_action(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
        """Continue pending cashier action right after successful auth."""
        action_name, expired = self._consume_cashier_resume_action(context)
        if expired:
            await update.message.reply_text("Sessiya muddati tugagan. Iltimos, menyudan qayta tanlang.")
            await self.show_cashier_menu(update, context)
            return True

        if not action_name:
            return False

        action = getattr(self, action_name, None)
        if not action:
            await self.show_cashier_menu(update, context)
            return True

        await action(update, context)
        return True

    async def _ask_for_phone_contact(self, message, prompt: str):
        """Ask the user to share a phone number via Telegram contact button."""
        await message.reply_text(prompt, reply_markup=self._build_contact_request_keyboard())

    def _extract_shared_phone(self, update: Update) -> tuple[Optional[str], Optional[str]]:
        """Extract shared phone number or return a validation error message."""
        contact = getattr(update.message, "contact", None)
        user_id = getattr(update.effective_user, "id", None)

        if not contact:
            return None, "Telefon raqamingizni oddiy matn emas, pastdagi tugma orqali ulashing:"

        contact_user_id = getattr(contact, "user_id", None)
        if contact_user_id and user_id and contact_user_id != user_id:
            return None, "Iltimos, aynan o'zingizning telefon raqamingizni ulashing:"

        phone = getattr(contact, "phone_number", "") or ""
        if not validate_phone_number(phone):
            return None, "Ulashilgan telefon raqami noto'g'ri. Qayta yuborish uchun tugmani bosing:"

        return phone, None

    async def initialize(self):
        if not await self.db.connect():
            raise RuntimeError("PostgreSQL connection could not be established.")
        await self._ensure_runtime_schema()

    async def shutdown(self):
        await self.export_utils.close_connection()
        await self.db.disconnect()

    async def _ensure_runtime_schema(self):
        """Apply the full PostgreSQL schema so a new database can bootstrap itself."""
        try:
            schema_sql = POSTGRES_SCHEMA_PATH.read_text(encoding="utf-8-sig").strip()
        except OSError as exc:
            raise RuntimeError(f"Failed to read PostgreSQL schema file: {POSTGRES_SCHEMA_PATH}") from exc

        if not schema_sql:
            raise RuntimeError(f"PostgreSQL schema file is empty: {POSTGRES_SCHEMA_PATH}")

        if not await self.db.execute_query(schema_sql):
            raise RuntimeError("Failed to apply PostgreSQL schema during bot startup.")

        logger.info("PostgreSQL schema is ready.")

    async def _get_group_chat_id(self):
        if self._group_chat_id_cache is not None:
            return self._group_chat_id_cache
        row = await self.db.fetch_one(CommonQueries.BOT_GROUP_CHAT_ID)
        if not row:
            return 0
        self._group_chat_id_cache = int(row.get("group_chat_id") or 0)
        return self._group_chat_id_cache

    async def _get_locations(self):
        if self._locations_cache is None:
            self._locations_cache = await self.db.fetch_all(CommonQueries.ACTIVE_LOCATION_LIST)
        return self._locations_cache or []

    async def _get_location_name(self, location_id: int) -> str:
        for location in await self._get_locations():
            if int(location.get("id")) == int(location_id):
                return location.get("name") or str(location_id)
        row = await self.db.fetch_one(CommonQueries.LOCATION_NAME_BY_ID, (location_id,))
        return row.get("name") if row else str(location_id)

    def _now_tashkent(self) -> datetime:
        return datetime.now(TASHKENT_TZ)

    def _day_bounds(self, start, end=None):
        start_day = datetime.fromisoformat(str(start)).date()
        end_day = datetime.fromisoformat(str(end or start)).date()
        start_bound = datetime.combine(start_day, datetime.min.time(), tzinfo=TASHKENT_TZ)
        end_bound = datetime.combine(end_day + timedelta(days=1), datetime.min.time(), tzinfo=TASHKENT_TZ)
        return start_bound, end_bound

    def _fmt_money(self, value) -> str:
        try:
            num = float(value or 0)
        except Exception:
            return str(value)
        return f"{num:,.0f}".replace(",", " ")

    def _fmt_datetime(self, value) -> str:
        if not value:
            return "-"
        try:
            return value.astimezone(TASHKENT_TZ).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(value)[:19]

    def _parse_report_data(self, value) -> dict:
        if isinstance(value, dict):
            return value
        if isinstance(value, str) and value.strip():
            try:
                parsed = json.loads(value)
            except Exception:
                return {}
            return parsed if isinstance(parsed, dict) else {}
        return {}

    def _normalize_payment_type(self, text: str) -> Optional[str]:
        normalized = (text or "").strip().lower()
        return {
            "naqd": "Naqd",
            "karta": "Karta",
            "p2p": "P2P",
        }.get(normalized)

    def _normalize_expense_payment_type(self, text: str) -> Optional[str]:
        return self._normalize_payment_type(text)

    def _normalize_expense_action(self, text: str) -> Optional[str]:
        normalized = (text or "").strip().lower()
        return {
            "➕ yana qo'shish": "add_more",
            "yana qo'shish": "add_more",
            "✅ yakunlash": "finish",
            "yakunlash": "finish",
        }.get(normalized)

    def _sverka_payment_method_labels(self) -> dict[str, str]:
        return {
            "debt_refunds": "Vozvrat qarzlar",
        }

    def _clear_generic_payment_method_state(self, context: ContextTypes.DEFAULT_TYPE) -> None:
        for key in (
            "other_payments_detail_stage",
            "other_payments_payment_type",
            "other_payments_comment",
            "debt_refunds_detail_stage",
            "debt_refunds_payment_type",
        ):
            context.user_data.pop(key, None)

    def _clear_debt_received_detail_state(self, context: ContextTypes.DEFAULT_TYPE) -> None:
        for key in (
            "debt_received_detail_stage",
            "debt_received_payment_type",
            "debt_received_counterparty_name",
            "debt_received_counterparty_phone",
            "debt_received_current_name",
            "debt_received_current_phone",
            "debt_received_current_amount",
            "debt_received_items",
        ):
            context.user_data.pop(key, None)

    def _debt_received_items(self, context: ContextTypes.DEFAULT_TYPE) -> list[dict]:
        items = context.user_data.get("debt_received_items")
        if not isinstance(items, list):
            items = []
            context.user_data["debt_received_items"] = items
        return items

    def _debt_received_items_total(self, context: ContextTypes.DEFAULT_TYPE) -> float:
        total = 0.0
        for item in self._debt_received_items(context):
            try:
                total += float(item.get("amount") or 0)
            except Exception:
                continue
        return total

    def _build_debt_received_loop_keyboard(self) -> ReplyKeyboardMarkup:
        return ReplyKeyboardMarkup(
            [[KeyboardButton("➕ Yana qo'shish"), KeyboardButton("✅ Yakunlash")]],
            resize_keyboard=True,
        )

    def _normalize_debt_received_action(self, text: str) -> Optional[str]:
        normalized = (text or "").strip().lower()
        return {
            "➕ yana qo'shish": "add_more",
            "yana qo'shish": "add_more",
            "✅ yakunlash": "finish",
            "yakunlash": "finish",
        }.get(normalized)

    def _clear_debt_payments_detail_state(self, context: ContextTypes.DEFAULT_TYPE) -> None:
        for key in (
            "debt_payments_detail_stage",
            "debt_payments_payment_type",
            "debt_payments_counterparty_name",
            "debt_payments_counterparty_phone",
            "debt_payments_current_name",
            "debt_payments_current_phone",
            "debt_payments_current_amount",
            "debt_payments_items",
        ):
            context.user_data.pop(key, None)

    def _debt_payments_items(self, context: ContextTypes.DEFAULT_TYPE) -> list[dict]:
        items = context.user_data.get("debt_payments_items")
        if not isinstance(items, list):
            items = []
            context.user_data["debt_payments_items"] = items
        return items

    def _debt_payments_items_total(self, context: ContextTypes.DEFAULT_TYPE) -> float:
        total = 0.0
        for item in self._debt_payments_items(context):
            try:
                total += float(item.get("amount") or 0)
            except Exception:
                continue
        return total

    def _build_debt_payments_loop_keyboard(self) -> ReplyKeyboardMarkup:
        return ReplyKeyboardMarkup(
            [[KeyboardButton("➕ Yana qo'shish"), KeyboardButton("✅ Yakunlash")]],
            resize_keyboard=True,
        )

    def _build_debt_payments_confirm_text(self, context: ContextTypes.DEFAULT_TYPE) -> str:
        name = context.user_data.get("debt_payments_current_name", "-")
        phone = context.user_data.get("debt_payments_current_phone", "-")
        amount = context.user_data.get("debt_payments_current_amount", 0)
        return (
            "📋 Qarz berish ma'lumotlarini tekshiring:\n\n"
            f"👤 Ism: {name}\n"
            f"📞 Telefon: {phone}\n"
            f"💰 Summa: {self._fmt_money(amount)}\n\n"
            "Tasdiqlash uchun \"Qarz berildi\" tugmasini bosing."
        )

    def _build_debt_payments_confirm_keyboard(self) -> ReplyKeyboardMarkup:
        return ReplyKeyboardMarkup(
            [[KeyboardButton("✅ Qarz berildi"), KeyboardButton("❌ Bekor qilish")]],
            resize_keyboard=True,
        )

    def _normalize_debt_payments_action(self, text: str) -> Optional[str]:
        normalized = (text or "").strip().lower()
        return {
            "➕ yana qo'shish": "add_more",
            "yana qo'shish": "add_more",
            "✅ yakunlash": "finish",
            "yakunlash": "finish",
        }.get(normalized)

    def _normalize_debt_payments_confirm(self, text: str) -> Optional[str]:
        normalized = (text or "").strip().lower()
        return {
            "✅ qarz berildi": "confirm",
            "qarz berildi": "confirm",
            "❌ bekor qilish": "cancel",
            "bekor qilish": "cancel",
        }.get(normalized)


    def _clear_expense_detail_state(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        *,
        keep_items: bool = False,
    ) -> None:
        keys = [
            "expense_detail_stage",
            "expense_payment_type",
            "expense_paid_to",
            "expense_recipient_name",
            "expense_recipient_phone",
            "expense_reason",
            "expense_cash_amount",
        ]
        if not keep_items:
            keys.append("expense_items")
        for key in keys:
            context.user_data.pop(key, None)

    def _expense_items(self, context: ContextTypes.DEFAULT_TYPE) -> list[dict]:
        items = context.user_data.get("expense_items")
        if not isinstance(items, list):
            items = []
            context.user_data["expense_items"] = items
        return items

    def _expense_items_total(self, context: ContextTypes.DEFAULT_TYPE) -> float:
        total = 0.0
        for item in context.user_data.get("expense_items") or []:
            if not isinstance(item, dict):
                continue
            try:
                total += float(item.get("amount") or 0)
            except Exception:
                continue
        return total

    def _parse_expense_item(self, text: str) -> dict:
        raw = (text or "").strip()
        if not raw:
            raise ValueError("Expense line is empty.")

        matches = list(re.finditer(r"\d[\d\s,._]*", raw))
        if not matches:
            raise ValueError("Expense line does not contain amount.")

        amount_text = re.sub(r"[^\d\s]", "", matches[-1].group(0))
        amount = self._parse_amount(amount_text)
        if amount <= 0:
            raise ValueError("Expense amount must be positive.")

        return {"text": raw, "amount": amount}

    def _build_expense_entry_prompt(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        note: Optional[str] = None,
    ) -> str:
        lines = []
        if note:
            lines.append(note)

        lines.extend(
            [
                "Chiqim sababini kiriting.",
                "Masalan:",
                "Mirshod Dastafka -- 10 000",
                "Ulug Paynet -- 100 000",
            ]
        )

        item_lines = []
        for item in context.user_data.get("expense_items") or []:
            if not isinstance(item, dict):
                continue
            text = (item.get("text") or "").strip()
            if text:
                item_lines.append(f"• {text}")

        if item_lines:
            lines.extend(
                [
                    "",
                    "Kiritilgan chiqimlar:",
                    *item_lines,
                    f"Jami chiqim: {self._fmt_money(self._expense_items_total(context))}",
                    "",
                    "Keyingi qatorni yuboring yoki Yakunlashni bosing.",
                ]
            )
        else:
            lines.extend(
                [
                    "",
                    "Agar chiqim bo'lmasa, Yakunlashni bosing.",
                ]
            )

        return "\n".join(lines)

    def _clear_sverka_value_state(self, context: ContextTypes.DEFAULT_TYPE) -> None:
        for key, *_ in self._sverka_config():
            context.user_data.pop(key, None)
        self._clear_debt_received_detail_state(context)
        self._clear_debt_payments_detail_state(context)
        self._clear_expense_detail_state(context)
        self._clear_generic_payment_method_state(context)
        self._clear_tax_info_state(context)

    def _clear_tax_info_state(self, context: ContextTypes.DEFAULT_TYPE) -> None:
        for key in (
            "tax_info_stage",
            "tax_info_check_image",
            "tax_info_cash_amount",
        ):
            context.user_data.pop(key, None)

    def _clear_sverka_flow_state(self, context: ContextTypes.DEFAULT_TYPE) -> None:
        context.user_data.pop("pending_sverka_key", None)
        context.user_data.pop("pending_sverka_state", None)
        context.user_data.pop("sverka_status", None)
        context.user_data.pop("sverka_entrypoint", None)

    async def _ensure_opening_requirements_completed(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        shift_id: int,
    ) -> bool:
        required_opening = [
            ("workplace_status", "Ish joyi holati rasmi", 2),
            ("terminal_power", "Terminal/ratsiya quvvati rasmi", 1),
            ("zero_report", "Uzcard/Humo nol hisobot rasmi", 1),
            ("opening_notification", "iiko/soliq smena ochilganlik rasmi", 1),
            ("receipt_roll", "Zaxira chek lenta rasmi", 1),
        ]
        missing = []
        for image_type, label, required_count in required_opening:
            current_count = await self._count_shift_images(shift_id, image_type)
            if current_count < required_count:
                remaining = required_count - current_count
                if required_count == 1:
                    missing.append(f"- {label}")
                else:
                    missing.append(f"- {label} ({remaining} ta qolgan)")
        if not missing:
            return True

        msg = (
            "Smena ochish bosqichidagi rasmlar to'liq emas.\n"
            "Quyidagilar yetishmayapti:\n"
            + "\n".join(missing)
        )
        await self.show_opening_requirements_menu(update, context, shift_id, note=msg)
        return False

    def _prepare_sverka_context(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        shift_id: int,
        entrypoint: str,
        force_reset: bool = False,
    ) -> None:
        if force_reset or context.user_data.get("sverka_shift_id") != shift_id:
            self._clear_sverka_value_state(context)

        context.user_data["current_shift_id"] = shift_id
        context.user_data["flow"] = "sverka"
        context.user_data["sverka_shift_id"] = shift_id
        context.user_data["sverka_entrypoint"] = entrypoint
        context.user_data.pop("pending_close_amount", None)
        context.user_data["sverka_status"] = {key: False for key, *_ in self._active_sverka_config(context)}
        context.user_data.pop("pending_sverka_key", None)
        context.user_data.pop("pending_sverka_state", None)
        self._init_sverka_status(context)

    async def _start_sverka_flow(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        shift_id: int,
        entrypoint: str,
        force_reset: bool = False,
        note: Optional[str] = None,
    ):
        self._prepare_sverka_context(context, int(shift_id), entrypoint, force_reset=force_reset)
        await self.show_sverka_menu(update, context, note=note)
        return SUBMIT_DAILY_REPORT

    async def _prompt_close_shift_amount(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        text: Optional[str] = None,
    ):
        context.user_data["flow"] = "closing"
        context.user_data.pop("pending_close_amount", None)
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=text or "Smenani yopish uchun yakuniy summani kiriting:",
            reply_markup=ReplyKeyboardRemove(),
        )
        return CLOSE_SHIFT

    def _build_report_data_payload(self, context: ContextTypes.DEFAULT_TYPE) -> dict:
        payload = {}

        if float(context.user_data.get("debt_received") or 0) > 0:
            items = context.user_data.get("debt_received_items")
            if isinstance(items, list) and items:
                payload["debt_received_detail"] = {"items": items}
            else:
                debt_received_detail = {
                    "counterparty_name": context.user_data.get("debt_received_counterparty_name"),
                    "counterparty_phone": context.user_data.get("debt_received_counterparty_phone"),
                    "payment_type": context.user_data.get("debt_received_payment_type"),
                }
                debt_received_detail = {
                    key: value for key, value in debt_received_detail.items() if value not in (None, "")
                }
                if debt_received_detail:
                    payload["debt_received_detail"] = debt_received_detail

        if float(context.user_data.get("debt_payments") or 0) > 0:
            items = context.user_data.get("debt_payments_items")
            if isinstance(items, list) and items:
                payload["debt_payments_detail"] = {"items": items}
            else:
                # Legacy single-entry fallback
                debt_payments_detail = {
                    "counterparty_name": context.user_data.get("debt_payments_counterparty_name"),
                    "counterparty_phone": context.user_data.get("debt_payments_counterparty_phone"),
                    "payment_type": context.user_data.get("debt_payments_payment_type"),
                }
                debt_payments_detail = {
                    key: value for key, value in debt_payments_detail.items() if value not in (None, "")
                }
                if debt_payments_detail:
                    payload["debt_payments_detail"] = debt_payments_detail

        if float(context.user_data.get("expenses") or 0) > 0:
            expense_items = []
            for item in context.user_data.get("expense_items") or []:
                if not isinstance(item, dict):
                    continue
                text = (item.get("text") or "").strip()
                if not text:
                    continue

                entry = {"text": text}
                try:
                    amount = float(item.get("amount") or 0)
                except Exception:
                    amount = 0
                if amount > 0:
                    entry["amount"] = amount
                expense_items.append(entry)

            expense_detail = {}
            if expense_items:
                expense_detail["items"] = expense_items
            if context.user_data.get("expense_cash_amount") is not None:
                expense_detail["cash_amount"] = context.user_data.get("expense_cash_amount")

            legacy_detail = {
                "payment_type": context.user_data.get("expense_payment_type"),
                "paid_to": context.user_data.get("expense_paid_to"),
                "recipient_name": context.user_data.get("expense_recipient_name"),
                "recipient_phone": context.user_data.get("expense_recipient_phone"),
                "reason": context.user_data.get("expense_reason"),
            }
            legacy_detail = {
                key: value for key, value in legacy_detail.items() if value not in (None, "")
            }
            expense_detail.update(legacy_detail)
            if expense_detail:
                payload["expense_detail"] = expense_detail

        other_payments_comment = (context.user_data.get("other_payments_comment") or "").strip()
        if other_payments_comment:
            payload["other_payments_comment"] = other_payments_comment

        tax_info = {
            "check_image": context.user_data.get("tax_info_check_image"),
            "cash_amount": context.user_data.get("tax_info_cash_amount"),
        }
        tax_info = {key: value for key, value in tax_info.items() if value not in (None, "")}
        if tax_info:
            payload["tax_info"] = tax_info

        payment_methods = {}
        for key in self._sverka_payment_method_labels():
            amount = float(context.user_data.get(key) or 0)
            payment_type = context.user_data.get(f"{key}_payment_type")
            if amount > 0 and payment_type:
                payment_methods[key] = payment_type
        if payment_methods:
            payload["payment_methods"] = payment_methods

        return payload

    def _build_debt_received_detail_lines(self, row) -> list[str]:
        report_data = self._parse_report_data((row or {}).get("report_data"))
        detail = report_data.get("debt_received_detail")
        if not isinstance(detail, dict) or not detail:
            return []

        items = detail.get("items")
        if isinstance(items, list) and items:
            lines = ["📌 Kelgan qarzlar tafsiloti"]
            for idx, item in enumerate(items, 1):
                if not isinstance(item, dict):
                    continue
                lines.append(f"  {idx}. 👤 {item.get('counterparty_name', '-')}")
                if item.get("counterparty_phone"):
                    lines.append(f"     📞 {item['counterparty_phone']}")
                if item.get("amount"):
                    lines.append(f"     💰 {self._fmt_money(item['amount'])}")
                if item.get("payment_type"):
                    lines.append(f"     💳 {item['payment_type']}")
            return lines

        lines = ["📌 Kelgan qarz tafsiloti"]
        if detail.get("counterparty_name"):
            lines.append(f"👤 Kimdan: {detail['counterparty_name']}")
        if detail.get("counterparty_phone"):
            lines.append(f"📞 Telefon: {detail['counterparty_phone']}")
        if detail.get("payment_type"):
            lines.append(f"💳 To'lov turi: {detail['payment_type']}")
        return lines

    def _build_debt_payments_detail_lines(self, row) -> list[str]:
        report_data = self._parse_report_data((row or {}).get("report_data"))
        detail = report_data.get("debt_payments_detail")
        if not isinstance(detail, dict) or not detail:
            return []

        items = detail.get("items")
        if isinstance(items, list) and items:
            lines = ["📌 Qarz berilganlar tafsiloti"]
            for idx, item in enumerate(items, 1):
                if not isinstance(item, dict):
                    continue
                lines.append(f"  {idx}. 👤 {item.get('counterparty_name', '-')}")
                if item.get("counterparty_phone"):
                    lines.append(f"     📞 {item['counterparty_phone']}")
                if item.get("amount"):
                    lines.append(f"     💰 {self._fmt_money(item['amount'])}")
                if item.get("check_image"):
                    lines.append("     🧾 Chek rasmi biriktilgan")
            return lines

        # Legacy single-entry format
        lines = ["📌 Qarz to'lovi tafsiloti"]
        if detail.get("counterparty_name"):
            lines.append(f"👤 Kimga: {detail['counterparty_name']}")
        if detail.get("counterparty_phone"):
            lines.append(f"📞 Telefon: {detail['counterparty_phone']}")
        if detail.get("payment_type"):
            lines.append(f"💳 To'lov turi: {detail['payment_type']}")
        return lines

    def _debt_payment_check_items(self, row) -> list[dict]:
        report_data = self._parse_report_data((row or {}).get("report_data"))
        detail = report_data.get("debt_payments_detail")
        if not isinstance(detail, dict):
            return []
        items = detail.get("items")
        if not isinstance(items, list):
            return []
        return [
            item for item in items
            if isinstance(item, dict) and (item.get("check_image") or "").strip()
        ]

    def _font(self, size: int, *, bold: bool = False):
        from PIL import ImageFont

        paths = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/dejavu/DejaVuSans.ttf",
            "/System/Library/Fonts/Supplemental/Arial Bold.ttf" if bold else "/System/Library/Fonts/Supplemental/Arial.ttf",
            "/Library/Fonts/Arial Bold.ttf" if bold else "/Library/Fonts/Arial.ttf",
        ]
        for path in paths:
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                continue
        return ImageFont.load_default()

    def _decorate_labeled_check_image(
        self,
        image_bytes,
        left_title: str,
        right_title: str,
        *,
        large_title: bool = False,
    ) -> BytesIO:
        from PIL import ImageDraw, ImageOps

        source = PILImage.open(BytesIO(image_bytes))
        source = ImageOps.exif_transpose(source).convert("RGB")

        max_image_width = 1200
        if source.width > max_image_width:
            ratio = max_image_width / float(source.width)
            source = source.resize((max_image_width, max(1, int(source.height * ratio))), PILImage.LANCZOS)

        border = 20
        padding = 28
        canvas_w = source.width + (border + padding) * 2
        available_w = canvas_w - (border + padding) * 2
        title = str(right_title or "").strip()
        font_size = 200

        measure_draw = ImageDraw.Draw(PILImage.new("RGB", (1, 1)))

        def text_size(text: str, font_obj) -> tuple[int, int]:
            bbox = measure_draw.textbbox((0, 0), text, font=font_obj)
            return bbox[2] - bbox[0], bbox[3] - bbox[1]

        def wrap_text(text: str, font_obj, max_width: int) -> list[str]:
            words = [word for word in str(text or "").split() if word]
            if not words:
                return [""]
            lines = []
            current = words[0]
            for word in words[1:]:
                candidate = f"{current} {word}"
                if text_size(candidate, font_obj)[0] <= max_width:
                    current = candidate
                else:
                    lines.append(current)
                    current = word
            lines.append(current)
            return lines

        if title:
            min_font_size = 72
            while font_size > min_font_size:
                font = self._font(font_size, bold=True)
                left_w, _ = text_size(left_title, font)
                right_w, _ = text_size(title, font)
                if left_w + right_w + padding <= available_w:
                    break
                font_size -= 2
            left_lines = [left_title]
        else:
            font = self._font(font_size, bold=True)
            left_lines = wrap_text(left_title, font, available_w)

        line_gap = max(8, font_size // 6)
        line_heights = [text_size(line, font)[1] for line in left_lines]
        left_text_h = sum(line_heights) + line_gap * max(0, len(left_lines) - 1)
        right_w, right_h = text_size(title, font) if title else (0, 0)
        text_h = max(left_text_h, right_h)
        header_h = text_h + padding * 2
        canvas_h = source.height + header_h + padding + border * 2

        canvas = PILImage.new("RGB", (canvas_w, canvas_h), "white")
        draw = ImageDraw.Draw(canvas)
        draw.rectangle(
            [border // 2, border // 2, canvas_w - border // 2 - 1, canvas_h - border // 2 - 1],
            outline=(20, 20, 20),
            width=border,
        )

        left_x = border + padding
        text_x = canvas_w - border - padding - right_w
        text_y = border + padding
        line_y = text_y
        for idx, line in enumerate(left_lines):
            draw.text((left_x, line_y), line, fill=(15, 23, 42), font=font)
            line_y += line_heights[idx] + line_gap
        if title:
            draw.text((text_x, text_y), title, fill=(15, 23, 42), font=font)

        image_x = border + padding
        image_y = border + header_h
        canvas.paste(source, (image_x, image_y))

        out = BytesIO()
        canvas.save(out, format="JPEG", quality=92)
        out.seek(0)
        return out

    def _decorate_debt_payment_check_image(self, image_bytes, debt_id: int, debt_name: str) -> BytesIO:
        return self._decorate_labeled_check_image(
            image_bytes,
            "Qarz chek",
            f"ID: {debt_id} | {str(debt_name or '-').strip() or '-'}",
        )

    def _decorate_tax_info_check_image(self, image_bytes, cash_amount) -> BytesIO:
        return self._decorate_labeled_check_image(
            image_bytes,
            "Soliq chek",
            "",
        )

    async def _send_debt_payment_check_images(self, context: ContextTypes.DEFAULT_TYPE, row) -> None:
        items = self._debt_payment_check_items(row)
        if not items:
            return

        group_chat_id = await self._get_group_chat_id()
        if not group_chat_id:
            logger.warning("Debt payment check images skipped: group_chat_id is not configured")
            return

        for idx, item in enumerate(items, 1):
            file_id = (item.get("check_image") or "").strip()
            try:
                tg_file = await context.bot.get_file(file_id)
                image_data = await tg_file.download_as_bytearray()
                decorated = self._decorate_debt_payment_check_image(
                    bytes(image_data),
                    idx,
                    item.get("counterparty_name") or "-",
                )
                await context.bot.send_photo(
                    chat_id=group_chat_id,
                    photo=InputFile(decorated, filename=f"qarz_cheki_{idx}.jpg"),
                )
            except Exception:
                logger.exception("Failed to send decorated debt payment check image: %s", file_id)
                try:
                    await context.bot.send_photo(chat_id=group_chat_id, photo=file_id)
                except Exception:
                    logger.exception("Failed to send original debt payment check image: %s", file_id)

    async def _send_tax_info_check_image(self, context: ContextTypes.DEFAULT_TYPE, row) -> None:
        detail = self._tax_info_detail(row)
        file_id = (detail.get("check_image") or "").strip()
        if not file_id:
            return

        group_chat_id = await self._get_group_chat_id()
        if not group_chat_id:
            logger.warning("Tax info check image skipped: group_chat_id is not configured")
            return

        try:
            tg_file = await context.bot.get_file(file_id)
            image_data = await tg_file.download_as_bytearray()
            decorated = self._decorate_tax_info_check_image(
                bytes(image_data),
                detail.get("cash_amount", 0),
            )
            await context.bot.send_photo(
                chat_id=group_chat_id,
                photo=InputFile(decorated, filename="soliq_malumotlari.jpg"),
            )
        except Exception:
            logger.exception("Failed to send decorated tax info check image: %s", file_id)
            try:
                await context.bot.send_photo(
                    chat_id=group_chat_id,
                    photo=file_id,
                    caption="Soliq ma'lumotlari",
                )
            except Exception:
                logger.exception("Failed to send original tax info check image: %s", file_id)

    def _build_expense_detail_lines(self, row) -> list[str]:
        report_data = self._parse_report_data((row or {}).get("report_data"))
        detail = report_data.get("expense_detail")
        if not isinstance(detail, dict) or not detail:
            return []

        lines = ["📌 Chiqim tafsiloti"]
        items = detail.get("items")
        if isinstance(items, list):
            for item in items:
                if isinstance(item, dict):
                    text = (item.get("text") or "").strip()
                elif isinstance(item, str):
                    text = item.strip()
                else:
                    text = ""
                if text:
                    lines.append(f"• {text}")
        if detail.get("payment_type"):
            lines.append(f"💳 To'lov turi: {detail['payment_type']}")
        if detail.get("cash_amount") is not None:
            lines.append(f"💵 Naqd summa: {self._fmt_money(detail['cash_amount'])}")
        if detail.get("paid_to"):
            lines.append(f"🏷️ Kimga berildi: {detail['paid_to']}")
        if detail.get("recipient_name"):
            lines.append(f"👤 Ism: {detail['recipient_name']}")
        if detail.get("recipient_phone"):
            lines.append(f"📞 Telefon: {detail['recipient_phone']}")
        if detail.get("reason"):
            lines.append(f"📝 Sabab: {detail['reason']}")
        return lines if len(lines) > 1 else []

    def _build_payment_method_lines(self, row) -> list[str]:
        report_data = self._parse_report_data((row or {}).get("report_data"))
        methods = report_data.get("payment_methods")
        if not isinstance(methods, dict) or not methods:
            return []
        labels = self._sverka_payment_method_labels()
        lines = ["💳 To'lov turlari"]
        for key, label in labels.items():
            method = methods.get(key)
            if method:
                lines.append(f"• {label}: {method}")
        return lines if len(lines) > 1 else []

    def _build_inline_payment_method_line(self, row, key: str) -> Optional[str]:
        report_data = self._parse_report_data((row or {}).get("report_data"))
        methods = report_data.get("payment_methods")
        if not isinstance(methods, dict):
            return None
        method = methods.get(key)
        if not method:
            return None
        return f"   ↳ To'lov turi: {method}"

    def _build_other_payments_comment_line(self, row) -> Optional[str]:
        report_data = self._parse_report_data((row or {}).get("report_data"))
        comment = report_data.get("other_payments_comment")
        if not isinstance(comment, str):
            return None
        comment = comment.strip()
        if not comment:
            return None
        return f"   ↳ Izoh: {comment}"

    def _tax_info_detail(self, row) -> dict:
        report_data = self._parse_report_data((row or {}).get("report_data"))
        detail = report_data.get("tax_info")
        return detail if isinstance(detail, dict) else {}

    def _build_tax_info_detail_lines(self, row) -> list[str]:
        detail = self._tax_info_detail(row)
        if not detail:
            return []
        lines = ["🧾 Soliq ma'lumotlari"]
        if detail.get("cash_amount") is not None:
            lines.append(f"💵 Naqd summa: {self._fmt_money(detail.get('cash_amount'))}")
        if detail.get("check_image"):
            lines.append("📷 Chek rasmi biriktirilgan")
        return lines

    def _format_other_payments_value(self, row) -> str:
        try:
            amount = float((row or {}).get("other_payments") or 0)
        except Exception:
            amount = 0.0
        if amount > 0:
            return self._fmt_money(amount)
        if self._build_other_payments_comment_line(row):
            return "izoh kiritilgan"
        return self._fmt_money(amount)

    def _calculate_total_balance(self, row) -> float:
        def _num(key: str) -> float:
            try:
                return float(row.get(key) or 0)
            except Exception:
                return 0.0

        return (
            _num("sales_amount")
            + _num("debt_received")
            + _num("debt_refunds")
            - _num("expenses")
            - _num("uzcard_amount")
            - _num("humo_amount")
            - _num("p2p_amount")
            - _num("uzcard_refund")
            - _num("humo_refund")
            - _num("other_payments")
            - _num("debt_payments")
        )

    def _calculate_sverka_summary_balance(self, row, *, closing: bool = False) -> float:
        if not closing:
            return self._calculate_total_balance(row)
        visible_row = dict(row or {})
        for key in (
            "uzcard_amount",
            "humo_amount",
            "p2p_amount",
            "uzcard_refund",
            "humo_refund",
            "debt_refunds",
        ):
            visible_row[key] = 0
        return self._calculate_total_balance(visible_row)

    async def _get_shift_summary(self, shift_id: int):
        return await self.db.fetch_one(
            """
            SELECT
                s.id,
                s.opened_at,
                s.closed_at,
                s.opening_amount,
                s.closing_amount,
                u.first_name,
                u.last_name,
                u.phone_number,
                l.name AS location,
                COALESCE(r.sales_amount, 0) AS sales_amount,
                COALESCE(r.debt_received, 0) AS debt_received,
                COALESCE(r.expenses, 0) AS expenses,
                COALESCE(r.uzcard_amount, 0) AS uzcard_amount,
                COALESCE(r.humo_amount, 0) AS humo_amount,
                COALESCE(r.p2p_amount, 0) AS p2p_amount,
                COALESCE(r.uzcard_refund, 0) AS uzcard_refund,
                COALESCE(r.humo_refund, 0) AS humo_refund,
                COALESCE(r.other_payments, 0) AS other_payments,
                COALESCE(r.debt_payments, 0) AS debt_payments,
                COALESCE(r.debt_refunds, 0) AS debt_refunds,
                COALESCE(r.report_data, '{}'::jsonb) AS report_data
            FROM shifts s
            JOIN users u ON s.user_id = u.id
            JOIN locations l ON s.location_id = l.id
            LEFT JOIN LATERAL (
                SELECT
                    sales_amount,
                    debt_received,
                    expenses,
                    uzcard_amount,
                    humo_amount,
                    p2p_amount,
                    uzcard_refund,
                    humo_refund,
                    other_payments,
                    debt_payments,
                    debt_refunds,
                    report_data
                FROM reports
                WHERE shift_id = s.id AND report_type = 'daily_report'
                ORDER BY id DESC
                LIMIT 1
            ) r ON TRUE
            WHERE s.id = %s
            """,
            (shift_id,),
        )

    def _build_shift_summary_message(self, row) -> str:
        cashier_name = f"{row.get('first_name', '')} {row.get('last_name') or ''}".strip() or "Kassir"
        report_date = str(row.get("opened_at") or "")[:10] or self._now_tashkent().strftime("%Y-%m-%d")
        total_balance = self._calculate_total_balance(row)
        lines = [
            "📊 Kunlik umumiy hisobot",
            f"👤 Kassir: {cashier_name}",
            f"🏬 Filial: {row.get('location') or '-'}",
            f"📅 Sana: {report_date}",
            f"🕘 Ochilish vaqti: {self._fmt_datetime(row.get('opened_at'))}",
            f"🕙 Yopilish vaqti: {self._fmt_datetime(row.get('closed_at'))}",
            f"💵 Ochilish summasi: {self._fmt_money(row.get('opening_amount'))}",
            f"💰 Yopish summasi: {self._fmt_money(row.get('closing_amount'))}",
            f"🧮 Naqd kutiladigan summa: {self._fmt_money(total_balance)}",
            "",
            "🧾 Sverka",
            f"💸 Savdo: {self._fmt_money(row.get('sales_amount'))}",
            f"📥 Kelgan qarz: {self._fmt_money(row.get('debt_received'))}",
            f"📉 Chiqim: {self._fmt_money(row.get('expenses'))}",
            f"💳 Uzcard: {self._fmt_money(row.get('uzcard_amount'))}",
            f"💳 Humo: {self._fmt_money(row.get('humo_amount'))}",
            f"💳 P2P: {self._fmt_money(row.get('p2p_amount'))}",
            f"↩️ Uzcard vozvrat: {self._fmt_money(row.get('uzcard_refund'))}",
            f"↩️ Humo vozvrat: {self._fmt_money(row.get('humo_refund'))}",
            f"🧷 Boshqa to'lovlar: {self._format_other_payments_value(row)}",
            f"🤝 Qarzga berilgan to'lovlar: {self._fmt_money(row.get('debt_payments'))}",
            f"🔁 Vozvrat qarzlar: {self._fmt_money(row.get('debt_refunds'))}",
        ]
        other_payments_comment_line = self._build_other_payments_comment_line(row)
        if other_payments_comment_line:
            other_payments_index = lines.index(f"🧷 Boshqa to'lovlar: {self._format_other_payments_value(row)}")
            lines.insert(other_payments_index + 1, other_payments_comment_line)
        debt_refunds_method_line = self._build_inline_payment_method_line(row, "debt_refunds")
        if debt_refunds_method_line:
            debt_refunds_index = lines.index(f"🔁 Vozvrat qarzlar: {self._fmt_money(row.get('debt_refunds'))}")
            lines.insert(debt_refunds_index + 1, debt_refunds_method_line)
        debt_received_lines = self._build_debt_received_detail_lines(row)
        if debt_received_lines:
            lines.extend(["", *debt_received_lines])
        debt_payments_lines = self._build_debt_payments_detail_lines(row)
        if debt_payments_lines:
            lines.extend(["", *debt_payments_lines])
        expense_lines = self._build_expense_detail_lines(row)
        if expense_lines:
            lines.extend(["", *expense_lines])
        tax_info_lines = self._build_tax_info_detail_lines(row)
        if tax_info_lines:
            lines.extend(["", *tax_info_lines])
        return "\n".join(lines)

    def _build_sverka_summary_message(
        self,
        row,
        *,
        title: str = "🧾 Sverka yakunlandi",
        closing: bool = False,
    ) -> str:
        cashier_name = f"{row.get('first_name', '')} {row.get('last_name') or ''}".strip() or "Kassir"
        report_date = str(row.get("opened_at") or "")[:10] or self._now_tashkent().strftime("%Y-%m-%d")
        total_balance = self._calculate_sverka_summary_balance(row, closing=closing)
        lines = [
            title,
            f"👤 Kassir: {cashier_name}",
            f"🏬 Filial: {row.get('location') or '-'}",
            f"📅 Sana: {report_date}",
            f"💸 Savdo: {self._fmt_money(row.get('sales_amount'))}",
            f"📥 Kelgan qarz: {self._fmt_money(row.get('debt_received'))}",
            f"📉 Chiqim: {self._fmt_money(row.get('expenses'))}",
        ]
        if not closing:
            lines.extend(
                [
                    f"💳 Uzcard: {self._fmt_money(row.get('uzcard_amount'))}",
                    f"💳 Humo: {self._fmt_money(row.get('humo_amount'))}",
                    f"💳 P2P: {self._fmt_money(row.get('p2p_amount'))}",
                    f"↩️ Uzcard vozvrat: {self._fmt_money(row.get('uzcard_refund'))}",
                    f"↩️ Humo vozvrat: {self._fmt_money(row.get('humo_refund'))}",
                ]
            )
        lines.extend(
            [
                f"🧷 Boshqa to'lovlar: {self._format_other_payments_value(row)}",
                f"🤝 Qarzga berilgan to'lovlar: {self._fmt_money(row.get('debt_payments'))}",
            ]
        )
        if not closing:
            lines.append(f"🔁 Vozvrat qarzlar: {self._fmt_money(row.get('debt_refunds'))}")
        lines.append(f"🧮 Naqd kutiladigan summa: {self._fmt_money(total_balance)}")
        other_payments_comment_line = self._build_other_payments_comment_line(row)
        if other_payments_comment_line:
            other_payments_index = lines.index(f"🧷 Boshqa to'lovlar: {self._format_other_payments_value(row)}")
            lines.insert(other_payments_index + 1, other_payments_comment_line)
        debt_refunds_method_line = self._build_inline_payment_method_line(row, "debt_refunds")
        if debt_refunds_method_line and not closing:
            debt_refunds_index = lines.index(f"🔁 Vozvrat qarzlar: {self._fmt_money(row.get('debt_refunds'))}")
            lines.insert(debt_refunds_index + 1, debt_refunds_method_line)
        debt_received_lines = self._build_debt_received_detail_lines(row)
        if debt_received_lines:
            lines.extend(["", *debt_received_lines])
        debt_payments_lines = self._build_debt_payments_detail_lines(row)
        if debt_payments_lines:
            lines.extend(["", *debt_payments_lines])
        expense_lines = self._build_expense_detail_lines(row)
        if expense_lines:
            lines.extend(["", *expense_lines])
        tax_info_lines = self._build_tax_info_detail_lines(row)
        if tax_info_lines:
            lines.extend(["", *tax_info_lines])
        return "\n".join(lines)

    def _build_shift_document_caption(self, title: str, row) -> str:
        cashier_name = f"{row.get('first_name', '')} {row.get('last_name') or ''}".strip() or "Kassir"
        report_date = str(row.get("opened_at") or "")[:10] or self._now_tashkent().strftime("%Y-%m-%d")
        return (
            f"{title}\n"
            f"👤 {cashier_name}\n"
            f"🏬 {row.get('location') or '-'}\n"
            f"📅 {report_date}"
        )

    def _build_export_caption(self, title: str, file_type: str) -> str:
        return f"📊 {title}\n📎 Format: {file_type}\n✅ Fayl tayyor."

    def _build_shift_opened_message(self, cashier_name: str, location_name: str, opening_amount, opening_time: str) -> str:
        return (
            f"Smena ochildi: {cashier_name}\n"
            f"Filial: {location_name}\n"
            f"Ochish summasi: {self._fmt_money(opening_amount)}\n"
            f"Vaqt: {opening_time or '-'}"
        )

    def _build_shift_closed_message(self, row, note: str) -> str:
        cashier_name = f"{row.get('first_name', '')} {row.get('last_name') or ''}".strip() or "Kassir"
        safe_note = (note or "").strip() or "Izoh qoldirilmadi"
        lines = [
            "🔒 Smena yopildi",
            f"👤 Kassir: {cashier_name}",
            f"🏬 Filial: {row.get('location') or '-'}",
            f"💰 Yopish summasi: {self._fmt_money(row.get('closing_amount'))}",
            f"🕙 Yopilish vaqti: {self._fmt_datetime(row.get('closed_at'))}",
            f"📝 Izoh: {safe_note}",
        ]
        try:
            expenses = float(row.get("expenses") or 0)
        except Exception:
            expenses = 0.0
        if expenses > 0:
            lines.extend(["", f"📉 Chiqim: {self._fmt_money(row.get('expenses'))}"])
            expense_lines = self._build_expense_detail_lines(row)
            if expense_lines:
                lines.extend(["", *expense_lines])
        return "\n".join(lines)

    def _build_payment_image_uploaded_message(self, image_title: str, shift_meta: dict, event_time) -> str:
        return (
            f"✅ To'lov rasmi qabul qilindi\n"
            f"📷 Tur: {image_title}\n"
            f"👤 Kassir: {shift_meta.get('cashier') or '-'}\n"
            f"🏬 Filial: {shift_meta.get('location') or '-'}\n"
            f"⏰ Vaqt: {self._format_telegram_time(event_time)}"
        )

    def _build_report_edit_success_message(
        self,
        cashier_name: str,
        location_name: str,
        field_label: str,
        amount,
        event_time=None,
    ) -> str:
        return (
            "✏️ Hisobot yangilandi\n"
            f"👤 Kassir: {cashier_name or '-'}\n"
            f"🏬 Filial: {location_name or '-'}\n"
            f"🧾 Band: {field_label}\n"
            f"💰 Yangi qiymat: {self._fmt_money(amount)}\n"
            f"⏰ Vaqt: {self._format_telegram_time(event_time)}"
        )
        
    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start command handler"""
        # Reset transient states on every /start to avoid stale flows
        context.user_data.pop('admin_reports_range_pending', None)
        context.user_data.pop('admin_reports_range_values', None)
        context.user_data.pop('pending_sverka_key', None)
        context.user_data.pop('pending_sverka_state', None)
        context.user_data.pop('pending_edit_key', None)
        context.user_data.pop('pending_payment_image', None)
        self._clear_debt_received_detail_state(context)
        self._clear_debt_payments_detail_state(context)
        self._clear_expense_detail_state(context)
        self._clear_generic_payment_method_state(context)
        self._clear_cashier_resume_action(context)
        context.user_data.pop("pending_opening_group_photos", None)
        context.user_data.pop("opening_stage_locked_media_group_id", None)
        context.user_data.pop("opening_stage_locked_name", None)
        context.user_data.pop("pending_next_opening_stage", None)
        context.user_data.pop("opening_stage_completed_prompt_sent", None)
        self._cancel_receipt_roll_finalize_task(context)
        context.user_data.pop("receipt_roll_finalize_token", None)
        context.user_data.pop("opening_finalize_done", None)
        context.user_data['flow'] = None
        try:
            # Best-effort: if DB restarted, reconnect automatically.
            await self.db._ensure_connection()
        except Exception:
            pass
        user = update.effective_user
        # If user already exists, skip registration flow
        try:
            existing = await self.db.fetch_one(
                CommonQueries.ACTIVE_USER_BY_TELEGRAM_ID,
                (user.id,)
            )
        except Exception as e:
            logger.exception("start(): DB error")
            if update.message:
                await update.message.reply_text("Serverda xatolik. Keyinroq qayta urinib ko'ring.")
            return ConversationHandler.END
        if not existing:
            approved_req = await self.db.fetch_one(
                CommonQueries.ACTIVE_APPROVED_REQUEST_BY_TELEGRAM_ID,
                (user.id,)
            )
            if approved_req:
                await self.db.execute_query(
                    CommonQueries.INSERT_APPROVED_CASHIER_USER,
                    (
                        approved_req['telegram_id'],
                        approved_req['first_name'],
                        approved_req['last_name'],
                        approved_req['phone_number'],
                        None
                    )
                )
                existing = await self.db.fetch_one(
                    CommonQueries.ACTIVE_USER_BY_TELEGRAM_ID,
                    (user.id,)
                )
        if existing:
            if existing['role'] == 'admin':
                await update.message.reply_text("Xush kelibsiz!")
                await update.message.reply_text("Administrator menyusi:")
                await self.show_admin_menu(update, context)
                return ConversationHandler.END
            # cashier: ask password each time
            cashier_name = (existing.get('first_name') or user.first_name or "").strip()
            await update.message.reply_text(f"Xush kelibsiz, {cashier_name}!")
            context.user_data['cashier_pending_password'] = True
            await update.message.reply_text("Parolni kiriting:")
            return ConversationHandler.END
        # New users go directly to role selection (no language step)
        context.user_data['language'] = 'uz'
        keyboard = [
            [InlineKeyboardButton("Admin", callback_data='role_admin')],
            [InlineKeyboardButton("Kassir", callback_data='role_cashier')]
        ]
        await update.message.reply_text(
            f"Assalomu alaykum, {user.first_name}!\nRol tanlang:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return SELECT_ROLE

    async def select_role(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle role selection"""
        query = update.callback_query
        await query.answer()
        
        role = query.data.split('_')[1]
        context.user_data['role'] = role

        await query.edit_message_text("Ismingizni kiriting:")
        if role == 'admin':
            return ADMIN_LOGIN
        else:
            return REGISTER_FIRSTNAME

    async def set_group(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Bind the current Telegram group to report forwarding."""
        message = update.effective_message
        chat = update.effective_chat
        user = update.effective_user

        if not message or not chat or not user:
            return

        if chat.type not in ("group", "supergroup"):
            await message.reply_text("Bu buyruqni faqat guruh ichida yuboring.")
            return

        admin = await self.db.fetch_one(
            AdminQueries.ACTIVE_ADMIN_ID_BY_TELEGRAM_ID,
            (user.id,)
        )
        if not admin:
            await message.reply_text("Bu buyruq faqat admin uchun ishlaydi.")
            return

        ok = await self.db.execute_query(
            CommonQueries.UPSERT_BOT_GROUP,
            (chat.id, user.id)
        )
        if not ok:
            await message.reply_text("Guruhni saqlashda xatolik bo'ldi. Keyinroq qayta urinib ko'ring.")
            return

        chat_title = chat.title or str(chat.id)
        self._group_chat_id_cache = int(chat.id)
        await message.reply_text(
            f"Ulandi: {chat_title}\nEndi hisobotlar shu guruhga yuboriladi."
        )

    async def cancel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Cancel active flow and clear deferred cashier action."""
        self._clear_cashier_resume_action(context)
        context.user_data.pop("pending_opening_group_photos", None)
        context.user_data.pop("opening_stage_locked_media_group_id", None)
        context.user_data.pop("opening_stage_locked_name", None)
        context.user_data.pop("pending_next_opening_stage", None)
        context.user_data.pop("opening_stage_completed_prompt_sent", None)
        self._cancel_receipt_roll_finalize_task(context)
        context.user_data.pop("receipt_roll_finalize_token", None)
        context.user_data.pop("opening_finalize_done", None)
        context.user_data['flow'] = None
        context.user_data.pop('pending_sverka_key', None)
        context.user_data.pop('pending_sverka_state', None)
        context.user_data.pop('pending_edit_key', None)
        context.user_data.pop('pending_payment_image', None)
        context.user_data.pop('opening_stage', None)
        context.user_data['cashier_pending_password'] = False
        context.user_data['cashier_set_password'] = False
        context.user_data['cashier_set_password_confirm'] = False
        self._clear_debt_received_detail_state(context)
        self._clear_debt_payments_detail_state(context)
        self._clear_expense_detail_state(context)
        self._clear_generic_payment_method_state(context)

        message = update.effective_message
        user = await self.db.fetch_one(
            CommonQueries.ACTIVE_USER_BY_TELEGRAM_ID,
            (update.effective_user.id,),
        )
        if not message:
            return ConversationHandler.END

        await message.reply_text("Jarayon bekor qilindi.")
        if user and user.get("role") == "admin":
            await self.show_admin_menu(update, context)
        elif user and user.get("role") == "cashier":
            await self.show_cashier_menu(update, context)
        return ConversationHandler.END

    async def register_firstname(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get user's first name"""
        context.user_data['first_name'] = update.message.text
        await update.message.reply_text("Familiyangizni kiriting:")
        return REGISTER_LASTNAME

    async def register_lastname(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get user's last name"""
        context.user_data['last_name'] = update.message.text
        await self._ask_for_phone_contact(
            update.message,
            "Telefon raqamingizni yuborish uchun pastdagi tugmani bosing:"
        )
        return REGISTER_PHONE

    async def register_phone(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get user's phone number"""
        phone, error_message = self._extract_shared_phone(update)

        if error_message:
            await self._ask_for_phone_contact(update.message, error_message)
            return REGISTER_PHONE

        context.user_data['phone'] = phone

        # Create or update approval request (password will be set after approval)
        user_data = {
            'telegram_id': update.effective_user.id,
            'first_name': context.user_data.get('first_name', ''),
            'last_name': context.user_data.get('last_name', ''),
            'phone_number': context.user_data.get('phone', ''),
            'role': 'cashier'
        }

        existing_req = await self.db.fetch_one(
            CommonQueries.PENDING_REQUEST_BY_TELEGRAM_ID,
            (user_data['telegram_id'],)
        )
        if existing_req:
            await self.db.execute_query(
                CommonQueries.UPDATE_PENDING_APPROVAL_REQUEST,
                (user_data['first_name'], user_data['last_name'], user_data['phone_number'], user_data['telegram_id'])
            )
        else:
            await self.db.execute_query(CommonQueries.INSERT_PENDING_APPROVAL_REQUEST, user_data)

        # Notify admins about new cashier request
        await self.notify_admins_new_request(context, user_data)

        lang = 'uz'
        if lang == 'uz':
            msg = "So'rovingiz administratorga jo'natildi. Tasdiqlanganidan keyin parol o'rnatasiz."
        else:
            msg = "Р В РІР‚в„ўР В Р’В°Р РЋРІвЂљВ¬ Р В Р’В·Р В Р’В°Р В РЎвЂ”Р РЋР вЂљР В РЎвЂўР РЋР С“ Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦ Р В Р’В°Р В РўвЂР В РЎВР В РЎвЂР В Р вЂ¦Р В РЎвЂР РЋР С“Р РЋРІР‚С™Р РЋР вЂљР В Р’В°Р РЋРІР‚С™Р В РЎвЂўР РЋР вЂљР РЋРЎвЂњ. Р В РЎСџР В РЎвЂўР РЋР С“Р В Р’В»Р В Р’Вµ Р В РЎвЂўР В РўвЂР В РЎвЂўР В Р’В±Р РЋР вЂљР В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ Р В Р вЂ Р РЋРІР‚в„– Р РЋРЎвЂњР РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂ”Р В Р’В°Р РЋР вЂљР В РЎвЂўР В Р’В»Р РЋР Р‰."

        await update.message.reply_text(msg, reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END

    async def register_password(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get user's password"""
        password = update.message.text
        context.user_data['password'] = hash_password(password)
        
        lang = 'uz'
        
        if lang == 'uz':
            msg = "Parolni qaytadan kiriting:"
        else:
            msg = "Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°:"
        
        await update.message.reply_text(msg)
        return VERIFY_PASSWORD

    async def verify_password(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Verify password match"""
        password = update.message.text
        stored_password = context.user_data['password']
        
        if verify_password(stored_password, password):
            # Store user data temporarily for admin approval
            user_data = {
                'telegram_id': update.effective_user.id,
                'first_name': context.user_data['first_name'],
                'last_name': context.user_data['last_name'],
                'phone_number': context.user_data['phone'],
                'role': 'cashier'
            }
            
            # Insert or update approval request
            user_data['password_hash'] = stored_password
            existing_req = await self.db.fetch_one(
                CommonQueries.PENDING_REQUEST_BY_TELEGRAM_ID,
                (user_data['telegram_id'],)
            )
            if existing_req:
                await self.db.execute_query(
                    CommonQueries.UPDATE_PENDING_APPROVAL_REQUEST_WITH_PASSWORD,
                    (
                        user_data['first_name'],
                        user_data['last_name'],
                        user_data['phone_number'],
                        user_data['password_hash'],
                        user_data['telegram_id']
                    )
                )
            else:
                await self.db.execute_query(CommonQueries.INSERT_PENDING_APPROVAL_REQUEST_WITH_PASSWORD, user_data)

            # Notify admins about new cashier request
            await self.notify_admins_new_request(context, user_data)
            
            lang = 'uz'
            
            if lang == 'uz':
                msg = "So'rovingiz administratorga jo'natildi. Tasdiqlanganidan keyin botdan foydalanishingiz mumkin."
            else:
                msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†РІР‚С™Р’В¬ Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р Р‹Р В РЎвЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦ Р В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р Р‹Р РЋРІР‚Сљ. Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РЎвЂњР В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В¶Р В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В Р РЏ Р В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р Р‹Р В РЎвЂњР В Р’В Р РЋР’ВР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°Р В Р’В Р вЂ™Р’В·Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В Р РЏ Р В Р’В Р вЂ™Р’В±Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р’В Р РЋР’В."
            
            await update.message.reply_text(msg)
            return ConversationHandler.END
        else:
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Parollar mos kelmadi! Iltimos, qaytadan kiriting:"
            else:
                msg = "Р В Р’В Р РЋРЎСџР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚В Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚в„–Р В Р Р‹Р Р†Р вЂљРЎв„ў! Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р В РЎвЂњР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°:"
            
            await update.message.reply_text(msg)
            return REGISTER_PASSWORD

    async def admin_login(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Admin login flow"""
        # If user is cashier, require admin promotion
        existing_user = await self.db.fetch_one(
            CommonQueries.USER_BY_TELEGRAM_ID,
            (update.effective_user.id,)
        )
        if existing_user and existing_user.get('role') == 'cashier':
            await update.message.reply_text("Admin bo'lish uchun mavjud admin tasdiqlashi kerak.")
            return ConversationHandler.END

        # Allow up to 2 admins
        admin_count = await self.db.fetch_one(AdminQueries.ACTIVE_ADMIN_COUNT)
        if admin_count and int(admin_count.get('cnt', 0)) >= 2:
            # If this user is already admin, allow login; otherwise block
            existing_admin = await self.db.fetch_one(
                AdminQueries.ACTIVE_ADMIN_BY_TELEGRAM_ID,
                (update.effective_user.id,)
            )
            if not existing_admin:
                await update.message.reply_text("Adminlar limiti 2 ta. Hozircha yangi admin qo'shib bo'lmaydi.")
                return ConversationHandler.END

        # Check if user is already registered as admin
        result = await self.db.fetch_one(AdminQueries.ACTIVE_ADMIN_BY_TELEGRAM_ID, (update.effective_user.id,))

        if result:
            # Admin already exists
            lang = 'uz'
            
            if lang == 'uz':
                msg = f"Xush kelibsiz, Administrator!"
            else:
                msg = f"Р В Р’В Р Р†Р вЂљРЎСљР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В±Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ°, Р В Р’В Р РЋРІР‚в„ўР В Р’В Р СћРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™!"
                
            await update.message.reply_text(msg)
            await self.show_admin_menu(update, context)
            return MAIN_MENU
        else:
            # New admin registration
            context.user_data['first_name'] = update.effective_user.first_name
            context.user_data['last_name'] = update.effective_user.last_name
            context.user_data['phone'] = ""  # Will be collected
            
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Telefon raqamingizni yuborish uchun pastdagi tugmani bosing:"
            else:
                msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†РІР‚С™Р’В¬ Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™ Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’ВµР В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎвЂєР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В° (Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™: +998901234567):"
                
            await self._ask_for_phone_contact(update.message, msg)
            return ADMIN_REGISTER_PHONE

    async def admin_register_phone(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get admin's phone number"""
        phone, error_message = self._extract_shared_phone(update)
        
        if error_message:
            lang = 'uz'
            
            if lang == 'uz':
                msg = error_message
            else:
                msg = "Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р Р‹Р Р†Р вЂљРЎвЂєР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ў Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В° Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’ВµР В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎвЂєР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°! Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р В РЎвЂњР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В° (Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™: +998901234567):"
                
            await self._ask_for_phone_contact(update.message, msg)
            return ADMIN_REGISTER_PHONE
        
        context.user_data['phone'] = phone
        
        lang = 'uz'
        
        if lang == 'uz':
            msg = "Parol kiriting:"
        else:
            msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°:"
        
        await update.message.reply_text(msg, reply_markup=ReplyKeyboardRemove())
        return ADMIN_REGISTER_PASSWORD

    async def admin_register_password(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get admin's password"""
        password = update.message.text
        context.user_data['password'] = hash_password(password)
        
        lang = 'uz'
        
        if lang == 'uz':
            msg = "Parolni qaytadan kiriting:"
        else:
            msg = "Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°:"
        
        await update.message.reply_text(msg)
        return ADMIN_VERIFY_PASSWORD

    async def admin_verify_password(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Verify admin password match"""
        password = update.message.text
        stored_password = context.user_data['password']
        
        if verify_password(stored_password, password):
            # Register admin user
            user_data = {
                'telegram_id': update.effective_user.id,
                'first_name': context.user_data['first_name'],
                'last_name': context.user_data['last_name'],
                'phone_number': context.user_data['phone'],
                'role': 'admin',
                'password_hash': context.user_data['password']
            }
            
            query = """
                INSERT INTO users (telegram_id, first_name, last_name, phone_number, role, password_hash)
                VALUES (%(telegram_id)s, %(first_name)s, %(last_name)s, %(phone_number)s, %(role)s, %(password_hash)s)
            """
            await self.db.execute_query(query, user_data)
            
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Siz muvaffaqiyatli ro'yxatdan o'tdingiz, Administrator!"
            else:
                msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р Р‹Р РЋРІР‚СљР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†РІР‚С™Р’В¬Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’ВµР В Р’В Р РЋРІР‚вЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р В Р вЂ°, Р В Р’В Р РЋРІР‚в„ўР В Р’В Р СћРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™!"
                
            await update.message.reply_text(msg)
            await self.show_admin_menu(update, context)
            return MAIN_MENU
        else:
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Parollar mos kelmadi! Iltimos, qaytadan kiriting:"
            else:
                msg = "Р В Р’В Р РЋРЎСџР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚В Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р РЋРІР‚вЂќР В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РІР‚в„–Р В Р Р‹Р Р†Р вЂљРЎв„ў! Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р В РЎвЂњР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°:"
            
            await update.message.reply_text(msg)
            return ADMIN_REGISTER_PASSWORD

    async def show_admin_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show admin menu"""
        context.user_data['admin_reports_range_pending'] = False
        menu_text = "Administrator menyusi:"
        message = getattr(update, "effective_message", None) or getattr(update, "message", None)
        if not message:
            return

        reply_markup = ReplyKeyboardMarkup(
            [[KeyboardButton(label) for label in row] for row in ADMIN_MENU_ROWS],
            resize_keyboard=True,
        )
        await message.reply_text(menu_text, reply_markup=reply_markup)

    async def show_admin_reports_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show admin report period options."""
        message = getattr(update, "effective_message", None) or getattr(update, "message", None)
        if not message:
            return
        reply_markup = ReplyKeyboardMarkup(
            [[KeyboardButton(label) for label in row] for row in ADMIN_REPORTS_MENU_ROWS],
            resize_keyboard=True,
        )
        await message.reply_text("Qaysi hisobot kerak?", reply_markup=reply_markup)

    async def show_cashier_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show cashier menu"""
        menu_text = "Kassir menyusi:"
        reply_markup = self._build_cashier_menu_keyboard()
        message = getattr(update, "effective_message", None) or getattr(update, "message", None)
        if not message:
            return
        await message.reply_text(menu_text, reply_markup=reply_markup)

    async def handle_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle general messages based on user state"""
        text = update.message.text
        user_id = update.effective_user.id

        # If cashier is mid-flow, let ConversationHandler handle and avoid menu spam
        if context.user_data.get('flow') in ['opening', 'sverka', 'closing', 'edit', 'payment_image']:
            if context.user_data.get('flow') == 'payment_image' and context.user_data.get('pending_payment_image'):
                if text in CASHIER_MENU_TEXTS:
                    context.user_data.pop('pending_payment_image', None)
                    context.user_data['flow'] = None
                    # oqimni tozaladik, endi pastdagi oddiy menyu dispatch ishlasin
                else:
                    await update.message.reply_text("Iltimos, rasm yuboring (foto yoki image fayl).")
                    return
            if context.user_data.get('flow') == 'sverka' and context.user_data.get('pending_sverka_key'):
                key = context.user_data.get('pending_sverka_key')
                handlers = {
                    'sales_amount': self.report_sales,
                    'debt_received': self.report_debt_received,
                    'expenses': self.report_expenses,
                    'uzcard_amount': self.report_uzcard,
                    'humo_amount': self.report_humo,
                    'p2p_amount': self.report_p2p,
                    'uzcard_refund': self.report_uzcard_refund,
                    'humo_refund': self.report_humo_refund,
                    'other_payments': self.report_other_payments,
                    'debt_payments': self.report_debt_payments,
                    'debt_refunds': self.report_debt_refunds,
                    'tax_info': self.report_tax_info,
                }
                handler = handlers.get(key)
                if handler:
                    await handler(update, context)
                return
            # Fallback: debt_payments multi-step davom etayotgan bo'lsa (ConversationHandler state yo'qolgan)
            if context.user_data.get('flow') == 'sverka' and context.user_data.get('debt_payments_detail_stage'):
                await self.report_debt_payments(update, context)
                return
            if context.user_data.get('flow') == 'edit' and context.user_data.get('pending_edit_key'):
                await self.edit_reports_value(update, context)
                return
            # Fallback: agar ConversationHandler state yo'qolgan bo'lsa ham oqim davom etsin
            if context.user_data.get('flow') == 'closing':
                if context.user_data.get("pending_close_amount") is not None:
                    await self.close_shift_note(update, context)
                else:
                    await self.close_shift(update, context)
                return
            if context.user_data.get('flow') == 'opening':
                # Location tanlanganidan keyin summa kiritish bosqichi
                if context.user_data.get('location_id') and not context.user_data.get('opening_stage'):
                    await self.open_shift_amount(update, context)
                    return
                # Rasm bosqichlarida matn yuborilsa tushunarli ogohlantirish beramiz
                await update.message.reply_text("Iltimos, rasm yuboring.")
                return
        
        # Check if user is admin or cashier
        user = await self.db.fetch_one(CommonQueries.ACTIVE_USER_BY_TELEGRAM_ID, (user_id,))
        
        if not user:
            approved_req = await self.db.fetch_one(
                CommonQueries.ACTIVE_APPROVED_REQUEST_BY_TELEGRAM_ID,
                (user_id,)
            )
            if approved_req:
                await self.db.execute_query(
                    CommonQueries.INSERT_APPROVED_CASHIER_USER,
                    (
                        approved_req['telegram_id'],
                        approved_req['first_name'],
                        approved_req['last_name'],
                        approved_req['phone_number'],
                        None
                    )
                )
                user = await self.db.fetch_one(CommonQueries.ACTIVE_USER_BY_TELEGRAM_ID, (user_id,))
                if user:
                    cashier_name = (user.get('first_name') or update.effective_user.first_name or "").strip()
                    await update.message.reply_text(f"Xush kelibsiz, {cashier_name}!")
                    context.user_data['cashier_pending_password'] = True
                    await update.message.reply_text("Parolni kiriting:")
                    return
            # Check if it's a pending approval request
            req = await self.db.fetch_one(CommonQueries.PENDING_REQUEST_BY_TELEGRAM_ID, (user_id,))
            if req:
                lang = 'uz'
                
                if lang == 'uz':
                    msg = "Siz hali administrator tomonidan tasdiqlanmadingiz."
                else:
                    msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљР’В°Р В Р’В Р вЂ™Р’Вµ Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В¶Р В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚ВР В Р’В Р РЋР’ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р РЋР’В."
                
                await update.message.reply_text(msg)
                return ConversationHandler.END
            else:
                if text in KNOWN_MENU_TEXTS:
                    await update.message.reply_text("Avval /start buyrug'ini yuboring.")
                else:
                    await update.message.reply_text("Siz ro'yxatdan o'tmagansiz. Davom etish uchun /start yuboring.")
                return ConversationHandler.END
        
        # Determine user role and handle accordingly
        if user['role'] == 'admin':
            await self.handle_admin_command(update, context, user)
        else:
            # If password missing, force set new password
            if not user.get('password_hash'):
                if context.user_data.get('cashier_set_password'):
                    context.user_data['new_password_hash'] = hash_password(text)
                    context.user_data['cashier_set_password'] = False
                    context.user_data['cashier_set_password_confirm'] = True
                    await update.message.reply_text("Parolni qaytadan kiriting:")
                    return
                if context.user_data.get('cashier_set_password_confirm'):
                    if verify_password(context.user_data['new_password_hash'], text):
                        await self.db.execute_query(
                            "UPDATE users SET password_hash = %s WHERE telegram_id = %s",
                            (context.user_data['new_password_hash'], user_id)
                        )
                        context.user_data['cashier_set_password_confirm'] = False
                        context.user_data['cashier_authenticated'] = True
                        await update.message.reply_text("Parol o'rnatildi.")
                        resumed = await self._resume_cashier_post_auth_action(update, context)
                        if not resumed:
                            await self.show_cashier_menu(update, context)
                    else:
                        context.user_data['cashier_set_password'] = True
                        context.user_data['cashier_set_password_confirm'] = False
                        await update.message.reply_text("Parollar mos kelmadi. Yangi parol kiriting:")
                    return
                context.user_data['cashier_set_password'] = True
                await update.message.reply_text("Parol o'rnatilmagan. Yangi parol kiriting:")
                return

            # Require password on each new /start or session
            if context.user_data.get('cashier_pending_password'):
                if text in CASHIER_MENU_TEXTS:
                    self._set_cashier_resume_action(context, CASHIER_DIRECT_ACTIONS.get(text))
                    await update.message.reply_text("Tanlangan amal saqlandi. Avval parolni kiriting.")
                    return
                if user.get('password_hash') and verify_password(user['password_hash'], text):
                    context.user_data['cashier_pending_password'] = False
                    context.user_data['cashier_authenticated'] = True
                    resumed = await self._resume_cashier_post_auth_action(update, context)
                    if not resumed:
                        await self.show_cashier_menu(update, context)
                else:
                    await update.message.reply_text(
                        "Parol noto'g'ri. Qayta kiriting.\n"
                        "Agar parol esdan chiqqan bo'lsa, adminga `reset <telegram_id>` yozdiring."
                    )
                return
            if not context.user_data.get('cashier_authenticated'):
                context.user_data['cashier_pending_password'] = True
                if text in CASHIER_MENU_TEXTS:
                    self._set_cashier_resume_action(context, CASHIER_DIRECT_ACTIONS.get(text))
                await update.message.reply_text("Parolni kiriting:")
                return
            await self.handle_cashier_command(update, context, user)

    async def handle_image_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle image messages even if ConversationHandler state was lost."""
        self._sync_opening_stage_with_media_group(update, context)
        flow = context.user_data.get('flow')
        if context.user_data.get('pending_payment_image'):
            await self.upload_payment_image(update, context)
            return

        if flow == 'opening' or context.user_data.get('opening_stage'):
            stage = context.user_data.get('opening_stage')
            if stage == 'workplace_status':
                await self.upload_workplace_status(update, context)
                return
            if stage == 'terminal_power':
                await self.upload_terminal_power(update, context)
                return
            if stage == 'zero_report':
                await self.upload_zero_report(update, context)
                return
            if stage == 'opening_notification':
                await self.upload_opening_notification(update, context)
                return
            if stage == 'receipt_roll':
                await self.upload_receipt_roll(update, context)
                return

        # Fallback: Qarz berish chek rasmi (ConversationHandler state yo'qolgan bo'lsa)
        if flow == 'sverka' and context.user_data.get('debt_payments_detail_stage') == 'check_image':
            await self.report_debt_payments(update, context)
            return

        if flow == 'sverka' and context.user_data.get('tax_info_stage') == 'check_image':
            await self.report_tax_info(update, context)
            return

    async def handle_admin_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user):
        """Handle admin commands"""
        text = (update.message.text or "").strip()
        lang = 'uz'

        # If we are waiting for a date range input, don't treat admin menu buttons as invalid format.
        # Cancel the pending range when user presses another menu item.
        if context.user_data.get('admin_reports_range_pending'):
            if text in (ADMIN_MENU_TEXTS | ADMIN_REPORT_TEXTS | EXPORT_MENU_TEXTS):
                context.user_data['admin_reports_range_pending'] = False
            else:
                await self.handle_admin_reports_range(update, context)
                return

        # Approve/reject cashier requests by command text
        if text:
            normalized = text.strip()
            lower = normalized.lower()
            if lower.startswith(("approve ", "tasdiq ", "tasdiqlash ", "confirm ", "accept ", "odobrit ", "Р В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚В˜Р В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В±Р В Р Р‹Р В РІР‚С™ ")):
                target_id = self._extract_telegram_id(normalized)
                if target_id:
                    await self.approve_cashier(update, context, target_id)
                else:
                    await update.message.reply_text("ID topilmadi. Masalan: approve 123456789")
                return
            if lower.startswith(("reject ", "rad ", "otkaz ", "otklon ", "deny ", "Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В· ", "Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В¦ ")):
                target_id = self._extract_telegram_id(normalized)
                if target_id:
                    await self.reject_cashier(update, context, target_id)
                else:
                    await update.message.reply_text("ID topilmadi. Masalan: reject 123456789")
                return
            if lower.startswith(("reset ", "parol ", "Р В РЎвЂ”Р В Р’В°Р РЋР вЂљР В РЎвЂўР В Р’В» ", "Р РЋР С“Р В Р’В±Р РЋР вЂљР В РЎвЂўР РЋР С“ ")):
                target_id = self._extract_telegram_id(normalized)
                if target_id:
                    await self.reset_cashier_password(update, context, target_id)
                else:
                    await update.message.reply_text("ID topilmadi. Masalan: reset 123456789")
                return

        report_period = ADMIN_REPORT_PERIODS.get(text)
        if report_period:
            await self._ask_report_location(update, context, report_period)
            return

        direct_action = ADMIN_DIRECT_ACTIONS.get(text)
        if direct_action:
            await getattr(self, direct_action)(update, context)
            return

        if text == "Vaqt oralig'i":
            await self.send_reports(update, context)
        elif text in EXPORT_MENU_TEXTS:
            await self.handle_export_choice(update, context)
        else:
            if lang == 'uz':
                msg = "Iltimos, menyudan birini tanlang."
            else:
                msg = "Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р вЂ™Р’В±Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦ Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В· Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р РЋРІР‚СљР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В  Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р Р‹Р В РІР‚в„–."
                
            await update.message.reply_text(msg)
            await self.show_admin_menu(update, context)

    async def handle_cashier_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user):
        """Handle cashier commands"""
        text = update.message.text
        lang = 'uz'

        # Let ConversationHandler handle these menu actions

        # Majburiy jarayonlar ishlayotgan paytda noto'g'ri bosqichga o'tishni bloklaymiz.
        # payment_image holatida esa menyu tugmalari bosilsa oqimni tozalab davom etishga ruxsat beramiz.
        active_flow = context.user_data.get('flow')
        if active_flow in ['opening', 'sverka', 'closing', 'payment_image']:
            if active_flow == 'payment_image':
                if text in CASHIER_MENU_TEXTS:
                    context.user_data.pop('pending_payment_image', None)
                    context.user_data['flow'] = None
                else:
                    await update.message.reply_text("Iltimos, rasm yuboring yoki menyudan tugma tanlang.")
                    return
            else:
                await update.message.reply_text("Jarayon davom etmoqda. Iltimos, avval joriy bosqichni yakunlang.")
                return

        direct_action = CASHIER_DIRECT_ACTIONS.get(text)
        if direct_action:
            await getattr(self, direct_action)(update, context)
        else:
            if lang == 'uz':
                msg = f"Xush kelibsiz, {user['first_name']}! Iltimos, menyudan birini tanlang."
            else:
                msg = f"Р В Р’В Р Р†Р вЂљРЎСљР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В±Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ°, {user['first_name']}! Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р вЂ™Р’В±Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦ Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В· Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р РЋРІР‚СљР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В  Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р Р‹Р В РІР‚в„–."
                
            await update.message.reply_text(msg)
            await self.show_cashier_menu(update, context)

    async def _ensure_cashier_authenticated(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        resume_action: Optional[str] = None,
    ) -> bool:
        """Kassir uchun parol o'rnatish/kirishni majburiy tekshiradi."""
        tg_id = update.effective_user.id
        user = await self.db.fetch_one(
            "SELECT * FROM users WHERE telegram_id=%s AND role='cashier' AND is_active=TRUE",
            (tg_id,)
        )
        if not user:
            await update.message.reply_text("Kassir topilmadi. /start bosing.")
            return False

        # Tasdiqlangan, lekin hali parol o'rnatilmagan kassir
        if not user.get('password_hash'):
            context.user_data['cashier_set_password'] = True
            context.user_data['cashier_set_password_confirm'] = False
            context.user_data['cashier_pending_password'] = False
            context.user_data['cashier_authenticated'] = False
            self._set_cashier_resume_action(context, resume_action)
            await update.message.reply_text("Avval parol o'rnating. Yangi parol kiriting:")
            return False

        # Sessiya uchun parol kiritilmagan bo'lsa
        if not context.user_data.get('cashier_authenticated'):
            context.user_data['cashier_pending_password'] = True
            self._set_cashier_resume_action(context, resume_action)
            await update.message.reply_text("Parolni kiriting:")
            return False

        return True

    async def start_shift_opening(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start the shift opening process"""
        if not await self._ensure_cashier_authenticated(update, context, resume_action="start_shift_opening"):
            return MAIN_MENU
        context.user_data.pop('blocked_media_group_id', None)
        context.user_data['pending_opening_group_photos'] = []
        context.user_data.pop("opening_stage_locked_media_group_id", None)
        context.user_data.pop("opening_stage_locked_name", None)
        context.user_data.pop("pending_next_opening_stage", None)
        context.user_data.pop("opening_stage_completed_prompt_sent", None)
        self._cancel_receipt_roll_finalize_task(context)
        context.user_data.pop("receipt_roll_finalize_token", None)
        context.user_data.pop("opening_finalize_done", None)

        lang = 'uz'
        user_row = await self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
        if user_row:
            # Only block if the user currently has an ACTIVE open shift.
            # A previously closed shift does NOT prevent opening a new one.
            active_shift = await self.db.fetch_one(
                "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC LIMIT 1",
                (user_row['id'],)
            )
            if active_shift:
                await update.message.reply_text("Sizda hozirda ochiq smena bor. Avval o'sha smenani yoping.")
                return MAIN_MENU
        
        context.user_data['workplace_status_uploaded_ids'] = []
        context.user_data['opening_stage'] = None
        await self.show_location_selection(update, context)
        context.user_data['flow'] = 'opening'
        return SELECT_LOCATION

    async def open_shift_amount(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get the opening amount for the shift"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data['opening_amount'] = amount
            context.user_data['opening_amount_time'] = self._format_telegram_time(getattr(update.message, "date", None))
            
            # Create shift now that we have location + amount
            location_id = context.user_data.get('location_id')
            if not location_id:
                await update.message.reply_text("Avval filialni tanlang.")
                await self.show_location_selection(update, context)
                return SELECT_LOCATION

            try:
                user_row = await self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
                if user_row:
                    today = self._now_tashkent().date().isoformat()
                    start_bound, end_bound = self._day_bounds(today)

                    # Block if the location already has any shift today (open or closed)
                    location_today_shift = await self.db.fetch_one(
                        """
                        SELECT id, is_open FROM shifts
                        WHERE location_id=%s AND opened_at >= %s AND opened_at < %s
                        ORDER BY id DESC LIMIT 1
                        """,
                        (location_id, start_bound, end_bound)
                    )
                    if location_today_shift:
                        if bool(location_today_shift.get('is_open')):
                            await update.message.reply_text(
                                "Bu filialda hozirda ochiq smena mavjud. "
                                "Avval o'sha smenani yoping."
                            )
                        else:
                            await update.message.reply_text(
                                "Bu filialda bugun smena allaqachon ochilgan va yopilgan. "
                                "Bir kunda bir filial uchun faqat 1 ta smena ochiladi."
                            )
                        await self.show_cashier_menu(update, context)
                        context.user_data['flow'] = None
                        return MAIN_MENU

                    await self.db.execute_query(
                        """
                        INSERT INTO shifts (user_id, location_id, opening_amount, is_open)
                        VALUES (%s, %s, %s, TRUE)
                        """,
                        (user_row['id'], location_id, amount)
                    )
                    shift = await self.db.fetch_one(
                        "SELECT id FROM shifts WHERE user_id=%s ORDER BY opened_at DESC LIMIT 1",
                        (user_row['id'],)
                    )
                    if shift:
                        context.user_data['current_shift_id'] = shift['id']
                        context.user_data['workplace_status_uploaded_ids'] = []
            except Exception:
                context.user_data['flow'] = None
                await update.message.reply_text("Xatolik: smena ma'lumotlarini saqlab bo'lmadi. Qayta urinib ko'ring.")
                await self.show_cashier_menu(update, context)
                return MAIN_MENU

            lang = 'uz'
            if lang == 'uz':
                await update.message.reply_text("Summa tasdiqlandi.")
                msg = "Ish joyingiz tayyorligini tasdiqlang. Ish stolingizni rasmga olib yuboring. (2 ta rasm)"
            else:
                await update.message.reply_text("Р В Р Р‹Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В° Р В РЎвЂ”Р В РЎвЂўР В РўвЂР РЋРІР‚С™Р В Р вЂ Р В Р’ВµР РЋР вЂљР В Р’В¶Р В РўвЂР В Р’ВµР В Р вЂ¦Р В Р’В°.")
                msg = "Р В РЎСџР В РЎвЂўР В РўвЂР РЋРІР‚С™Р В Р вЂ Р В Р’ВµР РЋР вЂљР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂ“Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂўР В Р вЂ Р В Р вЂ¦Р В РЎвЂўР РЋР С“Р РЋРІР‚С™Р РЋР Р‰ Р РЋР вЂљР В Р’В°Р В Р’В±Р В РЎвЂўР РЋРІР‚РЋР В Р’ВµР В РЎвЂ“Р В РЎвЂў Р В РЎВР В Р’ВµР РЋР С“Р РЋРІР‚С™Р В Р’В°. Р В РЎвЂєР РЋРІР‚С™Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р РЋР Р‰Р РЋРІР‚С™Р В Р’Вµ Р РЋРІР‚С›Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂў Р РЋР вЂљР В Р’В°Р В Р’В±Р В РЎвЂўР РЋРІР‚РЋР В Р’ВµР В РЎвЂ“Р В РЎвЂў Р РЋР С“Р РЋРІР‚С™Р В РЎвЂўР В Р’В»Р В Р’В°. (2 Р РЋРІР‚С›Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂў)"
            await update.message.reply_text(msg)
            context.user_data['opening_stage'] = 'workplace_status'
            return UPLOAD_WORKPLACE_STATUS
        except ValueError:
            lang = 'uz'
            
            if lang == 'uz':
                msg = self._invalid_amount_msg(context)
            else:
                msg = "Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р В РІР‚В Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°Р В Р’В Р В РІР‚В¦Р В Р Р‹Р РЋРІР‚СљР В Р Р‹Р В РІР‚в„– Р В Р Р‹Р В РЎвЂњР В Р Р‹Р РЋРІР‚СљР В Р’В Р РЋР’ВР В Р’В Р РЋР’ВР В Р Р‹Р РЋРІР‚Сљ."
                
            await update.message.reply_text(msg)
            return OPEN_SHIFT_AMOUNT

    async def show_location_selection(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show location selection for the shift"""
        # Get all locations from the database
        locations = await self._get_locations()
        
        keyboard = []
        for loc in locations:
            keyboard.append([InlineKeyboardButton(loc['name'], callback_data=f"loc_{loc['id']}")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        msg = "Filialni tanlang:"
        await update.message.reply_text(msg, reply_markup=reply_markup)

    async def ask_select_location_again(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Remind user to select location before entering amount"""
        msg = "Avval filialni tanlang."
        await update.message.reply_text(msg)
        await self.show_location_selection(update, context)

    async def select_location(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle location selection"""
        query = update.callback_query
        if not query or not query.data:
            return SELECT_LOCATION

        await query.answer()

        try:
            location_id = int(query.data.split('_')[1])
        except Exception:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="Filialni qaytadan tanlang.")
            await self.show_location_selection(update, context)
            return SELECT_LOCATION

        context.user_data['location_id'] = location_id
        context.user_data['flow'] = 'opening'

        msg = "Smenani ochish summasini kiriting (faqat summa):"
        try:
            await query.edit_message_text(msg)
        except Exception:
            # Fallback: agar edit ishlamasa ham keyingi bosqichga o'tkazamiz
            await context.bot.send_message(chat_id=update.effective_chat.id, text=msg)
        return OPEN_SHIFT_AMOUNT
    async def upload_workplace_status(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Upload workplace status image (2 ta rasm majburiy)."""
        self._sync_opening_stage_with_media_group(update, context)
        current_stage = context.user_data.get("opening_stage")
        if current_stage and current_stage != "workplace_status":
            if current_stage == "terminal_power":
                return await self.upload_terminal_power(update, context)
            if current_stage == "zero_report":
                return await self.upload_zero_report(update, context)
            if current_stage == "opening_notification":
                return await self.upload_opening_notification(update, context)
            if current_stage == "receipt_roll":
                return await self.upload_receipt_roll(update, context)
        if self._is_blocked_media_group(update, context):
            return UPLOAD_WORKPLACE_STATUS

        file_id = self._get_image_file_id(update)
        if not file_id:
            await update.message.reply_text("Iltimos, rasm yuboring.")
            return UPLOAD_WORKPLACE_STATUS

        shift_id = context.user_data.get('current_shift_id')
        if not shift_id:
            user_row = await self.db.fetch_one("SELECT id FROM users WHERE telegram_id=%s", (update.effective_user.id,))
            if user_row:
                active_shift = await self.db.fetch_one(
                    "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC LIMIT 1",
                    (user_row['id'],)
                )
                if active_shift:
                    shift_id = active_shift['id']
                    context.user_data['current_shift_id'] = shift_id

        if not shift_id:
            context.user_data['flow'] = None
            await update.message.reply_text("Ochiq smena topilmadi. Avval smena oching.")
            await self.show_cashier_menu(update, context)
            return MAIN_MENU

        uploaded_ids = context.user_data.get('workplace_status_uploaded_ids')
        if not isinstance(uploaded_ids, list):
            uploaded_ids = []

        uploaded_ids.append(file_id)
        await self._save_shift_image(shift_id, 'workplace_status', file_id)
        self._queue_opening_group_photo(
            context,
            file_id,
            "Ish joyi holati rasmi",
            event_time=getattr(update.message, "date", None),
            media_kind=self._get_image_media_kind(update),
        )
        context.user_data['workplace_status_uploaded_ids'] = uploaded_ids

        db_count = await self._count_shift_images(shift_id, 'workplace_status')
        count = max(len(uploaded_ids), db_count)

        if count < 2:
            await update.message.reply_text("Rasm qabul qilindi (1/2). Yana bitta rasm yuboring.")
            context.user_data['opening_stage'] = 'workplace_status'
            return UPLOAD_WORKPLACE_STATUS

        if not self._is_stage_prompted(context, "workplace_status"):
            await update.message.reply_text("Rasmlar qabul qilindi (2/2).")
            await update.message.reply_text(
                "Terminallar va ratsiyalar quvvatini tekshiring va ularning quvvatlanish jarayonini rasmga oling."
            )
            self._mark_stage_prompted(context, "workplace_status")

        if self._lock_opening_stage_for_media_group(update, context, "workplace_status", "terminal_power"):
            return UPLOAD_WORKPLACE_STATUS

        self._block_current_media_group(update, context)
        self._clear_stage_prompted(context, "workplace_status")
        context.user_data['opening_stage'] = 'terminal_power'
        return UPLOAD_TERMINAL_POWER

    async def upload_terminal_power(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Upload terminal power image."""
        self._sync_opening_stage_with_media_group(update, context)
        current_stage = context.user_data.get("opening_stage")
        if current_stage and current_stage != "terminal_power":
            if current_stage == "workplace_status":
                return await self.upload_workplace_status(update, context)
            if current_stage == "zero_report":
                return await self.upload_zero_report(update, context)
            if current_stage == "opening_notification":
                return await self.upload_opening_notification(update, context)
            if current_stage == "receipt_roll":
                return await self.upload_receipt_roll(update, context)
        if self._is_blocked_media_group(update, context):
            return UPLOAD_TERMINAL_POWER

        file_id = self._get_image_file_id(update)
        if not file_id:
            await update.message.reply_text("Iltimos, rasm yuboring.")
            context.user_data['opening_stage'] = 'terminal_power'
            return UPLOAD_TERMINAL_POWER

        shift_id = context.user_data.get('current_shift_id')
        if shift_id:
            await self._save_shift_image(shift_id, 'terminal_power', file_id)
            self._queue_opening_group_photo(
                context,
                file_id,
                "Terminal/ratsiya quvvat holati",
                event_time=getattr(update.message, "date", None),
                media_kind=self._get_image_media_kind(update),
            )

        if not self._is_stage_prompted(context, "terminal_power"):
            await update.message.reply_text("Rasm qabul qilindi.")
            await update.message.reply_text("Uzcard va Humo kartalaridagi nol hisobotni chiqaring va rasmga oling.")
            self._mark_stage_prompted(context, "terminal_power")
        if self._lock_opening_stage_for_media_group(update, context, "terminal_power", "zero_report"):
            return UPLOAD_TERMINAL_POWER
        self._block_current_media_group(update, context)
        self._clear_stage_prompted(context, "terminal_power")
        context.user_data['opening_stage'] = 'zero_report'
        return UPLOAD_ZERO_REPORT

    async def upload_zero_report(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Upload zero report image."""
        self._sync_opening_stage_with_media_group(update, context)
        current_stage = context.user_data.get("opening_stage")
        if current_stage and current_stage != "zero_report":
            if current_stage == "workplace_status":
                return await self.upload_workplace_status(update, context)
            if current_stage == "terminal_power":
                return await self.upload_terminal_power(update, context)
            if current_stage == "opening_notification":
                return await self.upload_opening_notification(update, context)
            if current_stage == "receipt_roll":
                return await self.upload_receipt_roll(update, context)
        if self._is_blocked_media_group(update, context):
            return UPLOAD_ZERO_REPORT

        file_id = self._get_image_file_id(update)
        if not file_id:
            await update.message.reply_text("Iltimos, rasm yuboring.")
            context.user_data['opening_stage'] = 'zero_report'
            return UPLOAD_ZERO_REPORT

        shift_id = context.user_data.get('current_shift_id')
        if shift_id:
            await self._save_shift_image(shift_id, 'zero_report', file_id)
            self._queue_opening_group_photo(
                context,
                file_id,
                "Uzcard/Humo nol hisobot",
                event_time=getattr(update.message, "date", None),
                media_kind=self._get_image_media_kind(update),
            )

        if not self._is_stage_prompted(context, "zero_report"):
            await update.message.reply_text("Rasm qabul qilindi.")
            await update.message.reply_text("Iiko va soliq check tizimlarida smenani oching. Ochilganlik haqidagi bildirishnomani rasmga oling.")
            self._mark_stage_prompted(context, "zero_report")
        if self._lock_opening_stage_for_media_group(update, context, "zero_report", "opening_notification"):
            return UPLOAD_ZERO_REPORT
        self._block_current_media_group(update, context)
        self._clear_stage_prompted(context, "zero_report")
        context.user_data['opening_stage'] = 'opening_notification'
        return UPLOAD_OPENING_NOTIFICATION

    async def upload_opening_notification(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Upload opening notification image."""
        self._sync_opening_stage_with_media_group(update, context)
        current_stage = context.user_data.get("opening_stage")
        if current_stage and current_stage != "opening_notification":
            if current_stage == "workplace_status":
                return await self.upload_workplace_status(update, context)
            if current_stage == "terminal_power":
                return await self.upload_terminal_power(update, context)
            if current_stage == "zero_report":
                return await self.upload_zero_report(update, context)
            if current_stage == "receipt_roll":
                return await self.upload_receipt_roll(update, context)
        if self._is_blocked_media_group(update, context):
            return UPLOAD_OPENING_NOTIFICATION

        file_id = self._get_image_file_id(update)
        if not file_id:
            await update.message.reply_text("Iltimos, rasm yuboring.")
            context.user_data['opening_stage'] = 'opening_notification'
            return UPLOAD_OPENING_NOTIFICATION

        shift_id = context.user_data.get('current_shift_id')
        if shift_id:
            await self._save_shift_image(shift_id, 'opening_notification', file_id)
            self._queue_opening_group_photo(
                context,
                file_id,
                "iiko/soliq ochilish bildirishnomasi",
                event_time=getattr(update.message, "date", None),
                media_kind=self._get_image_media_kind(update),
            )

        if not self._is_stage_prompted(context, "opening_notification"):
            await update.message.reply_text("Rasm qabul qilindi.")
            await update.message.reply_text("Zaxira chek lentalari mavjudligini rasm bilan jo'nating.")
            self._mark_stage_prompted(context, "opening_notification")
        if self._lock_opening_stage_for_media_group(update, context, "opening_notification", "receipt_roll"):
            return UPLOAD_OPENING_NOTIFICATION
        self._block_current_media_group(update, context)
        self._clear_stage_prompted(context, "opening_notification")
        context.user_data['opening_stage'] = 'receipt_roll'
        return UPLOAD_RECEIPT_ROLL

    async def upload_receipt_roll(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Upload receipt roll image and finish shift opening flow."""
        self._sync_opening_stage_with_media_group(update, context)
        if self._is_blocked_media_group(update, context):
            return UPLOAD_RECEIPT_ROLL

        file_id = self._get_image_file_id(update)
        if not file_id:
            await update.message.reply_text("Iltimos, rasm yuboring.")
            context.user_data['opening_stage'] = 'receipt_roll'
            return UPLOAD_RECEIPT_ROLL

        shift_id = context.user_data.get('current_shift_id')
        if shift_id:
            await self._save_shift_image(shift_id, 'receipt_roll', file_id)
            self._queue_opening_group_photo(
                context,
                file_id,
                "Zaxira chek lenta rasmi",
                event_time=getattr(update.message, "date", None),
                media_kind=self._get_image_media_kind(update),
            )

        await update.message.reply_text("Rasm qabul qilindi.")
        media_group_id = getattr(update.message, "media_group_id", None)
        if media_group_id:
            context.user_data['opening_stage'] = 'receipt_roll'
            self._schedule_receipt_roll_finalize(
                context,
                chat_id=update.effective_chat.id,
                cashier_first_name=update.effective_user.first_name,
                cashier_last_name=update.effective_user.last_name,
            )
            return UPLOAD_RECEIPT_ROLL

        self._block_current_media_group(update, context)
        await self._finalize_shift_opening_flow(
            context,
            chat_id=update.effective_chat.id,
            cashier_first_name=update.effective_user.first_name,
            cashier_last_name=update.effective_user.last_name,
        )
        return MAIN_MENU

    async def start_daily_reporting(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start the daily reporting process"""
        if not await self._ensure_cashier_authenticated(update, context, resume_action="start_daily_reporting"):
            return MAIN_MENU
        self._clear_debt_received_detail_state(context)
        self._clear_debt_payments_detail_state(context)
        self._clear_expense_detail_state(context)
        self._clear_generic_payment_method_state(context)

        # Check if there's an active shift
        user_row = await self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
        if not user_row:
            await update.message.reply_text("Foydalanuvchi topilmadi.")
            return MAIN_MENU
        active_shift = await self.db.fetch_one("SELECT * FROM shifts WHERE user_id=%s AND is_open=TRUE", (user_row['id'],))
        
        if not active_shift:
            lang = 'uz'
            
            if lang == 'uz':
                msg = "Avval smena ochishingiz kerak."
            else:
                msg = "Р В Р’В Р В Р вЂ№Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р’В Р вЂ™Р’В° Р В Р’В Р В РІР‚В¦Р В Р Р‹Р РЋРІР‚СљР В Р’В Р вЂ™Р’В¶Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СњР В Р Р‹Р В РІР‚С™Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ° Р В Р Р‹Р В РЎвЂњР В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р Р‹Р РЋРІР‚Сљ."
            
            await update.message.reply_text(msg)
            return MAIN_MENU
        
        context.user_data['current_shift_id'] = active_shift['id']
        if not await self._ensure_opening_requirements_completed(update, context, int(active_shift['id'])):
            return MAIN_MENU

        return await self._start_sverka_flow(
            update,
            context,
            int(active_shift['id']),
            entrypoint="standalone",
        )

    async def report_sales(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get sales amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data['sales_amount'] = amount
            context.user_data.pop("sales_amount_detail_stage", None)
            context.user_data.pop("sales_amount_payment_type", None)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            self._mark_sverka_done(context, 'sales_amount')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_SALES

    async def report_debt_received(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Collect received debts as multiple entries."""
        stage = context.user_data.get("debt_received_detail_stage")
        if stage == "counterparty_name":
            name = (update.message.text or "").strip()
            if not name:
                await update.message.reply_text("Kimdan kelganini kiriting (ism).")
                return REPORT_DEBT_RECEIVED
            context.user_data["debt_received_current_name"] = name
            context.user_data["debt_received_detail_stage"] = "counterparty_phone"
            await update.message.reply_text("Telefon raqamini kiriting.")
            return REPORT_DEBT_RECEIVED
        if stage == "counterparty_phone":
            phone = (update.message.text or "").strip()
            if not validate_phone_number(phone):
                await update.message.reply_text("Iltimos, to'g'ri telefon raqamini kiriting.")
                return REPORT_DEBT_RECEIVED
            context.user_data["debt_received_current_phone"] = phone
            context.user_data["debt_received_detail_stage"] = "payment_type"
            await update.message.reply_text(
                "Kelgan qarz uchun to'lov turini tanlang.",
                reply_markup=self._build_expense_payment_type_keyboard(),
            )
            return REPORT_DEBT_RECEIVED
        if stage == "payment_type":
            payment_type = self._normalize_payment_type(update.message.text)
            if not payment_type:
                await update.message.reply_text(
                    "Iltimos, to'lov turini tugmalardan tanlang.",
                    reply_markup=self._build_expense_payment_type_keyboard(),
                )
                return REPORT_DEBT_RECEIVED
            item = {
                "counterparty_name": context.user_data.get("debt_received_current_name", ""),
                "counterparty_phone": context.user_data.get("debt_received_current_phone", ""),
                "amount": context.user_data.get("debt_received_current_amount", 0),
                "payment_type": payment_type,
            }
            self._debt_received_items(context).append(item)
            context.user_data.pop("debt_received_current_name", None)
            context.user_data.pop("debt_received_current_phone", None)
            context.user_data.pop("debt_received_current_amount", None)
            context.user_data["debt_received"] = self._debt_received_items_total(context)
            context.user_data["debt_received_detail_stage"] = "loop"

            items = self._debt_received_items(context)
            summary_lines = ["To'lov turi qabul qilindi.\n", "Kiritilgan kelgan qarzlar:"]
            for idx, it in enumerate(items, 1):
                summary_lines.append(
                    f"  {idx}. {it.get('counterparty_name', '-')} — "
                    f"{self._fmt_money(it.get('amount', 0))} — {it.get('payment_type', '-')}"
                )
            summary_lines.append(f"\nJami: {self._fmt_money(self._debt_received_items_total(context))}")
            await update.message.reply_text(
                "\n".join(summary_lines),
                reply_markup=self._build_debt_received_loop_keyboard(),
            )
            return REPORT_DEBT_RECEIVED

        if stage == "loop":
            action = self._normalize_debt_received_action(update.message.text)
            if action == "add_more":
                context.user_data["debt_received_detail_stage"] = None
                await update.message.reply_text(
                    "Yangi kelgan qarz summasini kiriting:",
                    reply_markup=ReplyKeyboardRemove(),
                )
                return REPORT_DEBT_RECEIVED
            if action == "finish":
                context.user_data.pop("debt_received_detail_stage", None)
                context.user_data.pop('pending_sverka_key', None)
                context.user_data.pop('pending_sverka_state', None)
                self._mark_sverka_done(context, 'debt_received')
                await update.message.reply_text(
                    "Kelgan qarzlar ma'lumotlari saqlandi.",
                    reply_markup=ReplyKeyboardRemove(),
                )
                return await self._after_sverka_step(update, context)
            await update.message.reply_text(
                "Iltimos, tugmalardan birini tanlang.",
                reply_markup=self._build_debt_received_loop_keyboard(),
            )
            return REPORT_DEBT_RECEIVED

        try:
            amount = self._parse_amount(update.message.text)
            if amount <= 0:
                self._clear_debt_received_detail_state(context)
                context.user_data['debt_received'] = 0
                context.user_data.pop('pending_sverka_key', None)
                context.user_data.pop('pending_sverka_state', None)
                self._mark_sverka_done(context, 'debt_received')
                return await self._after_sverka_step(update, context)

            context.user_data["debt_received_current_amount"] = amount
            context.user_data["debt_received_detail_stage"] = "counterparty_name"
            await update.message.reply_text("Kelgan qarz kimdan keldi? Ismini kiriting.")
            return REPORT_DEBT_RECEIVED

        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_DEBT_RECEIVED

    async def report_expenses(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Collect expense lines and calculate the total automatically."""
        if context.user_data.get("expense_detail_stage") == "cash_amount":
            try:
                cash_amount = self._parse_amount(update.message.text)
            except ValueError:
                await update.message.reply_text(self._invalid_amount_msg(context))
                return REPORT_EXPENSES

            context.user_data["expense_cash_amount"] = cash_amount
            total = self._expense_items_total(context)
            context.user_data["expenses"] = total
            context.user_data.pop("expense_detail_stage", None)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            self._mark_sverka_done(context, 'expenses')
            await update.message.reply_text("Chiqimlar va naqd summa qabul qilindi.", reply_markup=ReplyKeyboardRemove())
            return await self._after_sverka_step(update, context)

        context.user_data["expense_detail_stage"] = "items"
        text = (update.message.text or "").strip()
        action = self._normalize_expense_action(text)

        if action == "add_more":
            if not self._expense_items(context):
                await update.message.reply_text(
                    self._build_expense_entry_prompt(context, note="Avval kamida bitta chiqim kiriting."),
                    reply_markup=self._build_expense_entry_keyboard(),
                )
                return REPORT_EXPENSES
            await update.message.reply_text(
                self._build_expense_entry_prompt(context, note="Keyingi chiqimni kiriting."),
                reply_markup=self._build_expense_entry_keyboard(),
            )
            return REPORT_EXPENSES

        if action == "finish":
            total = self._expense_items_total(context)
            context.user_data["expense_detail_stage"] = "cash_amount"
            await update.message.reply_text(
                (
                    f"Chiqimlar qabul qilindi. Jami chiqim: {self._fmt_money(total)}.\n"
                    "Naqd summani kiriting:"
                ) if total > 0 else "Chiqimlar kiritilmadi. Naqd summani kiriting:",
                reply_markup=ReplyKeyboardRemove(),
            )
            return REPORT_EXPENSES

        try:
            item = self._parse_expense_item(text)
        except ValueError:
            await update.message.reply_text(
                self._build_expense_entry_prompt(
                    context,
                    note="Chiqimni misoldagidek kiriting. Masalan: Mirshod Dastafka -- 10 000",
                ),
                reply_markup=self._build_expense_entry_keyboard(),
            )
            return REPORT_EXPENSES

        self._expense_items(context).append(item)
        await update.message.reply_text(
            self._build_expense_entry_prompt(context, note="Chiqim qo'shildi."),
            reply_markup=self._build_expense_entry_keyboard(),
        )
        return REPORT_EXPENSES

    async def report_uzcard(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get Uzcard amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['uzcard_amount'] = amount
            self._mark_sverka_done(context, 'uzcard_amount')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_UZCARD

    async def report_humo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get Humo amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['humo_amount'] = amount
            self._mark_sverka_done(context, 'humo_amount')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_HUMO

    async def report_p2p(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get P2P amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['p2p_amount'] = amount
            self._mark_sverka_done(context, 'p2p_amount')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_P2P

    async def report_uzcard_refund(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get Uzcard refund amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['uzcard_refund'] = amount
            self._mark_sverka_done(context, 'uzcard_refund')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_UZCARD_REFUND

    async def report_humo_refund(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get Humo refund amount"""
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            context.user_data['humo_refund'] = amount
            self._mark_sverka_done(context, 'humo_refund')
            return await self._after_sverka_step(update, context)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_HUMO_REFUND

    async def report_other_payments(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get other payments amount"""
        comment = (update.message.text or "").strip()
        if not comment:
            await update.message.reply_text("Boshqa to'lovlar bo'yicha izohni kiriting.")
            return REPORT_OTHER_PAYMENTS
        self._clear_generic_payment_method_state(context)
        context.user_data['other_payments'] = 0
        context.user_data["other_payments_comment"] = comment
        context.user_data.pop('pending_sverka_key', None)
        context.user_data.pop('pending_sverka_state', None)
        self._mark_sverka_done(context, 'other_payments')
        return await self._after_sverka_step(update, context)

    async def report_tax_info(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Collect closing tax check image, then cash amount."""
        stage = context.user_data.get("tax_info_stage") or "check_image"
        if stage == "check_image":
            file_id = self._get_image_file_id(update)
            if not file_id:
                await update.message.reply_text("Soliq cheki rasmini yuboring (foto yoki image fayl).")
                return REPORT_TAX_INFO

            context.user_data["tax_info_check_image"] = file_id
            context.user_data["tax_info_stage"] = "cash_amount"

            shift_id = context.user_data.get("current_shift_id")
            if shift_id:
                await self._save_shift_image(int(shift_id), "tax_info_check", file_id)

            await update.message.reply_text("Chek rasmi qabul qilindi. Naqd summani kiriting:")
            return REPORT_TAX_INFO

        if stage == "cash_amount":
            try:
                amount = self._parse_amount(update.message.text)
            except ValueError:
                await update.message.reply_text(self._invalid_amount_msg(context))
                return REPORT_TAX_INFO

            context.user_data["tax_info_cash_amount"] = amount
            context.user_data.pop("tax_info_stage", None)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            self._mark_sverka_done(context, 'tax_info')
            await update.message.reply_text("Soliq ma'lumotlari qabul qilindi.", reply_markup=ReplyKeyboardRemove())
            return await self._after_sverka_step(update, context)

        self._clear_tax_info_state(context)
        await update.message.reply_text("Soliq ma'lumotlarini qaytadan kiriting. Avval chek rasmini yuboring.")
        return REPORT_TAX_INFO

    async def report_debt_payments(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Qarz berish: Summa → Ism → Telefon → Tasdiqlash → Chek rasmi → Loop"""
        stage = context.user_data.get("debt_payments_detail_stage")
        has_photo = bool(getattr(update.message, "photo", None)) if update.message else False
        has_doc = bool(getattr(update.message, "document", None)) if update.message else False
        text = (update.message.text or "").strip() if update.message and update.message.text else ""
        logger.info(
            "report_debt_payments called: stage=%s, has_photo=%s, has_doc=%s, text=%r",
            stage, has_photo, has_doc, text[:50] if text else "",
        )

        # --- Stage: counterparty_name ---
        if stage == "counterparty_name":
            if not text:
                await update.message.reply_text("Qarzdor ismini kiriting.")
                return REPORT_DEBT_PAYMENTS
            context.user_data["debt_payments_current_name"] = text
            context.user_data["debt_payments_detail_stage"] = "counterparty_phone"
            await update.message.reply_text("Telefon raqamini kiriting.")
            return REPORT_DEBT_PAYMENTS

        # --- Stage: counterparty_phone ---
        if stage == "counterparty_phone":
            if not validate_phone_number(text):
                await update.message.reply_text("Iltimos, to'g'ri telefon raqamini kiriting.")
                return REPORT_DEBT_PAYMENTS
            context.user_data["debt_payments_current_phone"] = text
            context.user_data["debt_payments_detail_stage"] = "confirm"
            await update.message.reply_text(
                self._build_debt_payments_confirm_text(context),
                reply_markup=self._build_debt_payments_confirm_keyboard(),
            )
            return REPORT_DEBT_PAYMENTS

        # --- Stage: confirm ---
        if stage == "confirm":
            action = self._normalize_debt_payments_confirm(text)
            if action == "cancel":
                context.user_data.pop("debt_payments_current_name", None)
                context.user_data.pop("debt_payments_current_phone", None)
                context.user_data.pop("debt_payments_current_amount", None)
                context.user_data.pop("debt_payments_detail_stage", None)
                await update.message.reply_text(
                    "Bekor qilindi. Summani qayta kiriting yoki 0 kiriting.",
                    reply_markup=ReplyKeyboardRemove(),
                )
                return REPORT_DEBT_PAYMENTS
            if action == "confirm":
                context.user_data["debt_payments_detail_stage"] = "check_image"
                await update.message.reply_text(
                    "📸 To'lov/qarz cheki rasmini yuboring.",
                    reply_markup=ReplyKeyboardRemove(),
                )
                return REPORT_DEBT_PAYMENTS
            await update.message.reply_text(
                "Iltimos, tugmalardan birini tanlang.",
                reply_markup=self._build_debt_payments_confirm_keyboard(),
            )
            return REPORT_DEBT_PAYMENTS

        # --- Stage: check_image (photo handler) ---
        if stage == "check_image":
            photo = getattr(update.message, "photo", None)
            document = getattr(update.message, "document", None) if not photo else None
            if not photo and not document:
                await update.message.reply_text("Iltimos, chek rasmini yuboring (rasm yoki fayl).")
                return REPORT_DEBT_PAYMENTS
            if photo:
                file_id = photo[-1].file_id
            else:
                file_id = document.file_id
            item = {
                "counterparty_name": context.user_data.get("debt_payments_current_name", ""),
                "counterparty_phone": context.user_data.get("debt_payments_current_phone", ""),
                "amount": context.user_data.get("debt_payments_current_amount", 0),
                "check_image": file_id,
            }
            self._debt_payments_items(context).append(item)
            context.user_data.pop("debt_payments_current_name", None)
            context.user_data.pop("debt_payments_current_phone", None)
            context.user_data.pop("debt_payments_current_amount", None)
            context.user_data["debt_payments"] = self._debt_payments_items_total(context)
            context.user_data["debt_payments_detail_stage"] = "loop"
            items = self._debt_payments_items(context)
            summary_lines = ["✅ Chek rasmi qabul qilindi!\n", "Kiritilgan qarzlar:"]
            for idx, it in enumerate(items, 1):
                summary_lines.append(
                    f"  {idx}. {it.get('counterparty_name', '-')} — "
                    f"{self._fmt_money(it.get('amount', 0))}"
                )
            summary_lines.append(f"\nJami: {self._fmt_money(self._debt_payments_items_total(context))}")
            await update.message.reply_text(
                "\n".join(summary_lines),
                reply_markup=self._build_debt_payments_loop_keyboard(),
            )
            return REPORT_DEBT_PAYMENTS

        # --- Stage: loop ---
        if stage == "loop":
            action = self._normalize_debt_payments_action(text)
            if action == "add_more":
                context.user_data["debt_payments_detail_stage"] = None
                await update.message.reply_text(
                    "Yangi qarzdor uchun summani kiriting:",
                    reply_markup=ReplyKeyboardRemove(),
                )
                return REPORT_DEBT_PAYMENTS
            if action == "finish":
                context.user_data.pop("debt_payments_detail_stage", None)
                context.user_data.pop("pending_sverka_key", None)
                context.user_data.pop("pending_sverka_state", None)
                self._mark_sverka_done(context, "debt_payments")
                await update.message.reply_text(
                    "Qarz berish ma'lumotlari saqlandi.",
                    reply_markup=ReplyKeyboardRemove(),
                )
                return await self._after_sverka_step(update, context)
            await update.message.reply_text(
                "Iltimos, tugmalardan birini tanlang.",
                reply_markup=self._build_debt_payments_loop_keyboard(),
            )
            return REPORT_DEBT_PAYMENTS

        # --- Initial stage: amount entry ---
        try:
            amount = self._parse_amount(text)
            if amount <= 0:
                self._clear_debt_payments_detail_state(context)
                context.user_data["debt_payments"] = 0
                context.user_data.pop("pending_sverka_key", None)
                context.user_data.pop("pending_sverka_state", None)
                self._mark_sverka_done(context, "debt_payments")
                return await self._after_sverka_step(update, context)
            context.user_data["debt_payments_current_amount"] = amount
            context.user_data["debt_payments_detail_stage"] = "counterparty_name"
            await update.message.reply_text("Qarzdor ismini kiriting.")
            return REPORT_DEBT_PAYMENTS
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_DEBT_PAYMENTS

    async def report_debt_refunds(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Get debt refunds amount"""
        if context.user_data.get("debt_refunds_detail_stage") == "payment_type":
            payment_type = self._normalize_payment_type(update.message.text)
            if not payment_type:
                await update.message.reply_text(
                    "Iltimos, to'lov turini tugmalardan tanlang.",
                    reply_markup=self._build_expense_payment_type_keyboard(),
                )
                return REPORT_DEBT_REFUNDS
            context.user_data["debt_refunds_payment_type"] = payment_type
            context.user_data.pop("debt_refunds_detail_stage", None)
            context.user_data.pop('pending_sverka_key', None)
            context.user_data.pop('pending_sverka_state', None)
            self._mark_sverka_done(context, 'debt_refunds')
            await update.message.reply_text("To'lov turi qabul qilindi.", reply_markup=ReplyKeyboardRemove())
            return await self._after_sverka_step(update, context)
        try:
            amount = self._parse_amount(update.message.text)
            context.user_data['debt_refunds'] = amount
            if amount <= 0:
                context.user_data.pop("debt_refunds_detail_stage", None)
                context.user_data.pop("debt_refunds_payment_type", None)
                context.user_data.pop('pending_sverka_key', None)
                context.user_data.pop('pending_sverka_state', None)
                self._mark_sverka_done(context, 'debt_refunds')
                return await self._after_sverka_step(update, context)
            context.user_data.pop("debt_refunds_payment_type", None)
            context.user_data["debt_refunds_detail_stage"] = "payment_type"
            await update.message.reply_text(
                "Vozvrat qarzlar uchun to'lov turini tanlang.",
                reply_markup=self._build_expense_payment_type_keyboard(),
            )
            return REPORT_DEBT_REFUNDS
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return REPORT_DEBT_REFUNDS
    async def save_daily_report(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Save the daily report to the database"""
        shift_id = context.user_data['current_shift_id']
        report_data_json = json.dumps(self._build_report_data_payload(context), ensure_ascii=False)
        
        report_data = {
            'shift_id': shift_id,
            'report_type': 'daily_report',
            'sales_amount': context.user_data.get('sales_amount', 0),
            'debt_received': context.user_data.get('debt_received', 0),
            'expenses': context.user_data.get('expenses', 0),
            'uzcard_amount': context.user_data.get('uzcard_amount', 0),
            'humo_amount': context.user_data.get('humo_amount', 0),
            'p2p_amount': context.user_data.get('p2p_amount', 0),
            'uzcard_refund': context.user_data.get('uzcard_refund', 0),
            'humo_refund': context.user_data.get('humo_refund', 0),
            'other_payments': context.user_data.get('other_payments', 0),
            'debt_payments': context.user_data.get('debt_payments', 0),
            'debt_refunds': context.user_data.get('debt_refunds', 0),
            'report_data': report_data_json,
        }
        
        existing_report = await self.db.fetch_one(
            """
            SELECT id
            FROM reports
            WHERE shift_id = %s AND report_type = 'daily_report'
            ORDER BY id DESC
            LIMIT 1
            """,
            (shift_id,),
        )
        if existing_report:
            report_data['id'] = existing_report['id']
            query = """
                UPDATE reports
                SET
                    sales_amount = %(sales_amount)s,
                    debt_received = %(debt_received)s,
                    expenses = %(expenses)s,
                    uzcard_amount = %(uzcard_amount)s,
                    humo_amount = %(humo_amount)s,
                    p2p_amount = %(p2p_amount)s,
                    uzcard_refund = %(uzcard_refund)s,
                    humo_refund = %(humo_refund)s,
                    other_payments = %(other_payments)s,
                    debt_payments = %(debt_payments)s,
                    debt_refunds = %(debt_refunds)s,
                    report_data = %(report_data)s::jsonb
                WHERE id = %(id)s
            """
        else:
            query = """
                INSERT INTO reports (
                    shift_id, report_type, sales_amount, debt_received, expenses,
                    uzcard_amount, humo_amount, p2p_amount, uzcard_refund, humo_refund,
                    other_payments, debt_payments, debt_refunds, report_data
                ) VALUES (
                    %(shift_id)s, %(report_type)s, %(sales_amount)s, %(debt_received)s, %(expenses)s,
                    %(uzcard_amount)s, %(humo_amount)s, %(p2p_amount)s, %(uzcard_refund)s, %(humo_refund)s,
                    %(other_payments)s, %(debt_payments)s, %(debt_refunds)s, %(report_data)s::jsonb
                )
            """
        await self.db.execute_query(query, report_data)
        # Muhim: guruhga bitta yakuniy fayl faqat smena yopilganda yuboriladi.
        # Shu sababli bu yerda alohida sverka fayl yubormaymiz.

    async def _finalize_sverka(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        await self.save_daily_report(update, context)

        shift_id = context.user_data.get("current_shift_id")
        if shift_id:
            try:
                shift_summary = await self._get_shift_summary(shift_id) or {}
                await self._send_debt_payment_check_images(context, shift_summary)
                entrypoint = context.user_data.get("sverka_entrypoint")
                if entrypoint == "closing":
                    await self._send_tax_info_check_image(context, shift_summary)
                    summary_text = self._build_sverka_summary_message(
                        shift_summary,
                        title="🔒 Kassa yopilishi ma'lumotlari",
                        closing=True,
                    )
                else:
                    summary_text = self._build_sverka_summary_message(shift_summary)
                await self._send_group_message(context, summary_text)
            except Exception:
                logger.exception("sverka group send failed")

        entrypoint = context.user_data.get("sverka_entrypoint")
        self._clear_sverka_value_state(context)
        self._clear_sverka_flow_state(context)

        if entrypoint == "closing":
            return await self._prompt_close_shift_amount(
                update,
                context,
                text="Kassa yopilishi ma'lumotlari guruhga yuborildi. Endi smenani yopish uchun yakuniy summani kiriting:",
            )

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="Sverka yakunlandi! Barcha hisobotlar saqlandi.",
            reply_markup=ReplyKeyboardRemove(),
        )
        await self.show_cashier_menu(update, context)
        context.user_data['flow'] = None
        return MAIN_MENU

    async def start_shift_closing(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start the shift closing process"""
        if not await self._ensure_cashier_authenticated(update, context, resume_action="start_shift_closing"):
            return MAIN_MENU

        lang = 'uz'
        user_row = await self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
        if not user_row:
            await update.message.reply_text("Foydalanuvchi topilmadi.")
            return MAIN_MENU
        active_shift = await self.db.fetch_one(
            "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC LIMIT 1",
            (user_row['id'],)
        )
        if not active_shift:
            if lang == 'uz':
                msg = "Ochiq smena yo'q."
            else:
                msg = "Р В РЎСљР В Р’ВµР РЋРІР‚С™ Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂќР РЋР вЂљР РЋРІР‚в„–Р РЋРІР‚С™Р В РЎвЂўР В РІвЂћвЂ“ Р РЋР С“Р В РЎВР В Р’ВµР В Р вЂ¦Р РЋРІР‚в„–."
            await update.message.reply_text(msg)
            return MAIN_MENU
        if not await self._ensure_opening_requirements_completed(update, context, int(active_shift['id'])):
            return MAIN_MENU
        context.user_data['current_shift_id'] = active_shift['id']

        return await self._start_sverka_flow(
            update,
            context,
            int(active_shift['id']),
            entrypoint="closing",
            force_reset=True,
            note="Smenani yopishdan oldin yakuniy sverkani to'ldiring.",
        )

    async def start_payment_image_upload(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Ask cashier to choose Uzcard or Humo and then upload image"""
        if not await self._ensure_cashier_authenticated(update, context, resume_action="start_payment_image_upload"):
            return MAIN_MENU
        context.user_data.pop('blocked_media_group_id', None)

        lang = 'uz'
        user_row = await self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
        if not user_row:
            await update.message.reply_text("Foydalanuvchi topilmadi.")
            return MAIN_MENU

        active_shift = await self.db.fetch_one(
            "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE",
            (user_row['id'],)
        )
        if not active_shift:
            await update.message.reply_text("Ochiq smena yo'q.")
            return MAIN_MENU

        context.user_data['current_shift_id'] = active_shift['id']
        context.user_data['flow'] = 'payment_image'

        text = "Qaysi turdagi rasm yuborasiz?" if lang == 'uz' else "Qaysi turdagi rasm yuborasiz?"
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("💳 Uzcard", callback_data="payimg:uzcard")],
            [InlineKeyboardButton("💳 Humo", callback_data="payimg:humo")],
            [InlineKeyboardButton("⬅️ Orqaga", callback_data="payimg:back")],
        ])
        await update.message.reply_text(text, reply_markup=keyboard)
        return SELECT_PAYMENT_IMAGE

    async def select_payment_image_type(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
        key = (query.data or '').split(':', 1)[1] if query.data else ''
        if key == 'back':
            await self.show_cashier_menu(update, context)
            context.user_data['flow'] = None
            return MAIN_MENU
        if key not in ['uzcard', 'humo']:
            await self.start_payment_image_upload(update, context)
            return SELECT_PAYMENT_IMAGE

        context.user_data['pending_payment_image'] = key
        msg = "Rasmni yuboring:" if 'uz' == 'uz' else "Rasmni yuboring:"
        await context.bot.send_message(chat_id=query.message.chat_id, text=msg)
        return UPLOAD_PAYMENT_IMAGE

    async def upload_payment_image(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        try:
            if self._is_blocked_media_group(update, context):
                return UPLOAD_PAYMENT_IMAGE

            file_id = self._get_image_file_id(update)
            if not file_id:
                await update.message.reply_text("Iltimos, rasm yuboring (foto yoki image fayl).")
                return UPLOAD_PAYMENT_IMAGE

            key = context.user_data.get('pending_payment_image')
            shift_id = context.user_data.get('current_shift_id')
            if not shift_id:
                user_row = await self.db.fetch_one("SELECT id FROM users WHERE telegram_id=%s", (update.effective_user.id,))
                if user_row:
                    active_shift = await self.db.fetch_one(
                        "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC LIMIT 1",
                        (user_row['id'],)
                    )
                    if active_shift:
                        shift_id = active_shift['id']
                        context.user_data['current_shift_id'] = shift_id

            if key not in ('uzcard', 'humo') or not shift_id:
                await update.message.reply_text("Avval `Rasm jo'natish` tugmasini bosib, Uzcard yoki Humo ni tanlang.")
                context.user_data.pop('pending_payment_image', None)
                context.user_data['flow'] = None
                return MAIN_MENU

            image_type = 'uzcard_payment' if key == 'uzcard' else 'humo_payment'
            await self.db.execute_query(
                "INSERT INTO images (shift_id, image_url, image_type) VALUES (%s, %s, %s)",
                (shift_id, file_id, image_type)
            )
            if key == 'uzcard':
                await self._send_group_shift_photo(
                    context, shift_id, file_id, "Uzcard hisobot rasmi", event_time=getattr(update.message, "date", None)
                )
                shift_meta = await self._get_shift_meta(shift_id)
                await self._send_group_message(
                    context,
                    self._build_payment_image_uploaded_message(
                        "Uzcard hisobot rasmi",
                        shift_meta,
                        getattr(update.message, "date", None),
                    ),
                )
            else:
                await self._send_group_shift_photo(
                    context, shift_id, file_id, "Humo hisobot rasmi", event_time=getattr(update.message, "date", None)
                )
                shift_meta = await self._get_shift_meta(shift_id)
                await self._send_group_message(
                    context,
                    self._build_payment_image_uploaded_message(
                        "Humo hisobot rasmi",
                        shift_meta,
                        getattr(update.message, "date", None),
                    ),
                )

            context.user_data.pop('pending_payment_image', None)
            self._block_current_media_group(update, context)
            if key == 'uzcard':
                await update.message.reply_text("Uzcard hisobot rasmingiz qabul qilindi.")
            else:
                await update.message.reply_text("Humo hisobot rasmingiz qabul qilindi.")

            uzcard_img = await self._count_shift_images(shift_id, 'uzcard_payment')
            humo_img = await self._count_shift_images(shift_id, 'humo_payment')
            if uzcard_img < 1 or humo_img < 1:
                missing_key = "uzcard" if uzcard_img < 1 else "humo"
                missing_label = "Uzcard" if missing_key == "uzcard" else "Humo"
                context.user_data['pending_payment_image'] = missing_key
                context.user_data['flow'] = 'payment_image'
                await update.message.reply_text(f"{missing_label} rasmini ham yuboring.")
                return UPLOAD_PAYMENT_IMAGE

            await update.message.reply_text("Uzcard va Humo rasmlari to'liq qabul qilindi.")
            if context.user_data.pop("awaiting_payment_images_for_close", False):
                return await self._start_sverka_flow(
                    update,
                    context,
                    int(shift_id),
                    entrypoint="closing",
                    force_reset=True,
                    note="Rasmlar qabul qilindi. Endi yakuniy sverkani to'ldiring.",
                )

            context.user_data['flow'] = None
            await self.show_cashier_menu(update, context)
            return MAIN_MENU
        except Exception:
            logger.exception("upload_payment_image failed")
            context.user_data.pop('pending_payment_image', None)
            context.user_data.pop("awaiting_payment_images_for_close", None)
            context.user_data['flow'] = None
            await update.message.reply_text("Rasmni saqlashda xatolik bo'ldi. Qayta urinib ko'ring.")
            await self.show_cashier_menu(update, context)
            return MAIN_MENU
    async def edit_reports(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle report editing"""
        context.user_data['flow'] = 'edit'
        await self.show_edit_reports_menu(update, context)
        return EDIT_REPORT_SELECT

    async def show_edit_reports_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        lang = 'uz'
        user_row = await self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
        if not user_row:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="Foydalanuvchi topilmadi.")
            context.user_data['flow'] = None
            return MAIN_MENU

        # Prefer current shift; fallback to latest shift
        shift_id = context.user_data.get('current_shift_id')
        if not shift_id:
            last_shift = await self.db.fetch_one(
                "SELECT id FROM shifts WHERE user_id=%s ORDER BY id DESC LIMIT 1",
                (user_row['id'],)
            )
            shift_id = last_shift['id'] if last_shift else None

        report_row = None
        if shift_id:
            report_row = await self.db.fetch_one(
                "SELECT * FROM reports WHERE shift_id=%s AND report_type='daily_report' ORDER BY id DESC LIMIT 1",
                (shift_id,)
            )
        context.user_data['edit_report_id'] = report_row['id'] if report_row else None

        # Build list of editable fields
        fields = []
        for key, label_uz, label_ru, *_rest in self._sverka_config():
            label = label_uz if lang == 'uz' else label_ru
            if report_row:
                value = report_row.get(key)
                fields.append((key, label, value))
            else:
                if key in context.user_data:
                    value = context.user_data.get(key)
                    fields.append((key, label, value))

        if not fields:
            msg = "Tahrirlash uchun hisobot topilmadi. Avval sverka tugating." if lang == 'uz' else "Tahrirlash uchun hisobot topilmadi. Avval sverka tugating."
            await context.bot.send_message(chat_id=update.effective_chat.id, text=msg)
            context.user_data['flow'] = None
            return MAIN_MENU

        text_lines = ["Tahrirlanadigan hisobotlar:" if lang == 'uz' else "Tahrirlanadigan hisobotlar:"]
        for _, label, value in fields:
            text_lines.append(f"- {label}: {value if value is not None else 0}")

        keyboard = []
        row = []
        for key, label, value in fields:
            btn = InlineKeyboardButton(f"{label}", callback_data=f"edit:{key}")
            row.append(btn)
            if len(row) == 2:
                keyboard.append(row)
                row = []
        if row:
            keyboard.append(row)
        keyboard.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="edit:back")])

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="\n".join(text_lines),
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return EDIT_REPORT_SELECT

    async def edit_reports_select(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
        context.user_data['flow'] = 'edit'
        key = (query.data or '').split(':', 1)[1] if query.data else ''
        if key == 'back':
            await self.show_cashier_menu(update, context)
            context.user_data['flow'] = None
            return MAIN_MENU

        config = {c[0]: c for c in self._sverka_config()}
        if key not in config:
            await self.show_edit_reports_menu(update, context)
            return EDIT_REPORT_SELECT

        _, label_uz, label_ru, *_rest = config[key]
        label = label_uz if 'uz' == 'uz' else label_ru
        context.user_data['pending_edit_key'] = key
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=f"{label} uchun yangi summani kiriting:"
        )
        return EDIT_REPORT_VALUE

    async def edit_reports_value(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        try:
            amount = self._parse_amount(update.message.text)
        except ValueError:
            await update.message.reply_text(self._invalid_amount_msg(context))
            return EDIT_REPORT_VALUE

        key = context.user_data.get('pending_edit_key')
        if not key:
            await self.show_edit_reports_menu(update, context)
            return EDIT_REPORT_SELECT

        context.user_data.pop('pending_edit_key', None)
        context.user_data[key] = amount
        self._mark_sverka_done(context, key)

        report_id = context.user_data.get('edit_report_id')
        shift_id = context.user_data.get('current_shift_id')
        if report_id:
            await self.db.execute_query(
                f"UPDATE reports SET {key}=%s WHERE id=%s",
                (amount, report_id)
            )
            if not shift_id:
                report_row = await self.db.fetch_one("SELECT shift_id FROM reports WHERE id=%s", (report_id,))
                shift_id = report_row.get("shift_id") if report_row else None

            labels = self._sverka_payment_method_labels()
            field_label = labels.get(key)
            if not field_label:
                for cfg_key, label_uz, *_ in self._sverka_config():
                    if cfg_key == key:
                        field_label = label_uz
                        break
            field_label = field_label or key

            cashier_name = f"{update.effective_user.first_name} {update.effective_user.last_name or ''}".strip()
            location_name = "-"
            if shift_id:
                shift_meta = await self._get_shift_meta(int(shift_id))
                cashier_name = shift_meta.get("cashier") or cashier_name
                location_name = shift_meta.get("location") or "-"
            await self._send_group_message(
                context,
                self._build_report_edit_success_message(
                    cashier_name,
                    location_name,
                    field_label,
                    amount,
                    event_time=getattr(update.message, "date", None),
                ),
            )

        await update.message.reply_text("Saqlab qo'yildi.")
        await self.show_edit_reports_menu(update, context)
        return EDIT_REPORT_SELECT

    async def send_reports(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Ask for a date range and then show reports for that range."""
        context.user_data['admin_reports_range_pending'] = True
        await update.message.reply_text(
            "Qaysi vaqt oralig'ini ko'rasiz?\n"
            "Format:\n"
            "- 2026-03-01 2026-03-16\n"
            "yoki\n"
            "- 01.03.2026 16.03.2026"
        )
        return

    async def handle_reports_menu_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle report period selection from inline menu."""
        query = update.callback_query
        await query.answer()
        action = (query.data or "").split(":", 1)[-1]
        if action in ("daily", "weekly", "monthly"):
            await query.edit_message_text("Filialni tanlang:")
            await self._ask_report_location(update, context, action, chat_id=query.message.chat_id)
            return

        # custom date range
        context.user_data['admin_reports_range_pending'] = True
        await query.edit_message_text(
            "Qaysi vaqt oralig'ini ko'rasiz?\n"
            "Format:\n"
            "- 2026-03-01 2026-03-16\n"
            "yoki\n"
            "- 01.03.2026 16.03.2026"
        )

    async def _ask_report_location(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        period: str,
        chat_id: int = None,
    ):
        locations = await self._get_locations()
        keyboard = [[InlineKeyboardButton("🌐 Barcha filiallar", callback_data=f"reploc:all:{period}")]]
        for loc in locations:
            keyboard.append([InlineKeyboardButton(loc["name"], callback_data=f"reploc:{loc['id']}:{period}")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        target_chat = chat_id or update.effective_chat.id
        await context.bot.send_message(
            chat_id=target_chat,
            text="Qaysi filial hisobotini ko'rasiz?",
            reply_markup=reply_markup,
        )

    async def handle_report_location_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
        parts = (query.data or "").split(":")
        if len(parts) != 3:
            await query.edit_message_text("Noto'g'ri tanlov.")
            return

        _, loc_raw, period = parts
        location_id = None if loc_raw == "all" else int(loc_raw)
        today = self._now_tashkent().date()

        if period == "daily":
            start, end = today, today
        elif period == "weekly":
            start, end = today - timedelta(days=6), today
        elif period == "monthly":
            start, end = today - timedelta(days=29), today
        elif period == "range":
            rng = context.user_data.get("admin_reports_range_values")
            if not rng:
                await query.edit_message_text("Avval vaqt oralig'ini kiriting.")
                return
            start, end = rng
            context.user_data.pop("admin_reports_range_values", None)
        else:
            await query.edit_message_text("Noto'g'ri davr tanlandi.")
            return

        await query.edit_message_text("Hisobot tayyorlanmoqda...")
        await self._send_reports_for_range(
            query.message.chat_id,
            context,
            start,
            end,
            location_id=location_id,
        )

    async def _send_reports_for_range(self, chat_id: int, context: ContextTypes.DEFAULT_TYPE, start, end, location_id=None):
        location_filter_sql = ""
        range_start, range_end = self._day_bounds(start, end)
        params = [range_start, range_end]
        if location_id is not None:
            location_filter_sql = " AND s.location_id = %s"
            params.append(int(location_id))
        query = """
            SELECT
                r.id,
                u.first_name,
                u.last_name,
                l.name AS location,
                s.opened_at,
                s.closed_at,
                COALESCE(s.closing_amount,0) AS closing_amount,
                COALESCE(r.sales_amount,0) AS sales_amount,
                COALESCE(r.debt_received,0) AS debt_received,
                COALESCE(r.expenses,0) AS expenses,
                COALESCE(r.uzcard_amount,0) AS uzcard_amount,
                COALESCE(r.humo_amount,0) AS humo_amount,
                COALESCE(r.p2p_amount,0) AS p2p_amount,
                COALESCE(r.uzcard_refund,0) AS uzcard_refund,
                COALESCE(r.humo_refund,0) AS humo_refund,
                COALESCE(r.other_payments,0) AS other_payments,
                COALESCE(r.debt_payments,0) AS debt_payments,
                COALESCE(r.debt_refunds,0) AS debt_refunds
            FROM shifts s
            JOIN users u ON s.user_id = u.id
            JOIN locations l ON s.location_id = l.id
            JOIN LATERAL (
                SELECT
                    id,
                    sales_amount,
                    debt_received,
                    expenses,
                    uzcard_amount,
                    humo_amount,
                    p2p_amount,
                    uzcard_refund,
                    humo_refund,
                    other_payments,
                    debt_payments,
                    debt_refunds
                FROM reports
                WHERE shift_id = s.id AND report_type = 'daily_report'
                ORDER BY id DESC
                LIMIT 1
            ) r ON TRUE
            WHERE s.opened_at >= %s AND s.opened_at < %s
        """ + location_filter_sql + """
            ORDER BY s.opened_at DESC
            LIMIT 200
        """
        rows = await self.db.fetch_all(query, tuple(params))
        if not rows:
            # Daily so'rovda bugunda ma'lumot bo'lmasa oxirgi mavjud kunni ko'rsatamiz
            if str(start) == str(end):
                latest_q = """
                    SELECT DATE(MAX(s.opened_at)) AS d
                    FROM shifts s
                    JOIN LATERAL (
                        SELECT id
                        FROM reports
                        WHERE shift_id = s.id AND report_type = 'daily_report'
                        ORDER BY id DESC
                        LIMIT 1
                    ) r ON TRUE
                    WHERE 1=1
                """ + location_filter_sql + """
                """
                latest = await self.db.fetch_one(latest_q, tuple(params[2:]) if location_id is not None else ())
                if latest and latest.get('d'):
                    latest_day = str(latest['d'])
                    retry_start, retry_end = self._day_bounds(latest_day)
                    retry_params = [retry_start, retry_end]
                    if location_id is not None:
                        retry_params.append(int(location_id))
                    rows = await self.db.fetch_all(query, tuple(retry_params))
                    if rows:
                        start = latest_day
                        end = latest_day
                        await context.bot.send_message(
                            chat_id=chat_id,
                            text=f"Bugungi hisobot topilmadi. Oxirgi mavjud sana ({latest_day}) ko'rsatildi."
                        )
            if not rows:
                await context.bot.send_message(chat_id=chat_id, text="Hisobotlar topilmadi.")
                return

        title = f"Hisobotlar ({start} - {end})"
        location_name = None
        if location_id is not None:
            loc = await self.db.fetch_one("SELECT name FROM locations WHERE id=%s", (location_id,))
            if loc:
                location_name = loc['name']
                title += f" | Filial: {location_name}"
        lines = [title + ":"]

        def fmt(n):
            try:
                return f"{float(n or 0):,.0f}".replace(",", " ")
            except Exception:
                return str(n)

        for row in rows:
            cashier_name = f"{row['first_name']} {row['last_name'] or ''}".strip()
            total_balance = self._calculate_total_balance(row)
            day = str(row['opened_at'])[:10]
            closed_at = str(row.get('closed_at') or '')[:16]
            closing_amount = fmt(row.get('closing_amount', 0))
            lines.append(
                f"{day} | {cashier_name} | {row['location']} | "
                f"Sof: {fmt(total_balance)} | Yopish: {closing_amount} | Yopilgan: {closed_at or '-'}"
            )

        msg = "\n".join(lines)
        if len(msg) > 3800:
            msg = msg[:3800] + "\n...(qisqartirildi)"
        await context.bot.send_message(chat_id=chat_id, text=msg)

        # Shu oralig' bo'yicha Excel ham yuboramiz
        try:
            xlsx = await asyncio.to_thread(self._build_range_report_xlsx, rows, start, end, location_name)
            await context.bot.send_document(
                chat_id=chat_id,
                document=InputFile(xlsx, filename=f"hisobot_{start}_{end}.xlsx"),
                caption=self._build_export_caption(title, "Excel")
            )
        except Exception:
            logger.exception("Range report Excel yuborishda xatolik")

    def _build_range_report_xlsx(self, rows, start, end, location_name=None) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hisobot"

        headers = [
            "Sana",
            "Kassir",
            "Filial",
            "Yopilgan vaqt",
            "Yopish summasi",
            "Savdo",
            "Kelgan qarz",
            "Chiqim",
            "Uzcard",
            "Humo",
            "P2P",
            "Uzcard vozvrat",
            "Humo vozvrat",
            "Boshqa to'lovlar",
            "Qarzga berilgan to'lovlar",
            "Vozvrat qarzlar",
            "Naqd kutiladigan summa",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def _f(v):
            try:
                return float(v or 0)
            except Exception:
                return 0.0

        for row in rows:
            cashier_name = f"{row['first_name']} {row['last_name'] or ''}".strip()
            total_balance = self._calculate_total_balance(row)
            ws.append([
                str(row['opened_at'])[:10],
                cashier_name,
                row['location'],
                str(row.get('closed_at') or '')[:19],
                _f(row.get('closing_amount')),
                _f(row['sales_amount']),
                _f(row['debt_received']),
                _f(row['expenses']),
                _f(row['uzcard_amount']),
                _f(row['humo_amount']),
                _f(row['p2p_amount']),
                _f(row['uzcard_refund']),
                _f(row['humo_refund']),
                _f(row['other_payments']),
                _f(row['debt_payments']),
                _f(row['debt_refunds']),
                total_balance,
            ])

        for col in range(5, 18):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = "#,##0"

        for col in range(1, ws.max_column + 1):
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 34)

        meta = wb.create_sheet("Ma'lumot")
        meta.append(["Boshlanish", str(start)])
        meta.append(["Tugash", str(end)])
        meta.append(["Filial", location_name or "Barcha filiallar"])
        meta.append(["Yaratilgan vaqt", self._now_tashkent().strftime("%Y-%m-%d %H:%M:%S")])

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    def _parse_date_range(self, text: str):
        text = (text or "").strip()
        parts = re.findall(r"\d{4}-\d{2}-\d{2}", text)
        fmt = "%Y-%m-%d"
        if len(parts) < 2:
            parts = re.findall(r"\d{2}[./]\d{2}[./]\d{4}", text)
            fmt = "%d.%m.%Y" if "." in (parts[0] if parts else "") else "%d/%m/%Y"
        if len(parts) < 2:
            return None
        try:
            start = datetime.strptime(parts[0].replace("/", "."), fmt).date()
            end = datetime.strptime(parts[1].replace("/", "."), fmt).date()
        except Exception:
            return None
        if start > end:
            start, end = end, start
        return start, end

    async def handle_admin_reports_range(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        rng = self._parse_date_range(update.message.text if update.message else "")
        if not rng:
            await update.message.reply_text(
                "Format noto'g'ri. Masalan:\n"
                "2026-03-01 2026-03-16\n"
                "yoki\n"
                "01.03.2026 16.03.2026"
            )
            return

        context.user_data['admin_reports_range_pending'] = False
        start, end = rng
        context.user_data['admin_reports_range_values'] = (start, end)
        await self._ask_report_location(update, context, "range")
    async def send_all_cashiers(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Send all cashiers to admin"""
        query = "SELECT * FROM users WHERE role = 'cashier' AND is_active = TRUE"
        cashiers = await self.db.fetch_all(query)
        
        lang = 'uz'
        
        if lang == 'uz':
            if cashiers:
                msg = "Barcha kassirlar:\n"
                for cashier in cashiers:
                    msg += f"- {cashier['first_name']} {cashier['last_name']} ({cashier['phone_number']})\n"
            else:
                msg = "Hech qanday kassir topilmadi."
        else:
            if cashiers:
                msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р Р‹Р В РЎвЂњР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“:\n"
                for cashier in cashiers:
                    msg += f"- {cashier['first_name']} {cashier['last_name']} ({cashier['phone_number']})\n"
            else:
                msg = "Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’Вµ Р В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°Р В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚Сћ Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚В Р В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р РЋРІР‚вЂњР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°."
                
        await update.message.reply_text(msg)

    async def handle_approval_requests(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle cashier approval requests"""
        query = "SELECT * FROM approval_requests WHERE status = 'pending'"
        requests = await self.db.fetch_all(query)
        
        lang = 'uz'
        
        if lang == 'uz':
            if requests:
                msg = "Kassir so'rovlari:"
                for req in requests:
                    text = f"{req['first_name']} {req['last_name']} ({req['phone_number']}) | ID: {req['telegram_id']}"
                    keyboard = InlineKeyboardMarkup([
                        [
                            InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve:{req['telegram_id']}"),
                            InlineKeyboardButton("❌ Yo'q", callback_data=f"reject:{req['telegram_id']}")
                        ]
                    ])
                    await update.message.reply_text(text, reply_markup=keyboard)
            else:
                msg = "Yangi so'rovlar yo'q."
        else:
            if requests:
                msg = "Р В Р’В Р Р†Р вЂљРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРІвЂћвЂ“ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В :"
                for req in requests:
                    text = f"{req['first_name']} {req['last_name']} ({req['phone_number']}) | ID: {req['telegram_id']}"
                    keyboard = InlineKeyboardMarkup([
                        [
                            InlineKeyboardButton("Р В Р вЂ Р РЋРЎв„ўР Р†Р вЂљР’В¦ Р В Р’В Р РЋРІР‚С”Р В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В±Р В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р В Р вЂ°", callback_data=f"approve:{req['telegram_id']}"),
                            InlineKeyboardButton("Р В Р вЂ Р РЋРЎС™Р В Р вЂ° Р В Р’В Р РЋРІР‚С”Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В·", callback_data=f"reject:{req['telegram_id']}")
                        ]
                    ])
                    await update.message.reply_text(text, reply_markup=keyboard)
            else:
                msg = "Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў Р В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р Р‹Р Р†Р вЂљР’В¦ Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚СћР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В ."
                
        if msg:
            await update.message.reply_text(msg)

    async def handle_approval_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle inline approve/reject callbacks"""
        query = update.callback_query
        await query.answer()

        data = query.data or ""
        if not (data.startswith("approve:") or data.startswith("reject:")):
            return

        # Only admins can approve/reject
        admin = await self.db.fetch_one(
            "SELECT * FROM users WHERE telegram_id = %s AND role = 'admin' AND is_active = TRUE",
            (update.effective_user.id,)
        )
        if not admin:
            await query.edit_message_text("Faqat admin tasdiqlashi mumkin.")
            return

        try:
            telegram_id = int(data.split(":", 1)[1])
        except ValueError:
            await query.edit_message_text("Noto'g'ri ID.")
            return

        if data.startswith("approve:"):
            await self.approve_cashier(update, context, telegram_id)
            await query.edit_message_text("So'rov tasdiqlandi.")
        else:
            await self.reject_cashier(update, context, telegram_id)
            await query.edit_message_text("So'rov rad etildi.")

    async def notify_admins_new_request(self, context: ContextTypes.DEFAULT_TYPE, user_data: dict):
        """Notify all admins about new cashier approval request"""
        admins = await self.db.fetch_all("SELECT telegram_id FROM users WHERE role = 'admin' AND is_active = TRUE")
        if not admins:
            return

        text = (
            "Yangi kassir so'rovi:\n"
            f"{user_data['first_name']} {user_data['last_name']} ({user_data['phone_number']})\n"
            f"ID: {user_data['telegram_id']}"
        )
        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve:{user_data['telegram_id']}"),
                InlineKeyboardButton("❌ Yo'q", callback_data=f"reject:{user_data['telegram_id']}")
            ]
        ])

        for admin in admins:
            try:
                await context.bot.send_message(chat_id=admin['telegram_id'], text=text, reply_markup=keyboard)
            except Exception:
                continue

    def _extract_telegram_id(self, text: str):
        """Extract numeric telegram id from text."""
        match = re.search(r"\b(\d{5,20})\b", text)
        if match:
            try:
                return int(match.group(1))
            except ValueError:
                return None
        return None

    async def approve_cashier(self, update: Update, context: ContextTypes.DEFAULT_TYPE, telegram_id: int):
        """Approve cashier registration request"""
        # Check pending request
        query = "SELECT * FROM approval_requests WHERE telegram_id = %s AND status = 'pending'"
        req = await self.db.fetch_one(query, (telegram_id,))
        if not req:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="Bunday pending so'rov topilmadi.")
            return

        # Insert or reactivate user
        user = await self.db.fetch_one("SELECT * FROM users WHERE telegram_id = %s", (telegram_id,))
        if user:
            await self.db.execute_query(
                "UPDATE users SET role = 'cashier', is_active = TRUE, password_hash = NULL WHERE telegram_id = %s",
                (telegram_id,)
            )
        else:
            await self.db.execute_query(
                """
                INSERT INTO users (telegram_id, first_name, last_name, phone_number, role, password_hash, is_active)
                VALUES (%s, %s, %s, %s, 'cashier', %s, TRUE)
                """,
                (req['telegram_id'], req['first_name'], req['last_name'], req['phone_number'], None)
            )

        # Update request status
        await self.db.execute_query(
            "UPDATE approval_requests SET status = 'approved', approved_at = NOW() WHERE telegram_id = %s",
            (telegram_id,)
        )

        self._prime_cashier_password_setup(context, telegram_id)

        # Notify cashier
        try:
            await context.bot.send_message(
                chat_id=telegram_id,
                text="Sizning so'rovingiz tasdiqlandi. Endi yangi parol kiriting:",
                reply_markup=ReplyKeyboardRemove(),
            )
        except Exception:
            pass

        await context.bot.send_message(chat_id=update.effective_chat.id, text="So'rov tasdiqlandi.")

    async def reject_cashier(self, update: Update, context: ContextTypes.DEFAULT_TYPE, telegram_id: int):
        """Reject cashier registration request"""
        query = "SELECT * FROM approval_requests WHERE telegram_id = %s AND status = 'pending'"
        req = await self.db.fetch_one(query, (telegram_id,))
        if not req:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="Bunday pending so'rov topilmadi.")
            return

        await self.db.execute_query(
            "UPDATE approval_requests SET status = 'rejected', approved_at = NOW() WHERE telegram_id = %s",
            (telegram_id,)
        )

        try:
            await context.bot.send_message(
                chat_id=telegram_id,
                text="Sizning so'rovingiz rad etildi. Administrator bilan bog'laning."
            )
        except Exception:
            pass

        await context.bot.send_message(chat_id=update.effective_chat.id, text="So'rov rad etildi.")

    async def _save_shift_image(self, shift_id: int, image_type: str, file_id: str):
        """Persist image reference for a shift"""
        await self.db.execute_query(
            """
            INSERT INTO images (shift_id, image_url, image_type)
            VALUES (%s, %s, %s)
            """,
            (shift_id, file_id, image_type)
        )

    async def _count_shift_images(self, shift_id: int, image_type: str) -> int:
        row = await self.db.fetch_one(
            "SELECT COUNT(*) as cnt FROM images WHERE shift_id=%s AND image_type=%s",
            (shift_id, image_type)
        )
        return int(row['cnt']) if row else 0

    async def _today_shift_for_user(self, user_id: int):
        """Foydalanuvchining bugungi (ochilgan sanasi bugun bo'lgan) oxirgi smenasi."""
        today = self._now_tashkent().date().isoformat()
        start_bound, end_bound = self._day_bounds(today)
        return await self.db.fetch_one(
            """
            SELECT id, is_open, opened_at, closed_at
            FROM shifts
            WHERE user_id=%s AND opened_at >= %s AND opened_at < %s
            ORDER BY opened_at DESC, id DESC
            LIMIT 1
            """,
            (user_id, start_bound, end_bound)
        )

    def _get_image_file_id(self, update: Update):
        """Telegram'dan rasm file_id ni olish (photo yoki image document)."""
        msg = update.message
        if not msg:
            return None
        photo = getattr(msg, "photo", None)
        if photo:
            return photo[-1].file_id
        document = getattr(msg, "document", None)
        if document:
            mime = (getattr(document, "mime_type", "") or "").lower()
            name = (getattr(document, "file_name", "") or "").lower()
            if mime.startswith("image/") or name.endswith((".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif", ".heic")):
                return document.file_id
        return None

    def _is_blocked_media_group(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
        """Agar oldingi bosqichda qabul qilingan albomning qolgan rasmlari kelsa, ularni e'tiborsiz qoldirish."""
        msg = update.message
        if not msg:
            return False
        media_group_id = getattr(msg, "media_group_id", None)
        blocked = context.user_data.get("blocked_media_group_id")
        return bool(media_group_id and blocked and str(media_group_id) == str(blocked))

    def _block_current_media_group(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Bosqich yakunlanganda shu albomning qolgan rasmlari keyingi bosqichga o'tib ketmasin."""
        msg = update.message
        if not msg:
            return
        media_group_id = getattr(msg, "media_group_id", None)
        if media_group_id:
            context.user_data["blocked_media_group_id"] = str(media_group_id)

    def _sync_opening_stage_with_media_group(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        """
        Keep opening stage on the same question while one media-group is still arriving.
        This lets us accept all images from the same album (including duplicates).
        """
        locked_group_id = context.user_data.get("opening_stage_locked_media_group_id")
        if not locked_group_id:
            return

        msg = update.message
        current_group_id = getattr(msg, "media_group_id", None) if msg else None
        locked_stage_name = context.user_data.get("opening_stage_locked_name")
        if current_group_id and str(current_group_id) == str(locked_group_id):
            if locked_stage_name:
                context.user_data["opening_stage"] = locked_stage_name
            return

        next_stage = context.user_data.get("pending_next_opening_stage")
        if next_stage:
            context.user_data["opening_stage"] = next_stage
        context.user_data.pop("opening_stage_locked_media_group_id", None)
        context.user_data.pop("opening_stage_locked_name", None)
        context.user_data.pop("pending_next_opening_stage", None)
        context.user_data.pop("opening_stage_completed_prompt_sent", None)

    def _lock_opening_stage_for_media_group(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        current_stage: str,
        next_stage: str,
    ) -> bool:
        msg = update.message
        media_group_id = getattr(msg, "media_group_id", None) if msg else None
        if not media_group_id:
            return False
        context.user_data["opening_stage_locked_media_group_id"] = str(media_group_id)
        context.user_data["opening_stage_locked_name"] = current_stage
        context.user_data["pending_next_opening_stage"] = next_stage
        context.user_data["opening_stage"] = current_stage
        return True

    def _stage_prompt_once_key(self, stage: str) -> str:
        return f"{stage}:prompted"

    def _mark_stage_prompted(self, context: ContextTypes.DEFAULT_TYPE, stage: str) -> None:
        prompted = context.user_data.get("opening_stage_completed_prompt_sent")
        if not isinstance(prompted, dict):
            prompted = {}
        prompted[self._stage_prompt_once_key(stage)] = True
        context.user_data["opening_stage_completed_prompt_sent"] = prompted

    def _is_stage_prompted(self, context: ContextTypes.DEFAULT_TYPE, stage: str) -> bool:
        prompted = context.user_data.get("opening_stage_completed_prompt_sent")
        if not isinstance(prompted, dict):
            return False
        return bool(prompted.get(self._stage_prompt_once_key(stage)))

    def _clear_stage_prompted(self, context: ContextTypes.DEFAULT_TYPE, stage: str) -> None:
        prompted = context.user_data.get("opening_stage_completed_prompt_sent")
        if not isinstance(prompted, dict):
            return
        prompted.pop(self._stage_prompt_once_key(stage), None)
        if prompted:
            context.user_data["opening_stage_completed_prompt_sent"] = prompted
        else:
            context.user_data.pop("opening_stage_completed_prompt_sent", None)

    def _cancel_receipt_roll_finalize_task(self, context: ContextTypes.DEFAULT_TYPE) -> None:
        task = context.user_data.pop("receipt_roll_finalize_task", None)
        if task and not task.done():
            task.cancel()

    def _queue_opening_group_photo(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        file_id: str,
        image_title: str,
        event_time=None,
        media_kind: str = "photo",
    ) -> None:
        queue = context.user_data.get("pending_opening_group_photos")
        if not isinstance(queue, list):
            queue = []
        queue.append(
            {
                "file_id": file_id,
                "image_title": image_title,
                "event_time": event_time,
                "media_kind": media_kind,
            }
        )
        context.user_data["pending_opening_group_photos"] = queue

    def _get_image_media_kind(self, update: Update) -> str:
        msg = update.message
        if msg and getattr(msg, "photo", None):
            return "photo"
        return "document"

    async def _send_group_media_album(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        items: list[dict],
        caption: str = "",
    ) -> bool:
        group_chat_id = await self._get_group_chat_id()
        if not group_chat_id:
            logger.warning("Group media album skipped: group_chat_id is not configured")
            return False
        if not items:
            return False

        media = []
        for idx, item in enumerate(items):
            file_id = item.get("file_id")
            if not file_id:
                continue
            item_caption = caption if idx == 0 else None
            if item.get("media_kind") == "document":
                media.append(InputMediaDocument(media=file_id, caption=item_caption))
            else:
                media.append(InputMediaPhoto(media=file_id, caption=item_caption))
        if not media:
            return False

        try:
            # Telegram allows max 10 media items in one album.
            for i in range(0, len(media), 10):
                chunk = media[i : i + 10]
                if i > 0 and chunk:
                    # caption only for the first sent chunk
                    first = chunk[0]
                    if isinstance(first, InputMediaPhoto):
                        chunk[0] = InputMediaPhoto(media=first.media)
                    else:
                        chunk[0] = InputMediaDocument(media=first.media)
                await context.bot.send_media_group(chat_id=group_chat_id, media=chunk)
            return True
        except Exception:
            logger.exception("Failed to send media album to %s", group_chat_id)
            return False

    async def _add_image_label(self, bot, file_id: str, label: str) -> Optional[BytesIO]:
        """Download a Telegram photo and wrap it with a labeled border/header."""
        try:
            tg_file = await bot.get_file(file_id)
            buf = BytesIO()
            await tg_file.download_to_memory(buf)
            buf.seek(0)
            return self._decorate_labeled_check_image(buf.getvalue(), label, "")
        except Exception:
            logger.exception("_add_image_label failed: file_id=%s label=%s", file_id, label)
            return None

    async def _send_opening_group_photo_album(self, bot_client, queue: list) -> bool:
        """Send opening photos as one watermarked album without depending on user session state."""
        group_chat_id = await self._get_group_chat_id()
        if not group_chat_id:
            logger.warning("Opening photo album skipped: group_chat_id not configured")
            return False

        # Build one flat media list — all image types combined
        media = []
        upload_counter = 0
        for item in queue:
            file_id = item.get("file_id")
            image_title = item.get("image_title", "")
            media_kind = item.get("media_kind", "photo")
            if not file_id:
                continue

            if media_kind == "photo":
                try:
                    labeled_buf = await self._add_image_label(bot_client, file_id, image_title)
                    if labeled_buf:
                        fname = f"photo_{upload_counter}.jpg"
                        upload_counter += 1
                        raw_bytes = labeled_buf.getvalue()
                        media.append(InputMediaPhoto(media=raw_bytes, filename=fname))
                    else:
                        media.append(InputMediaPhoto(media=file_id))
                except Exception:
                    logger.exception("Skipping image label for file_id=%s, using original", file_id)
                    media.append(InputMediaPhoto(media=file_id))
            else:
                media.append(InputMediaDocument(media=file_id))

        if not media:
            return False

        try:
            # Telegram max = 10 per album; split automatically
            for i in range(0, len(media), 10):
                chunk = media[i : i + 10]
                await bot_client.send_media_group(chat_id=group_chat_id, media=chunk)
            return True
        except Exception:
            logger.exception("Failed to send opening photo album to group %s", group_chat_id)
            return False

    async def _flush_opening_group_photos(self, context: ContextTypes.DEFAULT_TYPE, shift_id: int) -> None:
        """Send all opening photos as a single watermarked album to the group."""
        queue = context.user_data.get("pending_opening_group_photos")
        if not isinstance(queue, list) or not queue:
            context.user_data.pop("pending_opening_group_photos", None)
            return

        bot_client = getattr(context, "bot", None)
        if bot_client:
            await self._send_opening_group_photo_album(bot_client, list(queue))

        context.user_data.pop("pending_opening_group_photos", None)

    async def _send_opening_group_notifications(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        opening_photo_queue: list,
        opening_message: str,
        chat_id: int,
    ) -> None:
        try:
            if opening_photo_queue:
                await self._send_opening_group_photo_album(context.bot, opening_photo_queue)

            sent = await self._send_group_message(context, opening_message)
            if not sent:
                await context.bot.send_message(
                    chat_id=chat_id,
                    text="Diqqat: smena ochilgani guruhga yuborilmadi. /setgroup va bot ruxsatlarini tekshiring.",
                )
        except Exception:
            logger.exception("Opening group notification background task failed")

    def _schedule_opening_group_notifications(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        opening_photo_queue: list,
        opening_message: str,
        chat_id: int,
    ) -> None:
        asyncio.create_task(
            self._send_opening_group_notifications(
                context,
                opening_photo_queue,
                opening_message,
                chat_id,
            )
        )

    async def _finalize_shift_opening_flow(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        chat_id: int,
        cashier_first_name: str,
        cashier_last_name: str,
    ) -> None:
        if context.user_data.get("opening_finalize_done"):
            return
        context.user_data["opening_finalize_done"] = True

        shift_id = context.user_data.get('current_shift_id')
        location_name = await self._get_location_name(context.user_data.get('location_id'))
        cashier_name = f"{cashier_first_name} {cashier_last_name or ''}".strip()
        opening_message = self._build_shift_opened_message(
            cashier_name,
            location_name,
            context.user_data.get('opening_amount', 0),
            context.user_data.get('opening_amount_time', ''),
        )
        opening_photo_queue = list(context.user_data.get("pending_opening_group_photos") or [])

        await context.bot.send_message(chat_id=chat_id, text="Smena muvaffaqiyatli ochildi! Endi sverka jarayonini boshlang.")
        await context.bot.send_message(
            chat_id=chat_id,
            text="Kassir menyusi:",
            reply_markup=self._build_cashier_menu_keyboard(),
        )

        self._cancel_receipt_roll_finalize_task(context)
        context.user_data.pop("receipt_roll_finalize_token", None)
        context.user_data['flow'] = None
        context.user_data.pop('opening_stage', None)
        context.user_data.pop("pending_opening_group_photos", None)
        context.user_data.pop("opening_stage_locked_media_group_id", None)
        context.user_data.pop("opening_stage_locked_name", None)
        context.user_data.pop("pending_next_opening_stage", None)
        context.user_data.pop("opening_stage_completed_prompt_sent", None)
        context.user_data.pop("opening_finalize_done", None)

        if shift_id:
            self._schedule_opening_group_notifications(
                context,
                opening_photo_queue,
                opening_message,
                chat_id,
            )

    def _schedule_receipt_roll_finalize(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        chat_id: int,
        cashier_first_name: str,
        cashier_last_name: str,
    ) -> None:
        self._cancel_receipt_roll_finalize_task(context)
        token = int(context.user_data.get("receipt_roll_finalize_token", 0)) + 1
        context.user_data["receipt_roll_finalize_token"] = token

        async def _runner():
            try:
                await asyncio.sleep(1.2)
            except asyncio.CancelledError:
                return
            if context.user_data.get("receipt_roll_finalize_token") != token:
                return
            if context.user_data.get("opening_stage") != "receipt_roll":
                return
            await self._finalize_shift_opening_flow(
                context,
                chat_id=chat_id,
                cashier_first_name=cashier_first_name,
                cashier_last_name=cashier_last_name,
            )

        context.user_data["receipt_roll_finalize_task"] = asyncio.create_task(_runner())

    def _build_close_shift_progress_text(self, done: int, total: int, current_step: str) -> str:
        total = max(int(total or 0), 1)
        done = max(0, min(int(done or 0), total))
        percent = int((done / total) * 100)
        remaining = max(total - done, 0)
        return (
            "Smena yopilmoqda. Iltimos, kuting.\n"
            f"Jarayon: {done}/{total} ({percent}%)\n"
            f"Hozir: {current_step}\n"
            f"Qoldi: {remaining} ta bosqich"
        )

    async def _update_close_shift_progress(self, progress_message, done: int, total: int, current_step: str):
        """Update one progress message so the cashier sees ongoing work."""
        if not progress_message:
            return
        text = self._build_close_shift_progress_text(done, total, current_step)
        try:
            await progress_message.edit_text(text)
        except Exception:
            pass

    async def _download_shift_image_blobs(self, context: ContextTypes.DEFAULT_TYPE, images: list, progress_callback=None) -> dict:
        """Download each unique Telegram image once and reuse it across Excel exports."""
        blobs = {}
        seen = set()

        for item in images or []:
            file_ref = (item.get("image_url") or "").strip()
            if not file_ref or file_ref in seen:
                continue
            seen.add(file_ref)
            try:
                tg_file = await context.bot.get_file(file_ref)
                data = await tg_file.download_as_bytearray()
                blobs[file_ref] = bytes(data)
            except Exception:
                logger.exception("Failed to download Telegram image for Excel export: %s", file_ref)
                blobs[file_ref] = None
            if progress_callback:
                await progress_callback(file_ref)

        return blobs

    def _embed_excel_image(self, worksheet, cell: str, image_bytes, *, width: int = 260, height: int = 180) -> bool:
        """Embed image bytes into an Excel sheet cell."""
        if not image_bytes:
            return False
        try:
            bio = BytesIO(image_bytes)
            img = PILImage.open(bio)
            out = BytesIO()
            img.convert("RGB").save(out, format="JPEG", quality=85)
            out.seek(0)
            xl_img = XLImage(out)
            xl_img.width = width
            xl_img.height = height
            # Keep the underlying buffer alive until workbook serialization finishes.
            xl_img._source_buffer = out
            worksheet.add_image(xl_img, cell)
            return True
        except Exception:
            return False

    def _parse_amount(self, text: str) -> float:
        """Parse amount values that contain digits only (spaces are allowed)."""
        raw = (text or "").strip()
        normalized = "".join(raw.split())
        if not normalized or not normalized.isdigit():
            raise ValueError("Amount must contain digits only.")
        return float(normalized)

    def _format_telegram_time(self, dt_value) -> str:
        """Telegram message vaqtini Asia/Tashkent ga o'tkazib formatlaydi."""
        if not dt_value:
            return self._now_tashkent().strftime("%Y-%m-%d %H:%M:%S")
        try:
            return dt_value.astimezone(TASHKENT_TZ).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(dt_value)[:19]

    def _sverka_config(self):
        return [
            ('sales_amount', "Savdo summasi", "Р В Р Р‹Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В° Р В РЎвЂ”Р РЋР вЂљР В РЎвЂўР В РўвЂР В Р’В°Р В Р’В¶", REPORT_SALES, "Bugungi savdo miqdorini kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР РЋРЎвЂњ Р В РЎвЂ”Р РЋР вЂљР В РЎвЂўР В РўвЂР В Р’В°Р В Р’В¶:"),
            ('debt_received', "Kelgan qarzlar", "Р В РЎСџР РЋР вЂљР В РЎвЂР РЋРІвЂљВ¬Р В Р’ВµР В РўвЂР РЋРІвЂљВ¬Р В РЎвЂР В Р’Вµ Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“Р В РЎвЂ", REPORT_DEBT_RECEIVED, "Kelgan qarzlarni kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂ”Р РЋР вЂљР В РЎвЂР РЋРІвЂљВ¬Р В Р’ВµР В РўвЂР РЋРІвЂљВ¬Р В РЎвЂР В Р’Вµ Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“Р В РЎвЂ (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('expenses', "Chiqimlar", "Р В Р’В Р В Р’В°Р РЋР С“Р РЋРІР‚В¦Р В РЎвЂўР В РўвЂР РЋРІР‚в„–", REPORT_EXPENSES, "Chiqim sababini kiriting.\nMasalan:\nMirshod Dastafka -- 10 000\nUlug Paynet -- 100 000\n\nAgar chiqim bo'lmasa, Yakunlashni bosing.", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Ռ Վ…Р В ՐВ°Р РЋՌ…Ռ Վ…Ռ В ӮՌ Վ…Ռ РЋРІР‚в„–Р РЋՌ…Р РЋРІР‚В° Ռ Ր…Ռ Վ…Ռ В ӮՌ РЋՌ…Р В ӨՌ Վ…Ռ В ��."),
            ('uzcard_amount', "Uzcard summasi", "Uzcard Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°", REPORT_UZCARD, "Uzcard orqali kiritilgan summani kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР РЋРЎвЂњ Р В РЎвЂ”Р В РЎвЂў Uzcard (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('humo_amount', "Humo summasi", "Humo Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°", REPORT_HUMO, "Humo orqali kiritilgan summani kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР РЋРЎвЂњ Р В РЎвЂ”Р В РЎвЂў Humo (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('p2p_amount', "P2P summasi", "P2P Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°", REPORT_P2P, "P2P orqali kiritilgan summani kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР РЋРЎвЂњ Р В РЎвЂ”Р В РЎвЂў P2P (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('uzcard_refund', "Uzcard vozvrat", "Р В РІР‚в„ўР В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Uzcard", REPORT_UZCARD_REFUND, "Uzcard orqali vozvrat bo'lgan summani kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В Р вЂ Р В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Р В РЎвЂ”Р В РЎвЂў Uzcard (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('humo_refund', "Humo vozvrat", "Р В РІР‚в„ўР В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Humo", REPORT_HUMO_REFUND, "Humo orqali vozvrat bo'lgan summani kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В Р вЂ Р В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Р В РЎвЂ”Р В РЎвЂў Humo (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('other_payments', "Boshqa to'lovlar", "Р В РІР‚СњР РЋР вЂљР РЋРЎвЂњР В РЎвЂ“Р В РЎвЂР В Р’Вµ Р В РЎвЂўР В РЎвЂ”Р В Р’В»Р В Р’В°Р РЋРІР‚С™Р РЋРІР‚в„–", REPORT_OTHER_PAYMENTS, "Boshqa to'lovlar bo'yicha izohni kiriting:", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РўвЂР РЋР вЂљР РЋРЎвЂњР В РЎвЂ“Р В РЎвЂР В Р’Вµ Ռ В Ռ…Р В Р№Р В Ռ…Ռ В ӨР Վ…Ռ В ՐВ°Ռ В ЎвЂұՌ В ЎвЂ”Р В Р’В°Р РЋРІР‚С™Ռ РЋՌ…:"),
            ('debt_payments', "Qarzga berilgan to'lovlar", "Р В РІР‚в„ўР РЋРІР‚в„–Р В РўвЂР В Р’В°Р В Р вЂ¦Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р вЂ  Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“", REPORT_DEBT_PAYMENTS, "Qarzga berilgan to'lovlarni kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В Р вЂ Р РЋРІР‚в„–Р В РўвЂР В Р’В°Р В Р вЂ¦Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р вЂ  Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“ (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):"),
            ('debt_refunds', "Vozvrat qarzlar", "Р В РІР‚в„ўР В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“Р В РЎвЂўР В Р вЂ ", REPORT_DEBT_REFUNDS, "Vozvrat qarzlarni kiriting (summa):", "Р В РІР‚в„ўР В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В Р вЂ Р В РЎвЂўР В Р’В·Р В Р вЂ Р РЋР вЂљР В Р’В°Р РЋРІР‚С™ Р В РўвЂР В РЎвЂўР В Р’В»Р В РЎвЂ“Р В РЎвЂўР В Р вЂ  (Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР В Р’В°):")
        ]

    def _active_sverka_config(self, context: ContextTypes.DEFAULT_TYPE):
        config = self._sverka_config()
        if context.user_data.get("sverka_entrypoint") != "closing":
            return config
        excluded = {
            "uzcard_amount",
            "humo_amount",
            "p2p_amount",
            "uzcard_refund",
            "humo_refund",
            "debt_refunds",
        }
        closing_config = [item for item in config if item[0] not in excluded]
        closing_config.append(
            (
                "tax_info",
                "Soliq ma'lumotlari",
                "Soliq ma'lumotlari",
                REPORT_TAX_INFO,
                "Soliq cheki rasmini yuboring:",
                "Soliq cheki rasmini yuboring:",
            )
        )
        return closing_config

    def _opening_requirements_config(self):
        return [
            ("workplace_status", "Ish joyi holati rasmi", 2, UPLOAD_WORKPLACE_STATUS, "Ish stolingizni rasmga olib yuboring (2 ta rasm)."),
            ("terminal_power", "Terminal/ratsiya quvvati rasmi", 1, UPLOAD_TERMINAL_POWER, "Terminallar va ratsiyalar quvvat holatini rasmga oling."),
            ("zero_report", "Uzcard/Humo nol hisobot rasmi", 1, UPLOAD_ZERO_REPORT, "Uzcard va Humo nol hisobot rasmini yuboring."),
            ("opening_notification", "iiko/soliq ochilish rasmi", 1, UPLOAD_OPENING_NOTIFICATION, "iiko/soliq tizimida smena ochilganlik rasmini yuboring."),
            ("receipt_roll", "Zaxira chek lenta rasmi", 1, UPLOAD_RECEIPT_ROLL, "Zaxira chek lenta mavjudligi rasmini yuboring."),
        ]

    async def _opening_missing_lines(self, shift_id: int):
        lines = []
        for key, label, required_count, *_ in self._opening_requirements_config():
            count = await self._count_shift_images(shift_id, key)
            if count < required_count:
                remain = required_count - count
                if required_count == 1:
                    lines.append(f"- {label}")
                else:
                    lines.append(f"- {label} ({remain} ta qolgan)")
        return lines

    async def show_opening_requirements_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE, shift_id: int, note: Optional[str] = None):
        cfg = self._opening_requirements_config()
        keyboard = []
        for key, label, required_count, *_ in cfg:
            count = await self._count_shift_images(shift_id, key)
            done = count >= required_count
            icon = "✅" if done else "❌"
            suffix = ""
            if required_count > 1:
                suffix = f" ({count}/{required_count})"
            keyboard.append([InlineKeyboardButton(f"{icon} {label}{suffix}", callback_data=f"op:{key}")])

        keyboard.append([
            InlineKeyboardButton("🔄 Yangilash", callback_data="op:refresh"),
            InlineKeyboardButton("⬅️ Orqaga", callback_data="op:back"),
        ])

        missing = await self._opening_missing_lines(shift_id)
        text = note or "Smena ochish rasmlari holati:"
        if missing:
            text += "\n\nYetishmayotganlar:\n" + "\n".join(missing)
        else:
            text += "\n\nBarcha kerakli rasmlar yuklangan."

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=text,
            reply_markup=InlineKeyboardMarkup(keyboard),
        )

    async def opening_select_step(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        if not query or not query.data:
            return MAIN_MENU
        await query.answer()

        key = query.data.split(":", 1)[1] if ":" in query.data else ""

        user_row = await self.db.fetch_one("SELECT id FROM users WHERE telegram_id=%s", (update.effective_user.id,))
        if not user_row:
            await context.bot.send_message(chat_id=query.message.chat_id, text="Foydalanuvchi topilmadi.")
            return MAIN_MENU
        active_shift = await self.db.fetch_one(
            "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC LIMIT 1",
            (user_row['id'],)
        )
        if not active_shift:
            await context.bot.send_message(chat_id=query.message.chat_id, text="Ochiq smena topilmadi.")
            return MAIN_MENU

        shift_id = int(active_shift["id"])
        context.user_data["current_shift_id"] = shift_id
        context.user_data["flow"] = "opening"

        if key == "back":
            await self.show_cashier_menu(update, context)
            return MAIN_MENU
        if key == "refresh":
            await self.show_opening_requirements_menu(update, context, shift_id)
            return MAIN_MENU

        cfg = {c[0]: c for c in self._opening_requirements_config()}
        if key not in cfg:
            await self.show_opening_requirements_menu(update, context, shift_id)
            return MAIN_MENU

        _, label, required_count, state, prompt = cfg[key]
        current_count = await self._count_shift_images(shift_id, key)
        if current_count >= required_count:
            await context.bot.send_message(chat_id=query.message.chat_id, text=f"✅ {label} allaqachon bajarilgan.")
            await self.show_opening_requirements_menu(update, context, shift_id)
            return MAIN_MENU

        context.user_data["opening_stage"] = key
        await context.bot.send_message(chat_id=query.message.chat_id, text=prompt, reply_markup=ReplyKeyboardRemove())
        return state

    def _init_sverka_status(self, context: ContextTypes.DEFAULT_TYPE):
        status = context.user_data.get('sverka_status')
        if not isinstance(status, dict):
            status = {}
        active_status = {}
        for key, *_ in self._active_sverka_config(context):
            active_status[key] = bool(status.get(key, False))
            if not active_status[key] and context.user_data.get(key) is not None:
                active_status[key] = True
        context.user_data['sverka_status'] = active_status

    def _mark_sverka_done(self, context: ContextTypes.DEFAULT_TYPE, key: str):
        self._init_sverka_status(context)
        context.user_data['sverka_status'][key] = True

    def _sverka_all_done(self, context: ContextTypes.DEFAULT_TYPE) -> bool:
        self._init_sverka_status(context)
        return all(context.user_data['sverka_status'].get(k, False) for k, *_ in self._active_sverka_config(context))

    def _invalid_amount_msg(self, context: ContextTypes.DEFAULT_TYPE) -> str:
        lang = 'uz'
        if lang == 'uz':
            return "Iltimos, faqat raqam kiriting. Masalan: 0 yoki 120000."
        return "Р В РЎСџР В РЎвЂўР В Р’В¶Р В Р’В°Р В Р’В»Р РЋРЎвЂњР В РІвЂћвЂ“Р РЋР С“Р РЋРІР‚С™Р В Р’В°, Р В Р вЂ Р В Р вЂ Р В Р’ВµР В РўвЂР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В РЎвЂР В Р’В»Р РЋР Р‰Р В Р вЂ¦Р РЋРЎвЂњР РЋР вЂ№ Р РЋР С“Р РЋРЎвЂњР В РЎВР В РЎВР РЋРЎвЂњ."

    def _sverka_menu_text(self, note: str) -> str:
        # Telegram sizes inline keyboards from the attached message bubble width.
        # A blank width hint keeps repeated sverka menus from becoming narrow.
        return f"{note}\n{chr(0x2800) * 36}"

    async def show_sverka_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE, note: Optional[str] = None):
        lang = 'uz'
        self._init_sverka_status(context)
        status = context.user_data.get('sverka_status', {})

        keyboard = []
        for key, label_uz, label_ru, *_rest in self._active_sverka_config(context):
            label = label_uz if lang == 'uz' else label_ru
            icon = "✅" if status.get(key) else "☐"
            keyboard.append([InlineKeyboardButton(f"{icon} {label}", callback_data=f"sv:{key}")])

        finish_text = "🟢 Yakunlash" if lang == 'uz' else "🟢 Yakunlash"
        cancel_text = "❌ Bekor qilish" if lang == 'uz' else "❌ Bekor qilish"
        keyboard.append(
            [
                InlineKeyboardButton(finish_text, callback_data="sv:finish"),
                InlineKeyboardButton(cancel_text, callback_data="sv:cancel"),
            ]
        )

        if not note:
            note = "Sverka bo'limlarini tanlang:" if lang == 'uz' else "Sverka bo'limlarini tanlang:"

        # Add missing items list at the bottom
        missing = []
        for key, label_uz, label_ru, *_rest in self._active_sverka_config(context):
            if not status.get(key):
                missing.append(label_uz if lang == 'uz' else label_ru)
        if missing:
            note += "\n\nTo'ldirilmagan bandlar:\n- " + "\n- ".join(missing)

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=self._sverka_menu_text(note),
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    async def sverka_select_step(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()

        key = (query.data or '').split(':', 1)[1] if query.data else ''
        if key == 'cancel':
            entrypoint = context.user_data.get("sverka_entrypoint")
            context.user_data['flow'] = None
            self._clear_sverka_flow_state(context)
            context.user_data.pop("awaiting_payment_images_for_close", None)
            self._clear_debt_received_detail_state(context)
            self._clear_debt_payments_detail_state(context)
            self._clear_expense_detail_state(context)
            self._clear_generic_payment_method_state(context)
            msg = "Kassa yopish jarayoni bekor qilindi." if entrypoint == "closing" else "Sverka jarayoni bekor qilindi."
            await context.bot.send_message(chat_id=query.message.chat_id, text=msg)
            await self.show_cashier_menu(update, context)
            return MAIN_MENU

        if key == 'finish':
            if not self._sverka_all_done(context):
                msg = "Hamma band to'ldirilmagan. Iltimos, qolganlarini to'ldiring." if 'uz' == 'uz' else "Р В РЎСљР В Р’Вµ Р В Р вЂ Р РЋР С“Р В Р’Вµ Р В РЎвЂ”Р РЋРЎвЂњР В Р вЂ¦Р В РЎвЂќР РЋРІР‚С™Р РЋРІР‚в„– Р В Р’В·Р В Р’В°Р В РЎвЂ”Р В РЎвЂўР В Р’В»Р В Р вЂ¦Р В Р’ВµР В Р вЂ¦Р РЋРІР‚в„–. Р В РІР‚вЂќР В Р’В°Р В РЎвЂ”Р В РЎвЂўР В Р’В»Р В Р вЂ¦Р В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂўР РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р вЂ Р РЋРІвЂљВ¬Р В РЎвЂР В Р’ВµР РЋР С“Р РЋР РЏ."
                await context.bot.send_message(chat_id=query.message.chat_id, text=msg)
                await self.show_sverka_menu(update, context)
                return SUBMIT_DAILY_REPORT
            return await self._finalize_sverka(update, context)

        config = {c[0]: c for c in self._active_sverka_config(context)}
        if key not in config:
            await self.show_sverka_menu(update, context)
            return SUBMIT_DAILY_REPORT

        _, label_uz, label_ru, state, prompt_uz, prompt_ru = config[key]
        context.user_data['pending_sverka_key'] = key
        context.user_data['pending_sverka_state'] = state
        prompt = prompt_uz if 'uz' == 'uz' else prompt_ru
        if key == "debt_received":
            self._clear_debt_received_detail_state(context)
        if key == "tax_info":
            self._clear_tax_info_state(context)
            context.user_data["tax_info_stage"] = "check_image"
        if key == "expenses":
            self._clear_expense_detail_state(context)
            context.user_data["expense_detail_stage"] = "items"
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text=self._build_expense_entry_prompt(context),
                reply_markup=self._build_expense_entry_keyboard(),
            )
            return state
        await context.bot.send_message(chat_id=query.message.chat_id, text=prompt, reply_markup=ReplyKeyboardRemove())
        return state

    async def _after_sverka_step(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if self._sverka_all_done(context):
            note = "✅ Barcha bandlar to'ldirildi. Yakunlash tugmasini bosing." if 'uz' == 'uz' else "Barcha bandlar to'ldirildi. Yakunlash tugmasini bosing."
            await self.show_sverka_menu(update, context, note=note)
            return SUBMIT_DAILY_REPORT

        note = "Qabul qilindi. Keyingi bandni tanlang." if 'uz' == 'uz' else "Р В РЎСџР РЋР вЂљР В РЎвЂР В Р вЂ¦Р РЋР РЏР РЋРІР‚С™Р В РЎвЂў. Р В РІР‚в„ўР РЋРІР‚в„–Р В Р’В±Р В Р’ВµР РЋР вЂљР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р РЋР С“Р В Р’В»Р В Р’ВµР В РўвЂР РЋРЎвЂњР РЋР вЂ№Р РЋРІР‚В°Р В РЎвЂР В РІвЂћвЂ“ Р В РЎвЂ”Р РЋРЎвЂњР В Р вЂ¦Р В РЎвЂќР РЋРІР‚С™."
        await self.show_sverka_menu(update, context, note=note)
        return SUBMIT_DAILY_REPORT

    async def close_shift(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Save closing amount and ask cashier for note."""
        try:
            try:
                amount = self._parse_amount(update.message.text)
            except ValueError:
                await update.message.reply_text(self._invalid_amount_msg(context))
                return CLOSE_SHIFT

            context.user_data["pending_close_amount"] = amount
            await update.message.reply_text("Izoh kiriting (xohlagan matn):")
            return CLOSE_SHIFT_NOTE
        except Exception:
            logger.exception("close_shift amount step failed")
            context.user_data.pop("pending_close_amount", None)
            context.user_data['flow'] = None
            await update.message.reply_text("Smena yopishda xatolik bo'ldi. Qayta urinib ko'ring.")
            await self.show_cashier_menu(update, context)
            return MAIN_MENU

    async def close_shift_note(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Close active shift after receiving cashier note."""
        try:
            amount = context.user_data.get("pending_close_amount")
            if amount is None:
                await update.message.reply_text("Avval yopish summasini kiriting.")
                return CLOSE_SHIFT

            user_row = await self.db.fetch_one("SELECT id FROM users WHERE telegram_id = %s", (update.effective_user.id,))
            if not user_row:
                await update.message.reply_text("Foydalanuvchi topilmadi.")
                return MAIN_MENU

            open_shifts = await self.db.fetch_all(
                "SELECT id FROM shifts WHERE user_id=%s AND is_open=TRUE ORDER BY id DESC",
                (user_row['id'],)
            ) or []
            if not open_shifts:
                await update.message.reply_text("Ochiq smena topilmadi.")
                return MAIN_MENU

            # Asosiy yopiladigan smena: eng oxirgisi
            shift_id = int(open_shifts[0]['id'])
            context.user_data['current_shift_id'] = shift_id

            await self.db.execute_query(
                "UPDATE shifts SET closing_amount=%s, closed_at=NOW(), is_open=FALSE WHERE id=%s",
                (amount, shift_id)
            )

            # Xavfsizlik: agar tasodifan bir nechta ochiq smena qolgan bo'lsa, ularni ham yopamiz
            stale_ids = [int(r['id']) for r in open_shifts[1:]]
            if stale_ids:
                for sid in stale_ids:
                    await self.db.execute_query(
                        "UPDATE shifts SET closed_at=COALESCE(closed_at, NOW()), is_open=FALSE WHERE id=%s",
                        (sid,)
                    )

            shift_summary = await self._get_shift_summary(shift_id) or {}
            note = (update.message.text or "").strip()
            sent = await self._send_group_message(
                context,
                self._build_shift_closed_message(shift_summary, note),
            )
            if not sent:
                await update.message.reply_text(
                    "Diqqat: smena yopilgani guruhga yuborilmadi. /setgroup va bot ruxsatlarini tekshiring."
                )

            await update.message.reply_text("Smena yopildi.")
            await self.show_cashier_menu(update, context)
            context.user_data['flow'] = None
            context.user_data.pop('pending_close_amount', None)
            context.user_data.pop('current_shift_id', None)
            context.user_data.pop('opening_stage', None)
            context.user_data.pop('pending_payment_image', None)
            context.user_data.pop("awaiting_payment_images_for_close", None)
            context.user_data.pop("sverka_shift_id", None)
            self._clear_sverka_flow_state(context)
            return MAIN_MENU
        except Exception:
            logger.exception("close_shift note step failed")
            context.user_data['flow'] = None
            context.user_data.pop('pending_close_amount', None)
            await update.message.reply_text("Smena yopishda xatolik bo'ldi. Qayta urinib ko'ring.")
            await self.show_cashier_menu(update, context)
            return MAIN_MENU

    def _build_shift_full_xlsx_workbook(self, shift: dict, report: dict, images: list, image_blobs: Optional[dict] = None) -> BytesIO:
        """
        Build one Excel file with all cashier data for the shift:
        - Smena (opened/closed, opening/closing amount, cashier, location)
        - Sverka (all numeric fields, one row)
        - Rasmlar (required photos + payment photos with embedded images)
        """
        wb = Workbook()
        ws_shift = wb.active
        ws_shift.title = "Smena"
        image_blobs = image_blobs or {}

        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_font = Font(bold=True, color="FFFFFF")
        bold = Font(bold=True)

        def _set_kv(row_idx: int, key: str, value):
            ws_shift.cell(row=row_idx, column=1, value=key).font = bold
            ws_shift.cell(row=row_idx, column=2, value=value)

        cashier_name = (f"{shift.get('first_name','')} {shift.get('last_name') or ''}".strip() or "Kassir")
        _set_kv(1, "Kassir", cashier_name)
        _set_kv(2, "Telefon", shift.get("phone_number") or "")
        _set_kv(3, "Filial", shift.get("location") or "")
        _set_kv(4, "Smena ochilgan vaqt", str(shift.get("opened_at") or ""))
        _set_kv(5, "Smena yopilgan vaqt", str(shift.get("closed_at") or ""))
        _set_kv(6, "Ochilish summasi", float(shift.get("opening_amount") or 0))
        _set_kv(7, "Yopish summasi", float(shift.get("closing_amount") or 0))

        ws_shift.column_dimensions["A"].width = 22
        ws_shift.column_dimensions["B"].width = 42
        for r in range(1, 8):
            ws_shift.cell(row=r, column=1).alignment = Alignment(vertical="center")
            ws_shift.cell(row=r, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        ws_shift.cell(row=6, column=2).number_format = "#,##0"
        ws_shift.cell(row=7, column=2).number_format = "#,##0"

        # Sverka sheet
        ws_rep = wb.create_sheet("Sverka")
        rep_headers = [
            "Savdo", "Kelgan qarz", "Chiqim", "Uzcard", "Humo",
            "P2P",
            "Uzcard vozvrat", "Humo vozvrat", "Boshqa to'lovlar",
            "Qarzga berilgan to'lovlar", "Vozvrat qarzlar", "Naqd kutiladigan summa",
        ]
        ws_rep.append(rep_headers)

        def _f(key: str) -> float:
            try:
                return float(report.get(key) or 0)
            except Exception:
                return 0.0

        total_balance = self._calculate_total_balance(report)

        ws_rep.append([
            _f("sales_amount"),
            _f("debt_received"),
            _f("expenses"),
            _f("uzcard_amount"),
            _f("humo_amount"),
            _f("p2p_amount"),
            _f("uzcard_refund"),
            _f("humo_refund"),
            _f("other_payments"),
            _f("debt_payments"),
            _f("debt_refunds"),
            float(total_balance),
        ])

        for c in range(1, len(rep_headers) + 1):
            cell = ws_rep.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws_rep.column_dimensions[get_column_letter(c)].width = min(max(12, len(rep_headers[c - 1]) + 2), 28)
            ws_rep.cell(row=2, column=c).number_format = "#,##0"

        # Images sheet
        ws_img = wb.create_sheet("Rasmlar")
        img_headers = ["Rasm turi", "Sana/Vaqt", "Rasm"]
        ws_img.append(img_headers)
        for c in range(1, len(img_headers) + 1):
            cell = ws_img.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_img.column_dimensions["A"].width = 32
        ws_img.column_dimensions["B"].width = 24
        ws_img.column_dimensions["C"].width = 42

        type_labels = {
            "workplace_status": "Ish joyi holati (2 ta)",
            "terminal_power": "Terminal/ratsiya quvvati",
            "zero_report": "Nol hisobot (Uzcard/Humo)",
            "opening_notification": "Iiko/soliq smena ochildi",
            "receipt_roll": "Zaxira chek lenta",
            "uzcard_payment": "Uzcard to'lov rasmi",
            "humo_payment": "Humo to'lov rasmi",
            "tax_info_check": "Soliq ma'lumotlari cheki",
        }
        ordered_types = [
            "workplace_status",
            "terminal_power",
            "zero_report",
            "opening_notification",
            "receipt_roll",
            "uzcard_payment",
            "humo_payment",
            "tax_info_check",
        ]

        ordered_images = sorted(
            images or [],
            key=lambda row: (
                ordered_types.index(row.get("image_type")) if row.get("image_type") in ordered_types else len(ordered_types),
                str(row.get("uploaded_at") or ""),
            ),
        )

        if not ordered_images:
            ws_img.append(["Rasmlar topilmadi.", "", ""])
        else:
            r = 2
            for item in ordered_images:
                ws_img.row_dimensions[r].height = 140
                image_type = item.get("image_type") or ""
                file_ref = item.get("image_url")
                ws_img.cell(row=r, column=1, value=type_labels.get(image_type, image_type))
                ws_img.cell(row=r, column=2, value=str(item.get("uploaded_at") or "")[:19])
                ok = self._embed_excel_image(ws_img, f"C{r}", image_blobs.get(file_ref))
                if not ok:
                    ws_img.cell(row=r, column=3, value="Rasmni yuklab bo'lmadi.")
                r += 1

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    async def _fetch_shift_export_data(self, shift_id: int):
        shift = await self.db.fetch_one(
            """
            SELECT
              s.id, s.opened_at, s.closed_at, s.opening_amount, s.closing_amount, s.is_open,
              u.first_name, u.last_name, u.phone_number,
              l.name AS location
            FROM shifts s
            JOIN users u ON s.user_id=u.id
            JOIN locations l ON s.location_id=l.id
            WHERE s.id=%s
            """,
            (shift_id,)
        ) or {}

        report = await self.db.fetch_one(
            "SELECT * FROM reports WHERE shift_id=%s AND report_type='daily_report' ORDER BY id DESC LIMIT 1",
            (shift_id,)
        ) or {}

        images = await self.db.fetch_all(
            """
            SELECT image_type, image_url, uploaded_at
            FROM images
            WHERE shift_id=%s
            ORDER BY uploaded_at ASC
            """,
            (shift_id,)
        ) or []

        return shift, report, images

    def _build_opening_images_xlsx_workbook(self, rows: list, image_blobs: Optional[dict] = None) -> BytesIO:
        """Build one Excel with only shift-opening images."""
        image_blobs = image_blobs or {}
        wb = Workbook()
        ws = wb.active
        ws.title = "Smena ochish rasmlari"

        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_font = Font(bold=True, color="FFFFFF")

        for col, label in enumerate(["Rasm nomi", "Rasm", "Sana/Vaqt"], start=1):
            c = ws.cell(row=1, column=col, value=label)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 24

        readable = {
            "workplace_status": "Ish joyi holati",
            "terminal_power": "Terminal/ratsiya holati",
            "zero_report": "Nol hisobot (Uzcard/Humo)",
            "opening_notification": "iiko/soliq ochilish bildirishnomasi",
            "receipt_roll": "Zaxira chek lenta",
        }

        if not rows:
            ws.cell(row=2, column=1, value="Smena ochish rasmlari topilmadi.")
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        r_idx = 2
        for item in rows:
            ws.row_dimensions[r_idx].height = 140
            t = (item.get("image_type") or "").strip()
            ws.cell(row=r_idx, column=1, value=readable.get(t, t))
            ok = self._embed_excel_image(ws, f"B{r_idx}", image_blobs.get(item["image_url"]))
            if not ok:
                ws.cell(row=r_idx, column=2, value="Rasmni yuklab bo'lmadi.")
            ws.cell(row=r_idx, column=3, value=str(item.get("uploaded_at") or "")[:19])
            r_idx += 1

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    def _build_shift_images_xlsx_workbook(self, rows: list, image_blobs: Optional[dict] = None) -> BytesIO:
        """Build one Excel with payment images and operational images."""
        image_blobs = image_blobs or {}
        wb = Workbook()
        ws_pay = wb.active
        ws_pay.title = "To'lov rasmlari"

        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_font = Font(bold=True, color="FFFFFF")

        for col, label in enumerate(["Uzcard rasm", "Humo rasm", "Sana/Vaqt"], start=1):
            c = ws_pay.cell(row=1, column=col, value=label)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws_pay.column_dimensions["A"].width = 42
        ws_pay.column_dimensions["B"].width = 42
        ws_pay.column_dimensions["C"].width = 24

        pay = defaultdict(list)
        other = []
        for r in rows:
            t = (r.get("image_type") or "").strip()
            if t in ("uzcard_payment", "humo_payment"):
                pay[t].append(r)
            else:
                other.append(r)

        uz_rows = pay.get("uzcard_payment", [])
        hu_rows = pay.get("humo_payment", [])
        max_len = max(len(uz_rows), len(hu_rows), 1)

        for i in range(max_len):
            excel_row = i + 2
            ws_pay.row_dimensions[excel_row].height = 140

            uz = uz_rows[i] if i < len(uz_rows) else None
            hu = hu_rows[i] if i < len(hu_rows) else None

            if uz:
                ok = self._embed_excel_image(ws_pay, f"A{excel_row}", image_blobs.get(uz["image_url"]))
                if not ok:
                    ws_pay.cell(row=excel_row, column=1, value="Rasmni yuklab bo'lmadi.")
            if hu:
                ok = self._embed_excel_image(ws_pay, f"B{excel_row}", image_blobs.get(hu["image_url"]))
                if not ok:
                    ws_pay.cell(row=excel_row, column=2, value="Rasmni yuklab bo'lmadi.")

            stamp = (uz or hu or {}).get("uploaded_at")
            ws_pay.cell(row=excel_row, column=3, value=str(stamp)[:19] if stamp else "")

        ws_other = wb.create_sheet("Ish jarayoni rasmlari")
        for col, label in enumerate(["Rasm turi", "Rasm", "Sana/Vaqt"], start=1):
            c = ws_other.cell(row=1, column=col, value=label)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws_other.column_dimensions["A"].width = 32
        ws_other.column_dimensions["B"].width = 42
        ws_other.column_dimensions["C"].width = 24

        readable = {
            "workplace_status": "Ish joyi holati",
            "terminal_power": "Terminal/ratsiya holati",
            "zero_report": "Nol hisobot",
            "opening_notification": "iiko/soliq ochilish",
            "receipt_roll": "Zaxira chek lenta",
            "tax_info_check": "Soliq ma'lumotlari cheki",
        }

        r_idx = 2
        for item in other:
            ws_other.row_dimensions[r_idx].height = 140
            t = item.get("image_type") or ""
            ws_other.cell(row=r_idx, column=1, value=readable.get(t, t))
            ok = self._embed_excel_image(ws_other, f"B{r_idx}", image_blobs.get(item["image_url"]))
            if not ok:
                ws_other.cell(row=r_idx, column=2, value="Rasmni yuklab bo'lmadi.")
            ws_other.cell(row=r_idx, column=3, value=str(item.get("uploaded_at") or "")[:19])
            r_idx += 1

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out
    async def reset_cashier_password(self, update: Update, context: ContextTypes.DEFAULT_TYPE, telegram_id: int):
        """Reset cashier password and ask them to set a new one"""
        user = await self.db.fetch_one(
            AdminQueries.CASHIER_BY_TELEGRAM_ID,
            (telegram_id,)
        )
        if not user:
            await context.bot.send_message(chat_id=update.effective_chat.id, text="Kassir topilmadi.")
            return

        await self.db.execute_query(AdminQueries.RESET_CASHIER_PASSWORD, (telegram_id,))

        try:
            await context.bot.send_message(
                chat_id=telegram_id,
                text="Parolingiz reset qilindi. /start bosing va yangi parol kiriting."
            )
        except Exception:
            pass

        await context.bot.send_message(chat_id=update.effective_chat.id, text="Parol reset qilindi.")

    async def _send_group_message(self, context: ContextTypes.DEFAULT_TYPE, text: str):
        group_chat_id = await self._get_group_chat_id()
        if not group_chat_id:
            logger.warning("Group message skipped: group_chat_id is not configured")
            return False
        try:
            await context.bot.send_message(chat_id=group_chat_id, text=text)
            return True
        except Exception:
            logger.exception("Failed to send group message to %s", group_chat_id)
            return False

    async def _send_group_photo(self, context: ContextTypes.DEFAULT_TYPE, file_id: str, caption: str = ""):
        group_chat_id = await self._get_group_chat_id()
        if not group_chat_id:
            logger.warning("Group photo skipped: group_chat_id is not configured")
            return False
        try:
            await context.bot.send_photo(chat_id=group_chat_id, photo=file_id, caption=caption)
            return True
        except Exception:
            logger.exception("Failed to send group photo to %s", group_chat_id)
            return False

    async def _send_group_document(self, context: ContextTypes.DEFAULT_TYPE, data: BytesIO, filename: str, caption: str = ""):
        group_chat_id = await self._get_group_chat_id()
        if not group_chat_id:
            logger.warning("Group document skipped: group_chat_id is not configured")
            return False
        try:
            data.seek(0)
            doc = InputFile(data, filename=filename)
            await context.bot.send_document(chat_id=group_chat_id, document=doc, caption=caption)
            return True
        except Exception:
            logger.exception("Failed to send group document to %s", group_chat_id)
            return False

    async def _get_shift_meta(self, shift_id: int):
        row = await self.db.fetch_one(
            """
            SELECT
                s.opened_at,
                l.name AS location,
                u.first_name,
                u.last_name
            FROM shifts s
            JOIN users u ON s.user_id = u.id
            JOIN locations l ON s.location_id = l.id
            WHERE s.id=%s
            """,
            (shift_id,),
        )
        if not row:
            return {"cashier": "", "location": "", "opened_at": ""}
        cashier = f"{row.get('first_name','')} {row.get('last_name') or ''}".strip()
        return {
            "cashier": cashier,
            "location": row.get("location") or "",
            "opened_at": str(row.get("opened_at") or "")[:19],
        }

    async def _send_group_shift_photo(
        self,
        context: ContextTypes.DEFAULT_TYPE,
        shift_id: int,
        file_id: str,
        image_title: str,
        event_time=None
    ):
        group_chat_id = await self._get_group_chat_id()
        if not group_chat_id:
            logger.warning("Group shift photo skipped: group_chat_id is not configured")
            return False

        shift_meta = await self._get_shift_meta(shift_id)
        caption = (
            f"📷 {image_title}\n"
            f"👤 Kassir: {shift_meta.get('cashier') or '-'}\n"
            f"🏬 Filial: {shift_meta.get('location') or '-'}\n"
            f"⏰ Vaqt: {self._format_telegram_time(event_time)}"
        )
        try:
            await context.bot.send_photo(chat_id=group_chat_id, photo=file_id, caption=caption)
            return True
        except Exception:
            logger.exception("Failed to send group shift photo as photo, fallback to document")

        try:
            await context.bot.send_document(chat_id=group_chat_id, document=file_id, caption=caption)
            return True
        except Exception:
            logger.exception("Failed to send group shift photo to %s", group_chat_id)
            return False

    def _build_sverka_xlsx(self, cashier_name: str, phone: str, location: str, opened_at, report_data: dict) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sverka"

        headers = [
            "Kassir", "Telefon", "Filial", "Smena ochilgan vaqt",
            "Savdo", "Kelgan qarz", "Chiqim", "Uzcard", "Humo",
            "P2P",
            "Uzcard vozvrat", "Humo vozvrat", "Boshqa to'lovlar",
            "Qarzga berilgan to'lovlar", "Vozvrat qarzlar", "Naqd kutiladigan summa",
        ]

        total_balance = self._calculate_total_balance(report_data)

        row = [
            cashier_name,
            phone,
            location,
            str(opened_at),
            float(report_data.get("sales_amount", 0) or 0),
            float(report_data.get("debt_received", 0) or 0),
            float(report_data.get("expenses", 0) or 0),
            float(report_data.get("uzcard_amount", 0) or 0),
            float(report_data.get("humo_amount", 0) or 0),
            float(report_data.get("p2p_amount", 0) or 0),
            float(report_data.get("uzcard_refund", 0) or 0),
            float(report_data.get("humo_refund", 0) or 0),
            float(report_data.get("other_payments", 0) or 0),
            float(report_data.get("debt_payments", 0) or 0),
            float(report_data.get("debt_refunds", 0) or 0),
            float(total_balance),
        ]

        ws.append(headers)
        ws.append(row)

        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_font = Font(bold=True, color="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        money_cols = list(range(5, 17))
        for col in money_cols:
            ws.cell(row=2, column=col).number_format = "#,##0"

        # Auto width
        for col in range(1, len(headers) + 1):
            max_len = 0
            for r in (1, 2):
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 40)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out
    async def modify_user_data(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle user data modification"""
        lang = 'uz'
        
        if lang == 'uz':
            msg = "Foydalanuvchi ma'lumotlarini o'zgartirish funksiyasi ishga tushirildi."
        else:
            msg = "Р В Р’В Р вЂ™Р’В¤Р В Р Р‹Р РЋРІР‚СљР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СњР В Р Р‹Р Р†Р вЂљР’В Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В Р РЏ Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В·Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚ВР В Р Р‹Р В Р РЏ Р В Р’В Р СћРІР‚ВР В Р’В Р вЂ™Р’В°Р В Р’В Р В РІР‚В¦Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р Р‹Р Р†Р вЂљР’В¦ Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р вЂ°Р В Р’В Р вЂ™Р’В·Р В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В Р В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’ВµР В Р’В Р вЂ™Р’В»Р В Р Р‹Р В Р РЏ Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р РЋРІР‚СљР В Р Р‹Р Р†Р вЂљР’В°Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’В°."
            
        await update.message.reply_text(msg)

    async def export_data(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle data export to Excel/PDF"""
        lang = 'uz'
        
        # Show export options
        if lang == 'uz':
            keyboard = [[KeyboardButton(label) for label in row] for row in EXPORT_MENU_ROWS]
            msg = "Eksport qilish formatini tanlang:"
        else:
            keyboard = [
                [KeyboardButton("Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р В РІР‚В° Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (Excel)"), KeyboardButton("Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р В РІР‚В° Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (PDF)")],
                [KeyboardButton("Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (Excel)"), KeyboardButton("Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (PDF)")],
                [KeyboardButton("Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎСљР Р†РІР‚С›РЎС› Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚В")]
            ]
            msg = "Р В Р’В Р Р†Р вЂљРІвЂћСћР В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р вЂ™Р’В±Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р Р‹Р Р†Р вЂљРЎвЂєР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’В°Р В Р Р‹Р Р†Р вЂљРЎв„ў Р В Р Р‹Р В Р Р‰Р В Р’В Р РЋРІР‚СњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°:"
            
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(msg, reply_markup=reply_markup)

    async def handle_export_choice(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle export choice from user"""
        text = update.message.text
        lang = 'uz'
        
        try:
            if text in ["Kunlik hisobot (Excel)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р В РІР‚В° Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (Excel)"]:
                # Generate Excel report
                await update.message.reply_text("Hisobot tayyorlanmoqda...")
                excel_data = await self.export_utils.generate_excel_report(report_type='daily')
                
                if lang == 'uz':
                    caption = self._build_export_caption("Kunlik hisobot", "Excel")
                else:
                    caption = "Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (Excel)"
                    
                excel_file = InputFile(excel_data, filename="kunlik_hisobot.xlsx")
                await update.message.reply_document(document=excel_file, caption=caption)
                
            elif text in ["Kunlik hisobot (PDF)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎС™Р В РІР‚В° Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (PDF)"]:
                # Generate PDF report
                await update.message.reply_text("Hisobot tayyorlanmoqda...")
                pdf_data = await self.export_utils.generate_pdf_report(report_type='daily')
                
                if lang == 'uz':
                    caption = self._build_export_caption("Kunlik hisobot", "PDF")
                else:
                    caption = "Р В Р’В Р Р†Р вЂљРЎС›Р В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’ВµР В Р’В Р СћРІР‚ВР В Р’В Р В РІР‚В¦Р В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В Р В Р’В Р В РІР‚В¦Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р Р†РІР‚С›РІР‚вЂњ Р В Р’В Р РЋРІР‚СћР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў (PDF)"
                    
                pdf_file = InputFile(pdf_data, filename="kunlik_hisobot.pdf")
                await update.message.reply_document(document=pdf_file, caption=caption)
                
            elif text in ["Kassirlar bo'yicha (Excel)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (Excel)"]:
                # Generate Excel report for cashiers
                await update.message.reply_text("Hisobot tayyorlanmoqda...")
                excel_data = await self.export_utils.generate_excel_report(report_type='cashier_performance')
                
                if lang == 'uz':
                    caption = self._build_export_caption("Kassirlar bo'yicha hisobot", "Excel")
                else:
                    caption = "Р В Р’В Р РЋРІР‚С”Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (Excel)"
                    
                excel_file = InputFile(excel_data, filename="kassirlar_hisobot.xlsx")
                await update.message.reply_document(document=excel_file, caption=caption)
                
            elif text in ["Kassirlar bo'yicha (PDF)", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљР’ВР СћРЎвЂ™ Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (PDF)"]:
                # Generate PDF report for cashiers
                await update.message.reply_text("Hisobot tayyorlanmoqda...")
                pdf_data = await self.export_utils.generate_pdf_report(report_type='cashier_performance')
                
                if lang == 'uz':
                    caption = self._build_export_caption("Kassirlar bo'yicha hisobot", "PDF")
                else:
                    caption = "Р В Р’В Р РЋРІР‚С”Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р Р‹Р Р†Р вЂљР Р‹Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р Р†Р вЂљРЎв„ў Р В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚Сћ Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В°Р В Р Р‹Р В РЎвЂњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚ВР В Р Р‹Р В РІР‚С™Р В Р’В Р вЂ™Р’В°Р В Р’В Р РЋР’В (PDF)"
                    
                pdf_file = InputFile(pdf_data, filename="kassirlar_hisobot.pdf")
                await update.message.reply_document(document=pdf_file, caption=caption)
                
            elif text in ["Orqaga", "Р РЋР вЂљР РЋРЎСџР Р†Р вЂљРЎСљР Р†РІР‚С›РЎС› Р В Р’В Р РЋРЎС™Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В·Р В Р’В Р вЂ™Р’В°Р В Р’В Р СћРІР‚В"]:
                # Return to admin menu
                await self.show_admin_menu(update, context)
                
            else:
                if lang == 'uz':
                    msg = "Iltimos, menyudan birini tanlang."
                else:
                    msg = "Р В Р’В Р РЋРЎСџР В Р’В Р РЋРІР‚СћР В Р’В Р вЂ™Р’В¶Р В Р’В Р вЂ™Р’В°Р В Р’В Р вЂ™Р’В»Р В Р Р‹Р РЋРІР‚СљР В Р’В Р Р†РІР‚С›РІР‚вЂњР В Р Р‹Р В РЎвЂњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’В°, Р В Р’В Р В РІР‚В Р В Р Р‹Р Р†Р вЂљРІвЂћвЂ“Р В Р’В Р вЂ™Р’В±Р В Р’В Р вЂ™Р’ВµР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚ВР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ Р В Р’В Р РЋРІР‚СћР В Р’В Р СћРІР‚ВР В Р’В Р РЋРІР‚ВР В Р’В Р В РІР‚В¦ Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В· Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р РЋРІР‚СљР В Р’В Р В РІР‚В¦Р В Р’В Р РЋРІР‚СњР В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р РЋРІР‚СћР В Р’В Р В РІР‚В  Р В Р’В Р РЋР’ВР В Р’В Р вЂ™Р’ВµР В Р’В Р В РІР‚В¦Р В Р Р‹Р В РІР‚в„–."
                    
                await update.message.reply_text(msg)
                
        except Exception as e:
            if lang == 'uz':
                msg = f"Eksport qilishda xatolik yuz berdi: {str(e)}"
            else:
                msg = f"Р В Р’В Р РЋРІР‚С”Р В Р Р‹Р Р†РІР‚С™Р’В¬Р В Р’В Р РЋРІР‚ВР В Р’В Р вЂ™Р’В±Р В Р’В Р РЋРІР‚СњР В Р’В Р вЂ™Р’В° Р В Р’В Р РЋРІР‚вЂќР В Р Р‹Р В РІР‚С™Р В Р’В Р РЋРІР‚В Р В Р Р‹Р В Р Р‰Р В Р’В Р РЋРІР‚СњР В Р Р‹Р В РЎвЂњР В Р’В Р РЋРІР‚вЂќР В Р’В Р РЋРІР‚СћР В Р Р‹Р В РІР‚С™Р В Р Р‹Р Р†Р вЂљРЎв„ўР В Р’В Р вЂ™Р’Вµ: {str(e)}"
                
            await update.message.reply_text(msg)


async def _global_error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Catch unexpected errors so the bot doesn't go silent."""
    from telegram.error import Conflict, NetworkError
    try:
        err = getattr(context, "error", None)
        # Conflict means another instance is running — log but don't spam users
        if isinstance(err, Conflict):
            logger.warning("Bot conflict: another instance is running. %s", err)
            return
        if isinstance(err, NetworkError):
            logger.warning("Network error (transient): %s", err)
            return
        if err:
            logger.exception("Unhandled exception", exc_info=err)
        else:
            logger.exception("Unhandled exception (no context.error)")
    except Exception:
        pass
    try:
        if isinstance(update, Update) and update.effective_chat:
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="Xatolik yuz berdi. Iltimos, qayta urinib ko'ring."
            )
    except Exception:
        pass

def main():
    """Run the bot."""
    import fcntl
    # --- PID lock: faqat bitta instance ishlashini ta'minlaydi ---
    _PID_FILE = "/tmp/sardoba_kassa_bot.lock"
    try:
        _pid_fh = open(_PID_FILE, 'w')
        fcntl.flock(_pid_fh, fcntl.LOCK_EX | fcntl.LOCK_NB)
        _pid_fh.write(str(os.getpid()))
        _pid_fh.flush()
    except BlockingIOError:
        import sys
        print("ERROR: Bot allaqachon ishlayapti! Avval eski jarayonni to'xtating.", flush=True)
        sys.exit(1)

    # Create the Application and pass it your bot's token
    TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', 'YOUR_BOT_TOKEN_HERE')
    request = HTTPXRequest(connect_timeout=20.0, read_timeout=60.0, write_timeout=60.0, pool_timeout=20.0)
    bot = SardobaBot()
    
    async def _post_init(_application):
        await bot.initialize()

    async def _post_shutdown(_application):
        await bot.shutdown()

    application = (
        Application.builder()
        .token(TOKEN)
        .request(request)
        .post_init(_post_init)
        .post_shutdown(_post_shutdown)
        .build()
    )

    # Create conversation handler for registration flow
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", bot.start), CallbackQueryHandler(bot.select_role, pattern='^role_')],
        states={
            SELECT_ROLE: [CallbackQueryHandler(bot.select_role)],
            REGISTER_FIRSTNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.register_firstname)],
            REGISTER_LASTNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.register_lastname)],
            REGISTER_PHONE: [MessageHandler(filters.CONTACT | (filters.TEXT & ~filters.COMMAND), bot.register_phone)],
            REGISTER_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.register_password)],
            VERIFY_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.verify_password)],
            ADMIN_LOGIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.admin_login)],
            ADMIN_REGISTER_PHONE: [MessageHandler(filters.CONTACT | (filters.TEXT & ~filters.COMMAND), bot.admin_register_phone)],
            ADMIN_REGISTER_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.admin_register_password)],
            ADMIN_VERIFY_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.admin_verify_password)],
        },
        fallbacks=[CommandHandler("cancel", bot.cancel)],
    )

    # Add handlers
    application.add_handler(conv_handler)
    application.add_handler(CommandHandler("setgroup", bot.set_group))
    application.add_handler(CommandHandler("cancel", bot.cancel))

    cashier_conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.TEXT & filters.Regex(r"^\s*Smena ochish\s*$"), bot.start_shift_opening),
            MessageHandler(filters.TEXT & filters.Regex(r"^\s*Sverka\s*$"), bot.start_daily_reporting),
            MessageHandler(filters.TEXT & filters.Regex(r"^\s*Smena yopish\s*$"), bot.start_shift_closing),
            CallbackQueryHandler(bot.select_location, pattern='^loc_'),
        ],
        states={
            OPEN_SHIFT_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.open_shift_amount)],
            SELECT_LOCATION: [
                CallbackQueryHandler(bot.select_location, pattern='^loc_'),
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.ask_select_location_again),
            ],
            SELECT_PAYMENT_IMAGE: [CallbackQueryHandler(bot.select_payment_image_type, pattern='^payimg:')],
            UPLOAD_PAYMENT_IMAGE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_payment_image),
            ],
            UPLOAD_WORKPLACE_STATUS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_workplace_status),
            ],
            UPLOAD_TERMINAL_POWER: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_terminal_power),
            ],
            UPLOAD_ZERO_REPORT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_zero_report),
            ],
            UPLOAD_OPENING_NOTIFICATION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_opening_notification),
            ],
            UPLOAD_RECEIPT_ROLL: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message),
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.upload_receipt_roll),
            ],
            SUBMIT_DAILY_REPORT: [
                CallbackQueryHandler(bot.sverka_select_step, pattern='^sv:'),
                CallbackQueryHandler(bot.opening_select_step, pattern='^op:'),
            ],
            REPORT_SALES: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_sales)],
            REPORT_DEBT_RECEIVED: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_debt_received)],
            REPORT_EXPENSES: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_expenses)],
            REPORT_UZCARD: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_uzcard)],
            REPORT_HUMO: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_humo)],
            REPORT_P2P: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_p2p)],
            REPORT_UZCARD_REFUND: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_uzcard_refund)],
            REPORT_HUMO_REFUND: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_humo_refund)],
            REPORT_OTHER_PAYMENTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_other_payments)],
            REPORT_DEBT_PAYMENTS: [
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.report_debt_payments),
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_debt_payments),
            ],
            REPORT_DEBT_REFUNDS: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_debt_refunds)],
            REPORT_TAX_INFO: [
                MessageHandler(filters.PHOTO | filters.Document.ALL, bot.report_tax_info),
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.report_tax_info),
            ],
            EDIT_REPORT_SELECT: [CallbackQueryHandler(bot.edit_reports_select, pattern='^edit:')],
            EDIT_REPORT_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.edit_reports_value)],
            CLOSE_SHIFT: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.close_shift)],
            CLOSE_SHIFT_NOTE: [MessageHandler(filters.TEXT & ~filters.COMMAND, bot.close_shift_note)],
        },
        fallbacks=[CommandHandler("cancel", bot.cancel)],
    )
    application.add_handler(cashier_conv)
    # Ensure sverka inline buttons always work even if conversation state was lost
    application.add_handler(CallbackQueryHandler(bot.sverka_select_step, pattern='^sv:'))
    # Ensure opening checklist inline buttons always work even if conversation state was lost
    application.add_handler(CallbackQueryHandler(bot.opening_select_step, pattern='^op:'))
    # Admin reports inline menu
    application.add_handler(CallbackQueryHandler(bot.handle_reports_menu_callback, pattern='^rep:'))
    application.add_handler(CallbackQueryHandler(bot.handle_report_location_callback, pattern='^reploc:'))
    # Filial tanlash callbackini global ham ushlaymiz (state yo'qolsa ham ishlasin)
    application.add_handler(CallbackQueryHandler(bot.select_location, pattern='^loc_'))
    # Ensure edit inline buttons always work even if conversation state was lost
    application.add_handler(CallbackQueryHandler(bot.edit_reports_select, pattern='^edit:'))
    # Ensure payment image selection works even if conversation state was lost
    application.add_handler(CallbackQueryHandler(bot.select_payment_image_type, pattern='^payimg:'))

    # Handle photos/documents globally as fallback when conversation state is lost
    application.add_handler(MessageHandler(filters.PHOTO | filters.Document.ALL, bot.handle_image_message))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, bot.handle_message))

    # Add callback query handler for approval selections
    application.add_handler(CallbackQueryHandler(bot.handle_approval_callback, pattern='^(approve|reject):'))
    application.add_error_handler(_global_error_handler)

    # Run the bot until the user presses Ctrl-C
    application.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=True,   # Eski/conflict xabarlarni o'chiradi
    )


if __name__ == '__main__':
    main()
